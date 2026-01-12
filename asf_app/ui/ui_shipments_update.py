# -*- coding: utf-8 -*-
"""
Onglet : Mise à Jour expéditions
Permet d'annuler ou de reprogrammer un BE déjà planifié (statut P) et
de préparer les brouillons Outlook associés.
"""

import re
import fnmatch
import datetime as dt
from typing import Optional, Tuple, List
from pathlib import Path

import pandas as pd
import streamlit as st

from asf_app.ui.loader import load_parameters
from asf_app.ui.ui_planning.state_planning import get_planning_state
from asf_app.ui.ui_communication.email_destinations_handler import _get_emails_for_destination
from asf_app.ui.ui_communication.email_expediteurs_handler import _get_emails_for_expediteur
from asf_app.ui.ui_communication.outlook import create_outlook_draft
from loaders.load_benevoles import get_benevoles_cached
from loaders.load_vols import get_vols_df_cached
from loaders.universal_loader import load_and_normalize
from scheduler.column_map import column_map_mag_central
from scheduler.config_paths import SHEET_MAG_CENTRAL, TABLEAU_DE_BORD
import scheduler.config_paths as cp
from utils.datetime_utils import parse_date_series, parse_time_series, normalize_hour_str, hour_min_from_series
from utils.ui_helpers import build_iata_city_maps, format_be_label
from utils.export_pdf import export_first_sheet_to_pdf
from scheduler.planning_schema import normalize_planning_df
from utils.identifiers import normalize_be_number
from asf_app.ui.ui_shipments_update_helpers import _norm_be, _fmt_date_long, _fmt_time, _fmt_vol, _wrap_body


def _dest_to_iata(dest_raw: str, df_paramdest: pd.DataFrame) -> str:
    dest = str(dest_raw).strip().upper()
    if len(dest) == 3:
        return dest
    try:
        mapping = (
            df_paramdest.dropna(subset=["Dest_Ville", "Dest_IATA"])
            .assign(Dest_Ville_UP=lambda d: d["Dest_Ville"].astype(str).str.upper().str.strip())
            .drop_duplicates(subset=["Dest_Ville_UP"])
            .set_index("Dest_Ville_UP")["Dest_IATA"]
            .astype(str)
            .str.upper()
            .to_dict()
        )
        return mapping.get(dest, dest)
    except Exception:
        return dest


# ---------------------------------------------------------------------------
# Chargement BE statut P
# ---------------------------------------------------------------------------
def _load_be_status(status_code: str) -> pd.DataFrame:
    try:
        xls = pd.ExcelFile(TABLEAU_DE_BORD)
        sheets = [
            name
            for name in xls.sheet_names
            if str(name).strip().upper().startswith("MAG CENTRAL")
        ]
    except Exception:
        sheets = []

    if not sheets:
        sheets = [SHEET_MAG_CENTRAL]

    def _rank(name: str) -> tuple[int, str]:
        match = re.search(r"(20\\d{2})", name)
        year = int(match.group(1)) if match else -1
        return (year, name)

    sheets = sorted(sheets, key=_rank)
    frames = []
    for sheet in sheets:
        df_sheet = load_and_normalize(
            path=TABLEAU_DE_BORD,
            sheet_name=sheet,
            mapping=column_map_mag_central,
            header=5,
        )
        if df_sheet is None or df_sheet.empty:
            continue
        df_sheet["_MAG_CENTRAL_SHEET"] = sheet
        frames.append(df_sheet)

    df = pd.concat(frames, ignore_index=True) if frames else pd.DataFrame()
    if df is None or df.empty:
        return pd.DataFrame(columns=["Week", "Year"])

    df = df.copy()
    if "BE_Statut" not in df.columns:
        df["BE_Statut"] = ""
    df["BE_Statut"] = df["BE_Statut"].astype(str).str.upper().str.strip()
    df = df[df["BE_Statut"] == status_code.upper()].copy()

    df["BE_Numero_Str"] = df.get("BE_Numero", "").astype(str).str.replace(r"\.0$", "", regex=True).str.strip()
    df["Date_Vol"] = pd.to_datetime(df.get("BE_Date_Vol", pd.NaT), errors="coerce")
    iso = df["Date_Vol"].dt.isocalendar()
    df["Week"] = iso.week.astype("Int64")
    df["Year"] = iso.year.astype("Int64")
    return df


def _load_be_status_d_for_week(week: int, year: int) -> pd.DataFrame:
    df = _load_be_status("D")
    if df.empty:
        return df
    mask_match = (df["Week"] == week) & (df["Year"] == year)
    mask_na = df["Week"].isna() | df["Year"].isna()
    df = df[mask_match | mask_na].copy()
    df["Source"] = "mag_central"
    return df


def _load_planning_preview(week: int, year: int) -> tuple[pd.DataFrame | None, str, Path | None]:
    """
    Charge un aperçu du planning validé (fichier exporté) pour la semaine choisie.
    Retourne (df, message, path).
    """
    return _load_planning_preview_with_path(week, year, None)


def _load_planning_preview_with_path(
    week: int, year: int, path_override: Optional[Path | str]
) -> tuple[pd.DataFrame | None, str, Path | None]:
    """
    Variante : accepte un chemin explicite si déjà sélectionné.
    """
    msg_missing = None
    if cp.is_graph_onedrive():
        if path_override:
            remote_path = str(path_override)
        else:
            candidates = _find_planning_files_for_week(week, year)
            if not candidates:
                return None, f"Fichier introuvable : S{week:02d}-{year}", None
            remote_path = str(candidates[0])
        local_path = cp.TMP_DIR / "onedrive_cache" / "planning_exports" / remote_path
        if not local_path.exists():
            cp.download_onedrive_file(remote_path, local_path, interactive=False)
        if not local_path.exists():
            return None, f"Fichier introuvable : {remote_path}", None
        path = local_path
    else:
        base_dir = cp.ASF_ONEDRIVE / "Planning MAB" / f"ASFmm PLANNING {year}"
        if path_override:
            path = Path(path_override)
        else:
            filename = f"ASFmm - PLANNING SEMAINE N° {week:02d} - {year}.xlsx"
            path = base_dir / filename
            if not path.exists():
                # tolérance sur nom : espaces, tirets, xlsm/xlsx
                patterns = [
                    f"ASFmm*{week:02d}*{year}.xls*",
                    f"*PLANNING*{week:02d}*{year}.xls*",
                ]
                candidates = []
                for pat in patterns:
                    candidates.extend(base_dir.glob(pat))
                if candidates:
                    # garder ceux qui contiennent la semaine précise
                    def _score(p):
                        name_up = p.name.upper()
                        return int(f"{name_up.count(str(week).zfill(2))}{name_up.count(str(year))}")
                    path = sorted(candidates, key=_score, reverse=True)[0]
                    msg_missing = f"Fichier exact introuvable, utilisation de : {path.name}"
                else:
                    return None, f"Fichier introuvable : {path}", None
            else:
                msg_missing = None

    sheet_candidates = [f"Planning S{week:02d}", "Export planning"]
    for sh in sheet_candidates:
        try:
            df_prev = pd.read_excel(path, sheet_name=sh)
            if df_prev is not None and not df_prev.empty:
                if msg_missing:
                    return df_prev, f"{msg_missing} — Aperçu basé sur la feuille « {sh} »", path
                return df_prev, f"Aperçu basé sur la feuille « {sh} »", path
        except Exception:
            continue
    if msg_missing:
        return None, f"{msg_missing} — Impossible de lire les feuilles {sheet_candidates} dans {path.name}", path
    return None, f"Impossible de lire les feuilles {sheet_candidates} dans {path.name}", path


def _available_weeks_from_exports() -> set[tuple[int, int]]:
    """
    Retourne les semaines dispo en inspectant les fichiers d'export Excel
    dans OneDrive (ASFmm PLANNING YYYY).
    """
    weeks: set[tuple[int, int]] = set()
    if cp.is_graph_onedrive():
        items = cp.list_onedrive_files("Planning MAB", recursive=True, suffixes=[".xls", ".xlsx", ".xlsm"])
        for item in items:
            name = item.get("name", "")
            m_week = re.search(r"N°\s*(\d+)", name)
            m_year = re.search(r"(20\\d{2})", name)
            if not (m_week and m_year):
                continue
            try:
                wk = int(m_week.group(1))
                yr = int(m_year.group(1))
                weeks.add((wk, yr))
            except Exception:
                continue
        return weeks

    base_dir = cp.ASF_ONEDRIVE / "Planning MAB"
    if not base_dir.exists():
        return weeks

    for sub in base_dir.iterdir():
        if not sub.is_dir():
            continue
        name_up = sub.name.upper()
        if not name_up.startswith("ASFMM PLANNING "):
            continue
        try:
            year = int(re.sub(r"\\D", "", sub.name)[-4:])
        except Exception:
            continue
        for f in sub.glob("ASFmm - PLANNING SEMAINE N° *.xls*"):
            m = re.search(r"N°\\s*(\\d+)", f.name)
            if not m:
                continue
            try:
                wk = int(m.group(1))
                weeks.add((wk, year))
            except Exception:
                continue
    return weeks


def _parse_version_from_name(path: Path) -> tuple[int, int]:
    """
    Extrait vXX[-YY] du nom de fichier. Par défaut retourne (1,0).
    """
    stem = path.stem.upper()
    m = re.search(r"V(\d+)(?:-(\d+))?", stem)
    if m:
        try:
            major = int(m.group(1))
            minor = int(m.group(2) or 0)
            return major, minor
        except Exception:
            pass
    return 1, 0


def _find_planning_files_for_week(week: int, year: int) -> List[Path | str]:
    """
    Liste les fichiers de planning correspondant à la semaine/année,
    triés par version décroissante (vXX[-YY]) puis date de modif.
    """
    if cp.is_graph_onedrive():
        remote_dir = cp.get_output_remote_dir(year)
        pattern = f"ASFmm - PLANNING SEMAINE N° {week:02d} - {year}*.xls*"
        items = cp.list_onedrive_files(remote_dir, recursive=False, suffixes=[".xls", ".xlsx", ".xlsm"])
        files = [i.get("path", "") for i in items if fnmatch.fnmatch(i.get("name", ""), pattern)]

        def _sort_key(p: str):
            major, minor = _parse_version_from_name(Path(p))
            return (major, minor, 0)

        files.sort(key=_sort_key, reverse=True)
        return files

    base_dir = cp.ASF_ONEDRIVE / "Planning MAB" / f"ASFmm PLANNING {year}"
    if not base_dir.exists():
        return []
    pattern = f"ASFmm - PLANNING SEMAINE N° {week:02d} - {year}*.xls*"
    files = list(base_dir.glob(pattern))
    files = [p for p in files if p.is_file()]

    def _sort_key(p: Path):
        major, minor = _parse_version_from_name(p)
        try:
            mtime = p.stat().st_mtime
        except Exception:
            mtime = 0
        return (major, minor, mtime)

    files.sort(key=_sort_key, reverse=True)
    return files


# ---------------------------------------------------------------------------
# Planning match (pour récupérer vol / bénévole actuels)
# ---------------------------------------------------------------------------
def _match_planning_row(df_planning: pd.DataFrame, be_value: str) -> Optional[pd.Series]:
    if df_planning is None or df_planning.empty:
        return None

    be_col_planning = None
    for cand in ["BE_Numero", "BE NUMERO", "BE_NUMERO", "BE_Num", "BE_numero"]:
        if cand in df_planning.columns:
            be_col_planning = cand
            break
    if be_col_planning is None:
        return None

    be_norm = _norm_be(be_value)
    df_tmp = df_planning.copy()
    df_tmp["_BE_KEY"] = df_tmp[be_col_planning].apply(_norm_be)

    rows = df_tmp[df_tmp["_BE_KEY"] == be_norm]
    if rows.empty:
        # tenter sur suffixe 3 chiffres
        short = be_norm[-3:] if len(be_norm) >= 3 else be_norm
        rows = df_tmp[df_tmp["_BE_KEY"].str.endswith(short, na=False)]
    if rows.empty:
        return None
    return rows.iloc[0]


# ---------------------------------------------------------------------------
# Vols disponibles pour une destination
# ---------------------------------------------------------------------------
def _build_vol_options(dest_iata: str, df_vols: pd.DataFrame, df_planning: pd.DataFrame) -> list[Tuple[str, Tuple[str, str, str]]]:
    if df_vols is None:
        df_vols = pd.DataFrame()
    try:
        vol_source = df_vols.copy()
    except Exception:
        vol_source = pd.DataFrame()

    if vol_source.empty:
        return []

    vol_source = vol_source.copy()
    if "Dest_IATA" not in vol_source.columns:
        vol_source["Dest_IATA"] = vol_source.get("Destination", "")
    if "Routing" not in vol_source.columns:
        vol_source["Routing"] = vol_source.get("Routing_Str", "")
    if "Date_Vol" not in vol_source.columns:
        vol_source["Date_Vol"] = vol_source.get("Date", "")
    if "Numero_Vol" not in vol_source.columns:
        vol_source["Numero_Vol"] = vol_source.get("Vol", vol_source.get("Numero_Vol", ""))
    if "Heure_Vol" not in vol_source.columns:
        vol_source["Heure_Vol"] = vol_source.get("Heure", "")

    vol_source["Routing_Str"] = vol_source["Routing"].astype(str)
    vol_source["Dest_IATA_UP"] = vol_source["Dest_IATA"].astype(str).str.upper()

    vol_filtered = pd.DataFrame(columns=vol_source.columns)
    if dest_iata:
        mask_dest = vol_source["Dest_IATA_UP"].str.contains(dest_iata, na=False) | vol_source["Routing_Str"].str.upper().str.contains(dest_iata, na=False)
        vol_filtered = vol_source[mask_dest]
        if vol_filtered.empty:
            vol_filtered = vol_source

    if vol_filtered.empty and df_planning is not None and not df_planning.empty:
        df_plan_vols = df_planning.copy()
        df_plan_vols["Destination_UP"] = df_plan_vols.get("Destination", "").astype(str).str.upper()
        df_plan_vols = df_plan_vols[df_plan_vols["Destination_UP"].str.contains(dest_iata, na=False)]
        if not df_plan_vols.empty:
            df_plan_vols = df_plan_vols.rename(columns={"Vol": "Numero_Vol"})
            df_plan_vols["Routing_Str"] = df_plan_vols.get("Routing", "")
            vol_filtered = df_plan_vols

    vols_unique = (
        vol_filtered[["Date_Vol", "Numero_Vol", "Heure_Vol", "Routing_Str"]]
        .dropna(how="all")
        .drop_duplicates()
        .sort_values(by=["Date_Vol", "Heure_Vol"])
    ) if not vol_filtered.empty else pd.DataFrame(columns=["Date_Vol", "Numero_Vol", "Heure_Vol", "Routing_Str"])

    options: list[Tuple[str, Tuple[str, str, str]]] = []
    for _, r in vols_unique.iterrows():
        vol_num_raw = r.get("Numero_Vol", "")
        date_raw = r.get("Date_Vol", "")
        heure_raw = r.get("Heure_Vol", "")
        if (str(vol_num_raw).strip() == "") and (str(date_raw).strip() == ""):
            continue
        label = f"{_fmt_date_long(date_raw)} — {_fmt_vol(vol_num_raw)} — {_fmt_time(heure_raw)}"
        value = (str(date_raw), str(vol_num_raw), str(heure_raw))
        options.append((label, value))

    return options


# ---------------------------------------------------------------------------
# Bénévole : statut pour la date/heure choisie
# ---------------------------------------------------------------------------
def _prepare_dispo(df_dispos: pd.DataFrame) -> pd.DataFrame:
    df = df_dispos.copy()
    # Parsing sans warnings : détecte déjà datetime
    def _parse_date_safe(x):
        if isinstance(x, (dt.date, dt.datetime, pd.Timestamp)):
            return pd.to_datetime(x).date()
        return pd.to_datetime(str(x), errors="coerce", dayfirst=True).date()

    df["Date"] = df.get("Date").apply(_parse_date_safe)

    def _time_only(val):
        # heures en HHhMM / HH:MM / HH:MM:SS
        for fmt in ("%Hh%M", "%H:%M:%S", "%H:%M"):
            try:
                return pd.to_datetime(str(val), format=fmt, errors="raise").time()
            except Exception:
                continue
        t = pd.to_datetime(str(val), errors="coerce")
        return t.time() if pd.notna(t) else None

    df["Arr"] = df.get("Heure_Arrivee", "").apply(_time_only)
    df["Dep"] = df.get("Heure_Depart", "").apply(_time_only)
    return df


def _coerce_display_types(df: pd.DataFrame) -> pd.DataFrame:
    """
    Force certains champs à être affichés comme texte (téléphones, etc.)
    pour éviter les erreurs Arrow lors de l'aperçu.
    """
    if df is None or df.empty:
        return df
    out = df.copy()
    for col in out.columns:
        col_l = str(col).lower()
        if "telephone" in col_l or "phone" in col_l:
            out[col] = out[col].astype(str)
    return out


def _bene_status(df_dispo: pd.DataFrame, df_planning: pd.DataFrame, name: str, date_str: str, heure_str: str, vol_str: Optional[str] = None) -> str:
    try:
        d = pd.to_datetime(date_str).date()
        h = pd.to_datetime(heure_str).time()
    except Exception:
        return "indisponible"

    rows = df_dispo[df_dispo["Benevole"] == name]
    rows_same_day = rows[rows["Date"] == d]
    already = False
    if df_planning is not None and not df_planning.empty:
        mask = (
            (df_planning.get("Benevole", pd.Series(dtype=str)).astype(str) == str(name))
            & (df_planning.get("Date_Vol", pd.Series(dtype=str)).astype(str) == str(date_str))
        )
        if vol_str is not None:
            mask = mask & (df_planning.get("Numero_Vol", pd.Series(dtype=str)).astype(str) == str(vol_str))
        already = mask.any()

    if already:
        return "déjà affecté sur ce créneau"
    if rows_same_day.empty:
        return "indisponible"

    has_info = False
    ok_dispo = False
    for _, r in rows_same_day.iterrows():
        arr = r["Arr"]
        dep = r["Dep"]
        if arr is None and dep is None:
            continue
        has_info = True
        arr = arr or h
        dep = dep or h
        if arr <= h <= dep:
            ok_dispo = True
            break

    if not has_info:
        return "inconnu"
    return "disponible" if ok_dispo else "indisponible"


# ---------------------------------------------------------------------------
# Phrase d'action
# ---------------------------------------------------------------------------
def _build_action_sentence(be_num: str, dest_iata: str, date_initial: str, action: str, new_date: str, vol_disp: str, bene_short: str) -> str:
    prefix = f"Le BE {be_num}, destination {dest_iata}, initialement prévu le {date_initial}"
    if action == "Annulation":
        return f"{prefix} sera annulé."
    if action == "Ajouter au planning":
        return f"Le BE {be_num}, destination {dest_iata}, sera ajouté le {new_date} sur le vol {vol_disp} avec {bene_short}."
    return f"{prefix} sera reprogrammé le {new_date} sur le vol {vol_disp} avec {bene_short}."


def _collect_be_from_planning(df_prev: pd.DataFrame, week: int, year: int) -> pd.DataFrame:
    """
    Extrait les BE présents dans le planning exporté sélectionné.
    """
    if df_prev is None or df_prev.empty:
        return pd.DataFrame()

    df = df_prev.copy()

    # BE
    be_cols = [c for c in df.columns if "BE" in str(c).upper() and "NUM" in str(c).upper()]
    be_col = be_cols[0] if be_cols else None

    if be_col is None:
        return pd.DataFrame()

    df_out = pd.DataFrame()
    df_out["BE_Numero_Str"] = df[be_col].astype(str).str.replace(r"\.0$", "", regex=True).str.strip()

    # Destination
    dest_col = None
    for cand in ["Destination", "Dest_IATA", "Ville", "DESTINATION"]:
        if cand in df.columns:
            dest_col = cand
            break
    df_out["Destination"] = df.get(dest_col, "")

    # Date vol
    date_col = None
    for cand in ["Date_Vol", "DATE", "Date", "Date Vol"]:
        if cand in df.columns:
            date_col = cand
            break
    df_out["Date_Vol"] = pd.to_datetime(
        df.get(date_col, ""),
        errors="coerce",
        dayfirst=True,
        format="%d/%m/%y",
    )
    iso = df_out["Date_Vol"].dt.isocalendar()
    df_out["Week"] = iso.week.astype("Int64")
    df_out["Year"] = iso.year.astype("Int64")
    df_out = df_out[(df_out["Week"] == week) & (df_out["Year"] == year)]

    # Vol / Heure
    vol_col = None
    for cand in ["Vol", "Numero_Vol", "NUMERO VOL", "Numero Vol"]:
        if cand in df.columns:
            vol_col = cand
            break
    heure_col = None
    for cand in ["Heure_Vol", "Heure", "HEURE VOL", "HEURE"]:
        if cand in df.columns:
            heure_col = cand
            break
    df_out["Numero_Vol"] = df.get(vol_col, "")
    df_out["Heure_Vol"] = df.get(heure_col, "")
    # Nb colis / Type si présents
    coli_col = None
    for cand in ["BE_Nb_Colis", "Nb_Colis", "NB COLIS", "NB_COLIS"]:
        if cand in df.columns:
            coli_col = cand
            break
    type_col = None
    for cand in ["BE_Type", "Type", "TYPE"]:
        if cand in df.columns:
            type_col = cand
            break
    df_out["BE_Nb_Colis"] = df.get(coli_col, 0)
    df_out["BE_Type"] = df.get(type_col, "")
    df_out["Source"] = "planning"
    return df_out


def _find_row_in_df(df: pd.DataFrame, be_num: str) -> Optional[pd.Series]:
    """Retourne la première ligne du df dont le numéro BE correspond."""
    if df is None or df.empty:
        return None
    df_tmp = df.copy()
    be_cols = [c for c in df_tmp.columns if "BE" in str(c).upper() and "NUM" in str(c).upper()]
    for col in be_cols:
        df_tmp["_BE_MATCH"] = df_tmp[col].astype(str).str.replace(r"\\.0$", "", regex=True).str.strip()
        match = df_tmp[df_tmp["_BE_MATCH"].str.endswith(str(be_num).strip())]
        if not match.empty:
            return match.iloc[0]
    return None


def _apply_planning_update(
    path: Path,
    action: str,
    be_num: str,
    dest_iata: str,
    date_new: str,
    vol_new: str,
    heure_new: str,
    bene_choice: str,
    be_info: pd.Series,
    plan_row: Optional[pd.Series] = None,
    plan_row_full: Optional[pd.Series] = None,
    bene_meta: Optional[dict] = None,
    bene_changed: bool = False,
):
    """
    Nouvelle approche : construit un DF consolidé, applique l'action, puis regénère Export/Planning.
    """
    import datetime as _dt
    import openpyxl
    from openpyxl.styles import PatternFill

    def _norm_be(val: str) -> str:
        return normalize_be_number(val) or str(val)

    # Lecture Export planning en DF
    try:
        df_export = pd.read_excel(path, sheet_name="Export planning")
    except Exception:
        df_export = pd.DataFrame()
    # Fallback Planning si Export planning absent
    if df_export.empty:
        try:
            df_export = pd.read_excel(path, sheet_name=0)
        except Exception:
            df_export = pd.DataFrame()

    # Normalisation minimale
    df_export = df_export.copy()
    df_export.columns = [str(c) for c in df_export.columns]
    if "BE_Numero" not in df_export.columns:
        df_export["BE_Numero"] = df_export.get("BE_NUM", df_export.get("BE", ""))
    df_export["BE_Key"] = df_export["BE_Numero"].apply(_norm_be)
    if "Date_Vol" in df_export.columns:
        df_export["Date_Vol"] = parse_date_series(df_export["Date_Vol"]).dt.date
    if "Heure_Vol" in df_export.columns:
        df_export["Heure_Vol"] = parse_time_series(df_export["Heure_Vol"]).dt.time
        df_export["HEURE_MIN"] = hour_min_from_series(df_export["Heure_Vol"])
    if "_STATUS" not in df_export.columns:
        df_export["_STATUS"] = "normal"

    # Appliquer l'action sur le DF
    if action == "Annulation":
        df_export.loc[df_export["BE_Key"] == _norm_be(be_num), "_STATUS"] = "old"
    else:
        # marquer l'ancienne ligne en old
        df_export.loc[df_export["BE_Key"] == _norm_be(be_num), "_STATUS"] = "old"
        # construire la nouvelle ligne
        new_row = {
            "BE_Numero": be_num,
            "BE_Key": _norm_be(be_num),
            "Destination": dest_iata,
            "IATA": dest_iata,
            "Date_Vol": parse_date_series(pd.Series([date_new])).iloc[0].date() if date_new else None,
            "Heure_Vol": parse_time_series(pd.Series([heure_new])).iloc[0].time() if heure_new else None,
            "Heure": normalize_hour_str(pd.Series([heure_new])).iloc[0],
            "HEURE_MIN": hour_min_from_series(pd.Series([heure_new])).iloc[0],
            "Numero_Vol": vol_new,
            "Numero_Vol": vol_new,
            "Routing": "",
            "Benevole": bene_choice,
            "BE_Nb_Colis": be_info.get("BE_Nb_Colis", be_info.get("Nb_Colis", "")),
            "BE_Nb_Equiv": be_info.get("BE_Nb_Equiv", be_info.get("Equiv_Colis", "")),
            "BE_Type": be_info.get("BE_Type", ""),
            "BE_Expediteur": be_info.get("BE_Expediteur", ""),
            "BE_Destinataire": be_info.get("BE_Destinataire", ""),
            "_STATUS": "new",
        }
        df_export = pd.concat([df_export, pd.DataFrame([new_row])], ignore_index=True)

    # Tri
    df_export = df_export.sort_values(by=["Date_Vol", "Heure_Vol", "BE_Numero"], kind="mergesort").reset_index(drop=True)

    # Écriture Excel
    wb = openpyxl.load_workbook(path)
    # incrémenter Q1 (version) si existant
    try:
        ws_plan = wb.worksheets[0]
        q1_val = ws_plan["Q1"].value
        ws_plan["Q1"].value = (int(q1_val) if q1_val not in (None, "") else 0) + 1
    except Exception:
        pass

    def _clear_values(ws, max_row: int, max_col: int):
        for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
            for cell in row:
                cell.value = None

    # Export planning
    ws_exp = (
        wb["Export planning"]
        if "Export planning" in wb.sheetnames
        else wb.create_sheet("Export planning", 1)
    )
    headers = list(df_export.columns)
    max_row = max(ws_exp.max_row, len(df_export) + 1)
    max_col = max(ws_exp.max_column, len(headers))
    _clear_values(ws_exp, max_row, max_col)
    for c_idx, h in enumerate(headers, start=1):
        ws_exp.cell(row=1, column=c_idx, value=h)
    for r_idx, (_, r) in enumerate(df_export.iterrows(), start=2):
        for c_idx, h in enumerate(headers, start=1):
            ws_exp.cell(row=r_idx, column=c_idx, value=r.get(h, ""))

    # Planning : reprendre uniquement les colonnes utiles et appliquer couleurs
    ws_plan_new = (
        wb["Planning"]
        if "Planning" in wb.sheetnames
        else wb.create_sheet("Planning", 0)
    )
    planning_cols = [
        "Benevole", "unusedE", "unusedF", "Destination", "IATA", "Routing",
        "Numero_Vol", "Heure_Vol", "BE_Numero", "BE_Nb_Colis", "BE_Type",
        "unusedN", "BE_Expediteur", "BE_Destinataire"
    ]
    # Squelettes colonnes D..Q (4..17)
    headers_plan = ["", "", "", "Benevole", "", "Destination", "IATA", "Routing", "Numero_Vol", "Heure_Vol", "BE_Numero", "BE_Nb_Colis", "BE_Type", "", "BE_Expediteur", "BE_Destinataire"]
    max_row = max(ws_plan_new.max_row, len(df_export) + 1)
    max_col = max(ws_plan_new.max_column, len(headers_plan))
    _clear_values(ws_plan_new, max_row, max_col)
    for row in ws_plan_new.iter_rows(min_row=2, max_row=max_row, min_col=4, max_col=17):
        for cell in row:
            cell.fill = PatternFill()
    for c_idx, val in enumerate(headers_plan, start=1):
        ws_plan_new.cell(row=1, column=c_idx, value=val)
    fill_red = PatternFill(fill_type="solid", fgColor="F8CBAD")
    fill_blue = PatternFill(fill_type="solid", fgColor="BDD7EE")
    for r_idx, (_, r) in enumerate(df_export.iterrows(), start=2):
        status = str(r.get("_STATUS", "normal")).lower()
        row_vals = ["", "", "", r.get("Benevole", ""), "", r.get("Destination", ""), r.get("IATA", ""), r.get("Routing", ""), r.get("Numero_Vol", ""), normalize_hour_str(pd.Series([r.get("Heure_Vol", "")])).iloc[0], r.get("BE_Numero", ""), "" if status.startswith("old") else r.get("BE_Nb_Colis", ""), r.get("BE_Type", ""), "", r.get("BE_Expediteur", ""), r.get("BE_Destinataire", "")]
        for c_idx, val in enumerate(row_vals, start=1):
            ws_plan_new.cell(row=r_idx, column=c_idx, value=val)
        if status.startswith("old") or status.startswith("new"):
            fill = fill_red if status.startswith("old") else fill_blue
            for c in range(4, 18):
                ws_plan_new.cell(row=r_idx, column=c).fill = fill

    wb.save(path)
    cp.sync_local_file_to_onedrive(path)


# ---------------------------------------------------------------------------
# UI principale
# ---------------------------------------------------------------------------
def render_tab_shipments_update():
    st.title("🚚 Mise à Jour expéditions")

    planning_state = get_planning_state()
    df_planning = (
        normalize_planning_df(planning_state.planning)
        if planning_state and planning_state.planning is not None
        else pd.DataFrame()
    )

    with st.spinner("Chargement des paramètres…"):
        df_paramdest, df_paramexpediteur, df_parambenev, _ = load_parameters()
    df_vols = get_vols_df_cached()
    df_dispos_raw = get_benevoles_cached()
    df_dispos = _prepare_dispo(df_dispos_raw)

    dest_city_map, city_to_iata_map = build_iata_city_maps(df_paramdest)

    # Sélecteur semaine (tri décroissant)
    weeks_set = _available_weeks_from_exports()
    df_be_p = None
    if not weeks_set:
        # fallback : déduire depuis MAG central (statut D)
        df_be_p = _load_be_status("D")
        weeks_set = {
            (int(w), int(y))
            for w, y in df_be_p[["Week", "Year"]].dropna().astype(int).itertuples(index=False, name=None)
        }

    weeks = sorted(weeks_set, key=lambda t: (t[1], t[0]), reverse=True)

    if not weeks:
        st.warning("Impossible d'extraire les numéros de semaine.")
        return

    week_labels = [f"{y} - Semaine {w:02d}" for w, y in weeks]
    week_map = {label: pair for label, pair in zip(week_labels, weeks)}

    choice_week_label = st.selectbox("Choisir la semaine", week_labels, index=0)
    selected_week, selected_year = week_map.get(choice_week_label, weeks[0])

    # Sélection de la version du planning (recherche fichiers vX)
    planning_candidates = _find_planning_files_for_week(selected_week, selected_year)
    chosen_path = None
    if planning_candidates:
        labels = []
        path_map = {}
        for p in planning_candidates:
            p_path = Path(p) if isinstance(p, str) else p
            major, minor = _parse_version_from_name(p_path)
            ver_label = f"v{major}" + (f"-{minor}" if minor else "")
            label = f"{ver_label} — {p_path.name}"
            labels.append(label)
            path_map[label] = p
        # ordre déjà décroissant, on garde index 0
        choice_label = st.selectbox("Choisir la version du planning", labels, index=0)
        chosen_path = path_map.get(choice_label)
    else:
        st.info("Aucune version de planning trouvée pour cette semaine, recherche par défaut.")

    # Aperçu du planning validé (OneDrive)
    df_preview, msg_preview, preview_path = _load_planning_preview_with_path(selected_week, selected_year, chosen_path)
    with st.expander("Aperçu du planning sélectionné", expanded=False):
        st.caption(msg_preview)
        if df_preview is not None and not df_preview.empty:
            df_preview = _coerce_display_types(df_preview)
            # Harmonise dates/heures pour éviter les erreurs Arrow
            for col in df_preview.columns:
                if "Date" in col:
                    df_preview[col] = pd.to_datetime(df_preview[col], errors="coerce", dayfirst=True)
                    df_preview[col] = df_preview[col].dt.strftime("%d/%m/%y")
                if "Heure" in col:
                    df_preview[col] = df_preview[col].apply(
                        lambda t: t.strftime("%Hh%M") if isinstance(t, (dt.time, pd.Timestamp)) else str(t)
                    )
            # Catch-all : toute valeur datetime/time restante -> str
            def _fmt_cell(v):
                if isinstance(v, (dt.datetime, dt.date, pd.Timestamp)):
                    return v.strftime("%d/%m/%y") if pd.notna(v) else ""
                if isinstance(v, dt.time):
                    return v.strftime("%Hh%M")
                return v

            df_preview = df_preview.apply(lambda col: col.map(_fmt_cell))
            st.dataframe(df_preview.head(100), height=360, hide_index=True, width="stretch")
        else:
            st.warning("Aucun aperçu disponible pour cette semaine.")

    # BE issus du planning sélectionné + BE statut D (MAG CENTRAL) sur la semaine
    df_be_plan = _collect_be_from_planning(df_preview, selected_week, selected_year)
    df_be_d = _load_be_status_d_for_week(selected_week, selected_year)

    df_be_all = pd.DataFrame()
    if df_be_plan is not None and not df_be_plan.empty:
        df_be_all = pd.concat([df_be_all, df_be_plan], ignore_index=True)
    if df_be_d is not None and not df_be_d.empty:
        df_be_all = pd.concat([df_be_all, df_be_d], ignore_index=True)

    if df_be_all.empty:
        st.info("Aucun BE trouvé pour cette semaine (planning + statut D).")
        return

    df_be_all["BE_Numero_Str"] = df_be_all["BE_Numero_Str"].fillna("").astype(str)
    df_be_all = df_be_all[df_be_all["BE_Numero_Str"].str.strip() != ""]
    if df_be_all.empty:
        st.info("Aucun BE identifié (numéro manquant).")
        return
    # Mapping IATA pour l'affichage
    def _dest_iata(val):
        return _dest_to_iata(val, df_paramdest)
    df_be_all["Dest_IATA_Label"] = df_be_all.get("Destination", "").apply(_dest_iata)
    # Normalisation BE : clé unique sur 6 chiffres, priorité aux lignes issues du planning et aux clés non "00xxxx"
    df_be_all["BE_Key"] = df_be_all["BE_Numero_Str"].apply(_norm_be)
    df_be_all = df_be_all[df_be_all["BE_Key"] != ""]
    df_be_all["BE_Num_Display"] = df_be_all["BE_Key"]  # pour affichage unique
    df_be_all["Source_rank"] = df_be_all.get("Source", "").astype(str).str.lower().eq("planning").astype(int)
    df_be_all["Prefix_rank"] = (~df_be_all["BE_Key"].str.startswith("00")).astype(int)
    df_be_all["TAIL4"] = df_be_all["BE_Key"].str[-4:]
    # Si une clé commence par 00 mais qu'une clé non-00 existe avec les mêmes 4 derniers digits, on retire la clé 00
    has_non_zero_tail = df_be_all.groupby("TAIL4")["Prefix_rank"].transform(lambda s: (s == 1).any())
    mask_drop = (df_be_all["BE_Key"].str.startswith("00")) & has_non_zero_tail
    df_be_all = df_be_all[~mask_drop]
    df_be_all = df_be_all.sort_values(
        by=["Dest_IATA_Label", "TAIL4", "Prefix_rank", "Source_rank", "BE_Key"],
        ascending=[True, True, False, False, True],
    )
    # Suppression stricte des doublons de clés (conserver la 1ère occurrence déjà triée)
    df_be_all = df_be_all[~df_be_all["BE_Key"].duplicated(keep="first")]
    be_lookup = df_be_all.set_index("BE_Key")

    def _format_be_option(num_str: str) -> str:
        if num_str in be_lookup.index:
            r = be_lookup.loc[num_str]
            dest = str(r.get("Dest_IATA_Label", r.get("Destination", "")) or "").upper()
            nb = pd.to_numeric(r.get("BE_Nb_Colis", 0), errors="coerce")
            nb_int = int(nb) if pd.notna(nb) else (r.get("BE_Nb_Colis", "") or "")
            be_type = str(r.get("BE_Type", "") or "").upper()
            date_disp = _fmt_date_long(r.get("Date_Vol", ""))
            date_part = date_disp if date_disp not in (None, "", "NaT") else "A planifier"
            be_disp = str(r.get("BE_Num_Display", num_str)).zfill(6)
            return f"{dest} - BE {be_disp} - {nb_int} colis - {be_type} - {date_part}"
        return str(num_str)

    selected_be = st.selectbox(
        "Sélectionner un BE",
        be_lookup.index.tolist(),
        format_func=_format_be_option,
    )

    if not selected_be:
        return

    be_row = be_lookup.loc[selected_be]
    dest_raw = be_row.get("Destination", be_row.get("Dest_IATA_Label", ""))
    dest_iata = _dest_to_iata(dest_raw, df_paramdest)
    date_initial = be_row.get("Date_Vol", "")
    date_initial_long = _fmt_date_long(date_initial)
    expediteur_name = str(be_row.get("BE_Expediteur", "") or "").strip()
    be_source = str(be_row.get("Source", "")).lower()

    plan_row = _match_planning_row(df_planning, selected_be)
    current_vol = ""
    current_heure = ""
    current_bene = ""
    bene_prenom_court = ""
    bene_nom = ""

    if plan_row is not None:
        date_initial = plan_row.get("Date_Vol", date_initial)
        date_initial_long = _fmt_date_long(date_initial)
        current_vol = plan_row.get("Numero_Vol", "")
        current_heure = plan_row.get("Heure_Vol", plan_row.get("Heure", ""))
        current_bene = plan_row.get("Benevole", "")
        bene_prenom_court = plan_row.get("Benevole_Prenom_Court", plan_row.get("Prenom_Court", ""))
        bene_nom = plan_row.get("Benevole_Nom", plan_row.get("Nom", ""))

    if not bene_prenom_court or not bene_nom:
        try:
            row_b = df_parambenev[df_parambenev["Benevole"] == current_bene]
            if not row_b.empty:
                bene_prenom_court = row_b.get("Prenom_Court", pd.Series([""])).iloc[0]
                bene_nom = row_b.get("Nom", pd.Series([""])).iloc[0]
        except Exception:
            pass

    if be_source == "planning":
        action_options = ["Annulation", "Changement de date ou bénévole"]
    else:
        action_options = ["Ajouter au planning"]
    action_choice = st.radio("Action", action_options, horizontal=True)

    vol_options = []
    vol_labels: list[str] = []
    vol_values: list[Tuple[str, str, str]] = []
    default_vol_tuple = (str(date_initial), str(current_vol), str(current_heure))
    action_requires_assignment = action_choice != "Annulation"
    if action_requires_assignment:
        col_vol, col_bene = st.columns(2)
        with col_vol:
            vol_options = _build_vol_options(dest_iata, df_vols, df_planning)
            vol_labels = [v[0] for v in vol_options] or ["Aucun vol disponible"]
            vol_values = [v[1] for v in vol_options] or [("", "", "")]
            default_idx = vol_values.index(default_vol_tuple) if default_vol_tuple in vol_values else 0
            vol_choice = st.selectbox("Sélectionner un vol", vol_labels, index=default_idx if vol_labels else 0)
            date_new, vol_new, heure_new = vol_values[vol_labels.index(vol_choice)] if vol_labels else ("", "", "")
        with col_bene:
            bene_options: list[str] = []
            bene_choice = ""
            if df_parambenev is not None and not df_parambenev.empty:
                for name in sorted(df_parambenev["Benevole"].dropna().unique()):
                    status = _bene_status(df_dispos, df_planning if df_planning is not None else pd.DataFrame(), name, date_new, heure_new, vol_new)
                    bene_options.append(f"{name} ({status})")
            default_bene_label = None
            if current_bene:
                status_def = _bene_status(df_dispos, df_planning if df_planning is not None else pd.DataFrame(), current_bene, date_new, heure_new, vol_new)
                default_bene_label = f"{current_bene} ({status_def})"
            bene_idx = bene_options.index(default_bene_label) if default_bene_label in bene_options else 0
            bene_choice_label = st.selectbox("Sélectionner un bénévole", bene_options or ["Aucun bénévole disponible"], index=bene_idx if bene_options else 0)
            bene_choice = bene_choice_label.split(" (")[0] if bene_options else ""
            if bene_choice and (not bene_prenom_court or not bene_nom):
                try:
                    row_b = df_parambenev[df_parambenev["Benevole"] == bene_choice]
                    if not row_b.empty:
                        bene_prenom_court = row_b.get("Prenom_Court", pd.Series([""])).iloc[0]
                        bene_nom = row_b.get("Nom", pd.Series([""])).iloc[0]
                except Exception:
                    pass
    else:
        date_new, vol_new, heure_new = str(date_initial), str(current_vol), str(current_heure)
        bene_choice = current_bene

    # Infos bénévole pour override
    bene_meta = {}
    if df_parambenev is not None and not df_parambenev.empty and bene_choice:
        row_bm = df_parambenev[df_parambenev["Benevole"] == bene_choice]
        if not row_bm.empty:
            rb = row_bm.iloc[0]
            bene_meta = {
                "Benevole": bene_choice,
                "ID": rb.get("ID", ""),
                "Telephone": rb.get("Telephone", ""),
                "Benevole_Prenom": rb.get("Prenom", ""),
                "Benevole_Prenom_Court": rb.get("Prenom_Court", ""),
                "Benevole_Nom": rb.get("Nom", ""),
                "ID_UP": rb.get("ID", ""),
                "Benev_UP": bene_choice.upper(),
                "ID_BEN_ID": rb.get("ID", ""),
                "Benevole_BEN_ID": rb.get("Benevole", bene_choice),
                "Nom": rb.get("Nom", ""),
                "Prenom": rb.get("Prenom", ""),
                "Prenom_Court": rb.get("Prenom_Court", ""),
                "Max_Jours_Semaine": rb.get("Max_Jours_Semaine", ""),
                "Max_Exp_Semaine": rb.get("Max_Exp_Semaine", ""),
                "Max_Exp_Jour": rb.get("Max_Exp_Jour", ""),
                "Attente_Max_Heures": rb.get("Attente_Max_Heures", ""),
                "Telephone_BEN_ID": rb.get("Telephone", rb.get("Telephone_BEN_ID", "")),
                "Email": rb.get("Email", ""),
                "Benev_UP_BEN_ID": str(rb.get("Benevole", bene_choice)).upper(),
                "ID_BENEVOLE": rb.get("ID", ""),
            }
    bene_changed = (bene_choice or "") != (current_bene or "")

    bene_short = f"{str(bene_prenom_court).strip()} {str(bene_nom).strip().upper()}".strip() if (bene_prenom_court or bene_nom) else bene_choice
    vol_disp = _fmt_vol(vol_new) if vol_new else (_fmt_vol(current_vol) if current_vol else "")
    date_new_long = _fmt_date_long(date_new)
    action_sentence = _build_action_sentence(
        be_num=str(selected_be),
        dest_iata=dest_iata,
        date_initial=date_initial_long,
        action=action_choice,
        new_date=date_new_long or date_initial_long,
        vol_disp=vol_disp or "(vol ?)",
        bene_short=bene_short or "(bénévole ?)",
    )

    st.markdown(f"**Action prévue :** {action_sentence}")
    btn_disabled = action_requires_assignment and (not date_new_long or not vol_new or not bene_choice)

    def _open_file(path_obj: Path | None):
        if not path_obj:
            return
        try:
            import platform, subprocess, os
            if platform.system() == "Darwin":
                subprocess.Popen(["open", str(path_obj)])
            elif platform.system() == "Windows":
                os.startfile(str(path_obj))  # type: ignore[attr-defined]
            else:
                subprocess.Popen(["xdg-open", str(path_obj)])
        except Exception:
            pass

    payload = {
        "week": selected_week,
        "year": selected_year,
        "dest_iata": dest_iata,
        "dest_label": dest_raw,
        "be": str(selected_be),
        "date_initial_long": date_initial_long,
        "date_new_long": date_new_long or date_initial_long,
        "vol_display": vol_disp,
        "bene_short": bene_short,
        "expediteur": expediteur_name,
        "action": action_choice,
        "action_sentence": action_sentence,
        "source": be_source,
        "planning_path": str(preview_path) if preview_path else "",
        "be_info": be_row.to_dict(),
    }

    if st.button("Valider le planning et exporter en Excel + PDF", type="primary", disabled=btn_disabled):
        if not preview_path or not Path(preview_path).exists():
            st.error("Impossible de trouver le fichier planning à mettre à jour.")
            return
        plan_row_full = _find_row_in_df(df_preview, selected_be) if df_preview is not None else None
        try:
            _apply_planning_update(
                Path(preview_path),
                action_choice,
                str(selected_be),
                dest_iata,
                date_new,
                vol_new,
                heure_new,
                bene_choice or current_bene,
                be_row,
                plan_row,
                plan_row_full,
                bene_meta=bene_meta if bene_meta is not None else {},
                bene_changed=bene_changed,
            )
            st.success("Planning Excel mis à jour.")
            try:
                pdf_path = export_first_sheet_to_pdf(Path(preview_path))
                st.success(f"PDF généré : {pdf_path.name}")
                _open_file(Path(preview_path))
                _open_file(pdf_path)
            except Exception as e_pdf:
                st.warning(f"PDF non généré automatiquement : {e_pdf}")
        except Exception as e:
            st.error(f"Erreur lors de la mise à jour du planning : {e}")
            return

        st.session_state["ship_update_payload"] = payload
        st.success("Mise à jour validée. Choisissez qui prévenir ci-dessous.")

    payload_state = st.session_state.get("ship_update_payload")
    if not payload_state:
        return

    st.divider()
    st.subheader("Notifications")

    def _body_lines() -> list[str]:
        action_line = f"{payload_state['dest_iata']} : {payload_state['action_sentence']}"
        return [
            "Bonjour,",
            "",
            f"Mise à jour du planning S{payload_state['week']:02d} - {payload_state['year']} :",
            "",
            action_line,
            "",
            "Cordialement,",
        ]

    # Emails bénévoles (selon action)
    bene_emails: list[str] = []
    def _bene_email(name: str) -> str:
        if df_parambenev is None or getattr(df_parambenev, "empty", True):
            return ""
        row = df_parambenev[df_parambenev["Benevole"] == name]
        if row.empty:
            return ""
        return str(row.get("Email", pd.Series([""])).iloc[0]).strip()

    if action_choice == "Changement de date ou bénévole":
        # Si bénévole change, prévenir l'ancien et le nouveau
        for ben in {current_bene, bene_choice}:
            mail = _bene_email(ben)
            if mail:
                bene_emails.append(mail)
    else:
        mail = _bene_email(bene_choice or current_bene)
        if mail:
            bene_emails.append(mail)

    col_asf, col_dest, col_exp = st.columns(3)

    with col_asf:
        to_list_asf = ["messmed@aviation-sans-frontieres-fr.org", *[m for m in bene_emails if m]]
        if st.button("Prévenir ASF + Bénévole", key="btn_mail_asf"):
            create_outlook_draft(
                to_list=to_list_asf,
                cc_list=None,
                subject=f"MAJ Planning S{payload_state['week']:02d}",
                body_html=_wrap_body(_body_lines()),
                attachments=None,
                use_signature=True,
            )
            st.success("Brouillon ASF ouvert.")

    with col_dest:
        dest_for_email = payload_state.get("dest_label") or payload_state.get("dest_iata")
        to_dest, cc_dest = _get_emails_for_destination(df_paramdest, dest_for_email)
        if not to_dest:
            # Fallback : essayer directement le code IATA
            to_dest, cc_dest = _get_emails_for_destination(df_paramdest, payload_state.get("dest_iata", ""))
        cc_dest = [*cc_dest] if isinstance(cc_dest, list) else [cc_dest] if cc_dest else []
        cc_dest.append("messmed@aviation-sans-frontieres-fr.org")
        if st.button("Prévenir Escale", key="btn_mail_dest", disabled=not to_dest):
            create_outlook_draft(
                to_list=to_dest,
                cc_list=cc_dest,
                subject=f"MAJ Planning S{payload_state['week']:02d} - {payload_state['dest_iata']}",
                body_html=_wrap_body(_body_lines()),
                attachments=None,
                use_signature=True,
            )
            st.success("Brouillon Escale ouvert." if to_dest else "Aucun email ParamDest trouvé.")

    with col_exp:
        exp_name_state = payload_state.get("expediteur", "")
        if exp_name_state and str(exp_name_state).upper() != "ASF":
            to_exp, cc_exp = _get_emails_for_expediteur(df_paramexpediteur, exp_name_state)
            cc_exp = [*cc_exp] if isinstance(cc_exp, list) else [cc_exp] if cc_exp else []
            cc_exp.append("messmed@aviation-sans-frontieres-fr.org")
            subject_exp = f"{exp_name_state} - MAJ Planning S{payload_state['week']:02d} - {payload_state['dest_iata']}"
            if st.button("Prévenir Expéditeur", key="btn_mail_exp", disabled=not to_exp):
                be_fmt = _norm_be(payload_state.get("be", ""))
                action_part = "est annulé."
                if payload_state.get("action") == "Changement de date ou bénévole":
                    action_part = (
                        f"est re-planifié au {payload_state['date_new_long']} / "
                        f"{payload_state['vol_display'] or '(vol ?)'}"
                    )
                if payload_state.get("action") == "Ajouter au planning":
                    action_part = (
                        f"sera ajouté le {payload_state['date_new_long']} / "
                        f"{payload_state['vol_display'] or '(vol ?)'}"
                    )
                body_lines_exp = [
                    "Bonjour,",
                    "",
                    f"Nous tenons à vous informer d'une Mise à jour du planning S{payload_state['week']:02d} - {payload_state['year']} :",
                    "",
                    f"{payload_state['dest_iata']} : Le BE {be_fmt}, initialement prévu le {payload_state['date_initial_long']} {action_part}",
                    "",
                    "Cordialement,",
                ]
                create_outlook_draft(
                    to_list=to_exp,
                    cc_list=cc_exp,
                    subject=subject_exp,
                    body_html=_wrap_body(body_lines_exp),
                    attachments=None,
                    use_signature=True,
                )
                st.success("Brouillon Expéditeur ouvert.")
        else:
            st.info("Expéditeur ASF : pas de notification expéditeur.")

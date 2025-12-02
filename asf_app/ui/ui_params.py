# asf_app/ui/ui_params.py
# -*- coding: utf-8 -*-

import os
from pathlib import Path

import pandas as pd
import streamlit as st

from asf_app.state import get_state

from scheduler.config_paths import (
    TABLEAU_DE_BORD_SRC,
    PLANNING_BENEVOLES_SRC,
    SHEET_PARAM_DEST,
    SHEET_PARAM_BE,
    SHEET_PARAM_BENEV,
    detect_onedrive_asf,
    ASF_ONEDRIVE,
)

from loaders.universal_loader import load_and_normalize
from scheduler.column_map import (
    column_map_param_dest,
    column_map_param_be,
    column_map_param_benev,
)


# ================================================================
# HELPERS
# ================================================================

def load_param_df(state, attr_name, path, sheet, mapping, header=0):
    """
    Charge un DF Param* dans STATE si vide.
    """
    df = getattr(state, attr_name)
    if df is None:
        df = load_and_normalize(path, sheet, mapping, header=header)
        df = df.reset_index(drop=True)
        setattr(state, attr_name, df)
    return df


def reload_param_df(state, attr_name, path, sheet, mapping, header=0):
    """
    Recharge explicitement un DF Param* depuis Excel vers STATE.
    """
    df = load_and_normalize(path, sheet, mapping, header=header)
    df = df.reset_index(drop=True)
    setattr(state, attr_name, df)
    return df


def write_excel_sheet(path: Path, sheet_name: str, df: pd.DataFrame):
    """
    Réécriture propre d’une feuille dans un Excel.
    """
    from openpyxl import load_workbook

    try:
        wb = load_workbook(path)
        if sheet_name in wb.sheetnames:
            wb.remove(wb[sheet_name])

        ws = wb.create_sheet(title=sheet_name)

        ws.append(list(df.columns))
        for row in df.itertuples(index=False, name=None):
            ws.append(list(row))

        wb.save(path)
        return True

    except Exception as e:
        st.error(f"❌ Erreur écriture Excel ({sheet_name}) : {e}")
        return False


# ================================================================
# UI PRINCIPALE
# ================================================================

def render_tab_params():

    state = get_state()

    st.header("⚙️ Paramètres & Tables de configuration")

    st.info("⚠️ Toutes les modifications sont valables **uniquement pour la session**.\n"
            "Aucune modification permanente n'est écrite dans le moteur ASF.")

    st.divider()

    # Bloc OneDrive (détection + override)
    with st.expander("🌥️ Chemins OneDrive / Sources", expanded=False):
        autodetect = detect_onedrive_asf()
        st.markdown(f"**OneDrive détecté :** `{autodetect}`")
        st.markdown(f"**OneDrive courant (ASF_ONEDRIVE) :** `{ASF_ONEDRIVE}`")
        new_root = st.text_input(
            "Chemin OneDrive (override pour la session)",
            value=str(ASF_ONEDRIVE),
            help="Laisse vide pour conserver la détection automatique. Exemple : ~/Library/CloudStorage/OneDrive-XXX",
        )
        if st.button("🔄 Appliquer le chemin OneDrive (session)", key="btn_onedrive_override"):
            if new_root.strip():
                os.environ["ASF_ONEDRIVE_ROOT"] = new_root.strip()
                st.success(f"Chemin OneDrive surchargé pour la session : {new_root}")
                st.rerun()
            else:
                st.info("Aucun chemin saisi, la détection automatique reste active.")

    # ---------------------------------------------------------
    # SELECTION DU BLOC
    # ---------------------------------------------------------
    if "active_block" not in st.session_state:
        st.session_state.active_block = None

    def set_block(name):
        st.session_state.active_block = name if st.session_state.active_block != name else None

    col1, col2, col3, col4 = st.columns(4)

    with col1:
        if st.button("⚙️ Paramoteur", width="stretch"):
            set_block("paramoteur")

    with col2:
        if st.button("🗂️ ParamDest", width="stretch"):
            set_block("paramdest")

    with col3:
        if st.button("📦 ParamBE", width="stretch"):
            set_block("parambe")

    with col4:
        if st.button("👥 ParamBenev", width="stretch"):
            set_block("parambenev")

    st.markdown("")

    # ============================================================
    # BLOCS UI
    # ============================================================

    # ------------------------------------------------------------
    # 1) PARAMOTEUR (SESSION ONLY)
    # ------------------------------------------------------------
    if st.session_state.active_block == "paramoteur":

        st.subheader("⚙️ Paramètres moteur (SESSION UNIQUEMENT)")

        # Initialisation runtime
        if not hasattr(state, "config_moteur"):
            state.config_moteur = {
                "MAX_BE_PER_FLIGHT": 20,
                "MAX_EQUIV_PER_VOLUNTEER": 20,
                "MAX_BENEV_PER_VOL": 0,
                "DUREE_MISSION_HEURES": 3,
                "MIN_HOURS_BETWEEN_FLIGHTS": 2,
            }

        cfg = state.config_moteur

        with st.form("form_param_moteur"):

            colA, colB = st.columns(2)
            with colA:
                cfg["MAX_BE_PER_FLIGHT"] = st.number_input(
                    "Nb max BE par vol",
                    min_value=1,
                    value=int(cfg["MAX_BE_PER_FLIGHT"]),
                )

            with colB:
                cfg["MAX_EQUIV_PER_VOLUNTEER"] = st.number_input(
                    "Equivalents max par bénévole",
                    min_value=1,
                    value=int(cfg["MAX_EQUIV_PER_VOLUNTEER"]),
                )

            colC, colD = st.columns(2)
            with colC:
                cfg["MAX_BENEV_PER_VOL"] = st.number_input(
                    "Nb max bénévoles par vol (0 = illimité)",
                    min_value=0,
                    value=int(cfg["MAX_BENEV_PER_VOL"]),
                )

            with colD:
                cfg["DUREE_MISSION_HEURES"] = st.number_input(
                    "Durée mission (h avant vol)",
                    min_value=1,
                    value=int(cfg["DUREE_MISSION_HEURES"]),
                )

            cfg["MIN_HOURS_BETWEEN_FLIGHTS"] = st.number_input(
                "Ecart min entre 2 missions (heures)",
                min_value=0,
                value=int(cfg["MIN_HOURS_BETWEEN_FLIGHTS"]),
            )

            if st.form_submit_button("💾 Appliquer"):
                state.config_moteur = cfg
                st.success("✔ Paramoteur mis à jour (SESSION ONLY).")


    # ------------------------------------------------------------
    # 2) PARAMDEST
    # ------------------------------------------------------------
    if st.session_state.active_block == "paramdest":

        st.subheader("🗂️ ParamDest (SESSION)")

        try:
            df = load_param_df(
                state,
                "df_param_dest",
                Path(state.tdb_tmp),
                SHEET_PARAM_DEST,
                column_map_param_dest,
                header=0
            )

            edited = st.data_editor(df, width="stretch", num_rows="dynamic")

            if st.button("💾 Valider ParamDest"):
                state.df_param_dest = edited
                st.success("✔ ParamDest mis à jour pour la session.")

        except Exception as e:
            st.error(f"❌ Erreur ParamDest : {e}")


    # ------------------------------------------------------------
    # 3) PARAMBE
    # ------------------------------------------------------------
    if st.session_state.active_block == "parambe":

        st.subheader("📦 ParamBE (SESSION)")

        try:
            df = load_param_df(
                state,
                "df_param_be",
                Path(state.tdb_tmp),
                SHEET_PARAM_BE,
                column_map_param_be,
                header=0
            )

            edited = st.data_editor(df, width="stretch", num_rows="dynamic")

            if st.button("💾 Valider ParamBE"):
                state.df_param_be = edited
                st.success("✔ ParamBE mis à jour pour la session.")

        except Exception as e:
            st.error(f"❌ Erreur ParamBE : {e}")


    # ------------------------------------------------------------
    # 4) PARAMBenev
    # ------------------------------------------------------------
    if st.session_state.active_block == "parambenev":

        st.subheader("👥 ParamBenev (SESSION)")

        try:
            df = load_param_df(
                state,
                "df_param_benev",
                Path(state.benev_tmp),
                SHEET_PARAM_BENEV,
                column_map_param_benev,
                header=0
            )

            edited = st.data_editor(df, width="stretch", num_rows="dynamic")

            if st.button("💾 Valider ParamBenev"):
                state.df_param_benev = edited
                st.success("✔ ParamBenev mis à jour pour la session.")

        except Exception as e:
            st.error(f"❌ Erreur ParamBenev : {e}")

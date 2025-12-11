# asf_app/ui/ui_inputs.py
# -*- coding: utf-8 -*-

import os
from datetime import datetime, date, timedelta
import pandas as pd
import streamlit as st
from pathlib import Path
import shutil

from asf_app.state import get_state, get_tmp_dir
from openpyxl import load_workbook

import scheduler.config_paths as cp
from scheduler.config_paths import (
    TABLEAU_DE_BORD,
    PLANNING_BENEVOLES,
    VOLS,
    SHEET_MAG_CENTRAL,
    SHEET_PARAM_BE,
    SHEET_PARAM_DEST,
    SHEET_PARAM_BENEV,
    SHEET_BENEV_DISPO,
    SHEET_VOLS,
    TABLEAU_DE_BORD_SRC,
    PLANNING_BENEVOLES_SRC,
    VOLS_SRC,
    prepare_paths,
    IS_STREAMLIT_CLOUD,
    CLOUD_MESSAGE,
)

# Loaders normalisés
from loaders.universal_loader import load_and_normalize
from loaders.load_shipments import load_shipments
from loaders.load_vols_api import load_vols_api, store_vols_api_sheet
from scheduler.column_map import (
    column_map_mag_central,
    column_map_param_be,
    column_map_param_dest,
    column_map_param_benev,
    column_map_benev_dispo,
    column_map_vols,
)

from scheduler import be_manager


# -------------------------------------------------------------------------
# HELPERS
# -------------------------------------------------------------------------

def pretty_mtime(path: Path) -> str:
    try:
        ts = path.stat().st_mtime
        dt = datetime.fromtimestamp(ts)
        return dt.strftime("%d/%m/%Y à %H:%M")
    except Exception:
        return "N/A"


def ensure_tmp_file(src_path: Path, filename: str) -> Path:
    """
    Assure la présence d’un fichier dans le TMP moteur.
    Copie si absent. Ne remplace pas si déjà là.
    """
    tmp_dir = get_tmp_dir()
    dst = tmp_dir / filename
    if not dst.exists() and src_path.exists():
        dst.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(src_path, dst)
    return dst


def pick_planning_dates(state):
    """
    Sélecteur de période de planning (début/fin) stocké dans le state.
    """
    today = date.today()
    next_monday = today + timedelta(days=(7 - today.weekday()))
    default_start = state.api_start_date or next_monday
    default_end = state.api_end_date or (default_start + timedelta(days=6))

    st.subheader("🗓️ Période du planning")
    col_s1, col_s2 = st.columns(2)
    with col_s1:
        start = st.date_input("Date de début", value=default_start, key="plan_start")
    with col_s2:
        # La date de fin suit automatiquement la date de début (+6)
        auto_end = start + timedelta(days=6)
        st.write(f"Date de fin (auto) : **{auto_end:%d/%m/%Y}**")
        end = auto_end

    if end < start:
        st.error("La date de fin doit être après la date de début.")
    state.api_start_date = start
    state.api_end_date = end


def benev_last_message(path: Path) -> str:
    """
    Lit D2 (date) et E2 (heure) dans la feuille 'Source' du planning bénévoles.
    Retourne une chaîne "DD/MM/YY à HHhMM" ou "N/A" si non disponible.
    """
    try:
        wb = load_workbook(path, data_only=True)
        if "Source" not in wb.sheetnames:
            return "N/A"
        ws = wb["Source"]
        d = ws["D2"].value
        h = ws["E2"].value
        def _fmt_date(x):
            if isinstance(x, datetime):
                return x.strftime("%d/%m/%y")
            if hasattr(x, "strftime"):
                return x.strftime("%d/%m/%y")
            return str(x) if x else ""
        def _fmt_time(x):
            if isinstance(x, datetime):
                return x.strftime("%Hh%M")
            if isinstance(x, str):
                # essayer HH:MM
                try:
                    return datetime.strptime(x.strip(), "%H:%M").strftime("%Hh%M")
                except Exception:
                    return x
            return str(x) if x else ""
        d_str = _fmt_date(d)
        h_str = _fmt_time(h)
        if d_str or h_str:
            return f"{d_str} à {h_str}".strip(" à ")
        return "N/A"
    except Exception:
        return "N/A"


# -------------------------------------------------------------------------
# LOADING FONCTIONS — écrivent dans state
# -------------------------------------------------------------------------

def load_tdb_file(state, force=False):
    if not force and state.df_be is not None and state.df_param_be is not None:
        return

    try:
        df_mag = load_and_normalize(state.tdb_tmp, SHEET_MAG_CENTRAL,
                                    column_map_mag_central, header=5)
        df_param_be = load_and_normalize(state.tdb_tmp, SHEET_PARAM_BE,
                                         column_map_param_be, header=0)
        df_param_dest = load_and_normalize(state.tdb_tmp, SHEET_PARAM_DEST,
                                           column_map_param_dest, header=0)

        state.df_be = df_mag
        state.df_param_be = df_param_be
        state.df_param_dest = df_param_dest

    except Exception as e:
        st.error(f"❌ Erreur chargement TABLEAU DE BORD : {e}")


def load_benev_file(state, force=False):
    if not force and state.df_benev is not None and state.df_param_benev is not None:
        return

    try:
        df_param_benev = load_and_normalize(state.benev_tmp, SHEET_PARAM_BENEV,
                                            column_map_param_benev, header=0)
        df_dispo = load_and_normalize(state.benev_tmp, SHEET_BENEV_DISPO,
                                      column_map_benev_dispo, header=0)

        state.df_param_benev = df_param_benev
        state.df_benev = df_dispo

    except Exception as e:
        st.error(f"❌ Erreur chargement Bénévoles : {e}")


def load_vols_file(state, force=False):
    if not force and state.df_vols is not None:
        return

    try:
        # Préférence : charger via loader complet (inclut onglets API + normalisation)
        from loaders.load_vols import load_vols_df
        state.df_vols = load_vols_df()
    except Exception:
        # Fallback minimal : onglet Vols normalisé
        try:
            df_vols = load_and_normalize(state.vols_tmp, SHEET_VOLS,
                                         column_map_vols, header=0)
            state.df_vols = df_vols
        except Exception as e:
            st.error(f"❌ Erreur chargement Vols : {e}")


# -------------------------------------------------------------------------
# TMP REWRITE / REFRESH
# -------------------------------------------------------------------------

def overwrite_tmp_file(uploaded_file, state, key_name, reload_func):
    """
    key_name ∈ {"tdb","benev","vols"}
    """
    if uploaded_file is None:
        return

    tmp_path = getattr(state, f"{key_name}_tmp")

    try:
        with open(tmp_path, "wb") as f:
            f.write(uploaded_file.read())

        # Invalidate dataframes
        if key_name == "tdb":
            state.df_be = None
            state.df_param_be = None
            state.df_param_dest = None
        elif key_name == "benev":
            state.df_benev = None
            state.df_param_benev = None
        elif key_name == "vols":
            state.df_vols = None

        reload_func(state, force=True)
        st.success("✔ Fichier mis à jour dans le dossier TMP")

    except Exception as e:
        st.error(f"❌ Erreur mise à jour TMP : {e}")


def refresh_from_onedrive(state, src_path, key_name, reload_func):
    """
    Copie depuis OneDrive vers TMP et recharge les df.
    """
    try:
        dst = ensure_tmp_file(src_path, Path(src_path).name)
        setattr(state, f"{key_name}_tmp", dst)

        # reset dfs
        if key_name == "tdb":
            state.df_be = None
            state.df_param_be = None
            state.df_param_dest = None
        elif key_name == "benev":
            state.df_benev = None
            state.df_param_benev = None
        elif key_name == "vols":
            state.df_vols = None

        reload_func(state, force=True)
        st.success(f"✔ Rechargé depuis OneDrive : {src_path.name}")

    except Exception as e:
        st.error(f"❌ Erreur refresh OneDrive : {e}")


def refresh_all(state):
    prepare_paths(copy_sources=True)
    state.tdb_tmp = TABLEAU_DE_BORD
    state.benev_tmp = PLANNING_BENEVOLES
    state.vols_tmp = VOLS
    state.df_be = state.df_param_be = state.df_param_dest = None
    state.df_benev = state.df_param_benev = None
    state.df_vols = None
    load_tdb_file(state, force=True)
    load_benev_file(state, force=True)
    load_vols_file(state, force=True)
    st.success("✔ Tous les fichiers ont été rechargés depuis OneDrive.")


# -------------------------------------------------------------------------
# UI PRINCIPALE
# -------------------------------------------------------------------------

def render_tab_inputs():
    st.header("📁 Fichiers d’entrée — OneDrive + TMP")
    state = get_state()

    cloud_mode = IS_STREAMLIT_CLOUD
    if cloud_mode:
        st.warning(CLOUD_MESSAGE)

    # Initialisation TMP si besoin (copies déjà préparées côté moteur)
    if state.tdb_tmp is None:
        state.tdb_tmp = TABLEAU_DE_BORD
    if state.benev_tmp is None:
        state.benev_tmp = PLANNING_BENEVOLES
    if state.vols_tmp is None:
        state.vols_tmp = VOLS

    # Chargements (évite d'afficher des erreurs sur Cloud si fichiers vides)
    if not cloud_mode or state.tdb_tmp.stat().st_size > 0:
        load_tdb_file(state)
    if not cloud_mode or state.benev_tmp.stat().st_size > 0:
        load_benev_file(state)
    if not cloud_mode or state.vols_tmp.stat().st_size > 0:
        load_vols_file(state)

    # ----- bouton refresh global ----
    if not cloud_mode and st.button("🔄 Recharger TOUS les fichiers depuis OneDrive"):
        refresh_all(state)

    # Sélecteur de période
    pick_planning_dates(state)

    col_tdb, col_benev, col_vols = st.columns(3)

    # ---------------------------------------------------------------------
    # TDB
    # ---------------------------------------------------------------------
    with col_tdb:
        st.subheader("📘 Tableau de bord")
        st.write(f"TMP : `{state.tdb_tmp.name}`")
        st.write(f"🕒 Modifié : {pretty_mtime(state.tdb_tmp)}")

        # BE planifiables
        try:
            param_be = be_manager.load_param_be()
            be_raw = load_shipments(param_be)
            be_filt = be_manager.filter_shipments(be_raw)
            be_sorted = be_manager.sort_shipments(be_filt)

            if be_sorted:
                counts = {}
                for s in be_sorted:
                    counts[s.dest] = counts.get(s.dest, 0) + 1
                st.write("📦 BE planifiables : " +
                         ", ".join([f"{d} ({c})" for d, c in counts.items()]))
            else:
                st.write("📦 Aucun BE planifiable.")
        except Exception as e:
            st.error(f"❌ Erreur BE : {e}")

        if (not cloud_mode) and st.button("🔄 Recharger TDB depuis OneDrive"):
            refresh_from_onedrive(state, TABLEAU_DE_BORD_SRC, "tdb", load_tdb_file)

        file = st.file_uploader("Importer TABLEAU_DE_BORD.xlsx", type=["xlsx"], key="up_tdb")
        if file:
            overwrite_tmp_file(file, state, "tdb", load_tdb_file)

    # ---------------------------------------------------------------------
    # BENEVOLES
    # ---------------------------------------------------------------------
    with col_benev:
        st.subheader("👥 Bénévoles")
        st.write(f"TMP : `{state.benev_tmp.name}`")
        st.write(f"🕒 Modifié : {pretty_mtime(state.benev_tmp)}")
        st.write(f"Dernier message traité : {benev_last_message(state.benev_tmp)}")

        if (not cloud_mode) and st.button("🔄 Recharger Bénévoles depuis OneDrive"):
            refresh_from_onedrive(state, PLANNING_BENEVOLES_SRC, "benev", load_benev_file)

        file = st.file_uploader("Importer Planning Bénévoles.xlsx", type=["xlsx"], key="up_benev")
        if file:
            overwrite_tmp_file(file, state, "benev", load_benev_file)

    # ---------------------------------------------------------------------
    # VOLS
    # ---------------------------------------------------------------------
    with col_vols:
        st.subheader("✈️ Vols")
        st.write(f"TMP : `{state.vols_tmp.name}`")
        st.write(f"🕒 Modifié : {pretty_mtime(state.vols_tmp)}")

        # Semaine détectée
        try:
            dfv = state.df_vols
            if dfv is not None and "Date_Vol" in dfv.columns:
                dates = pd.to_datetime(dfv["Date_Vol"], errors="coerce", dayfirst=True, format="%d/%m/%y")
                dates = dates.dropna()
                if not dates.empty:
                    dmin, dmax = dates.min(), dates.max()
                    st.write(f"🗓️ Du {dmin:%d/%m} au {dmax:%d/%m} "
                             f"(sem. {dmin.isocalendar()[1]})")
        except Exception as e:
            st.error(f"❌ Erreur lecture Vols : {e}")

        source_choice = st.radio(
            "Source vols",
            ["Fichier Excel", "API Air France (CDG ➜ ParamDest, AF)"],
            index=0 if state.vols_source != "api" else 1,
        )
        state.vols_source = "api" if source_choice.endswith("AF)") else "excel"

        if state.vols_source == "api":
            st.info(
                "Appel direct API Air France (origin CDG, operating AF) pour les destinations ParamDest. "
                "Limite 1 requête/s et 100/jour. Clé attendue dans `AF_API_KEY` (env ou secrets)."
            )
            cache_info = ""
            cache_path = get_tmp_dir() / "vols_api_cache.parquet"
            if cache_path.exists():
                mtime = datetime.fromtimestamp(cache_path.stat().st_mtime)
                cache_info = f"(cache du {mtime:%d/%m à %Hh%M})"
                st.write(f"Cache disponible {cache_info}")
                if st.button(f"Charger le dernier cache"):
                    try:
                        state.df_vols = pd.read_parquet(cache_path)
                        st.success(f"Vols chargés depuis cache {cache_info}")
                    except Exception as e:
                        st.error(f"❌ Erreur lecture cache : {e}")
            else:
                st.write("Aucun cache API disponible pour l'instant.")

            if state.api_start_date and state.api_end_date:
                if st.button("Appeler l'API Air France"):
                    try:
                        df_api = load_vols_api(state.api_start_date, state.api_end_date)
                        state.df_vols = df_api
                        if df_api is not None:
                            st.success(f"{len(df_api)} vols chargés via API (du {state.api_start_date:%d/%m} au {state.api_end_date:%d/%m}).")
                            try:
                                df_api.to_parquet(cache_path, index=False)
                            except Exception:
                                pass
                            # Sauvegarde dans Vols.xlsx (feuille API-SXX-YYYY en dernière position)
                            try:
                                sheet_name = store_vols_api_sheet(df_api, state.api_start_date)
                                st.info(f"Données API sauvegardées dans VOLS.xlsx (onglet {sheet_name}).")
                                # Copie dans la version TMP (VOLS) pour affichage immédiat
                                try:
                                    from loaders.load_vols_api import copy_api_sheet_to_tmp
                                    copy_api_sheet_to_tmp(sheet_name)
                                    from loaders.load_vols import load_vols_df
                                    state.df_vols = load_vols_df()
                                except Exception:
                                    pass
                            except Exception as e:
                                st.warning(f"Vols API chargés mais non sauvegardés dans Excel : {e}")
                        else:
                            st.warning("Aucun vol retourné par l'API.")
                    except Exception as e:
                        st.error(f"❌ Erreur API AF : {e}")
            else:
                st.warning("Sélectionne une période avant d'appeler l'API.")
        else:
            if (not cloud_mode) and st.button("🔄 Recharger Vols depuis OneDrive"):
                refresh_from_onedrive(state, VOLS_SRC, "vols", load_vols_file)

            file = st.file_uploader("Importer Vols.xlsx", type=["xlsx"], key="up_vols")
            if file:
                overwrite_tmp_file(file, state, "vols", load_vols_file)

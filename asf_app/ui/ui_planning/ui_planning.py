# asf_app/ui/ui_planning/ui_planning.py
# -*- coding: utf-8 -*-

import streamlit as st
import pandas as pd
from datetime import datetime, date, timedelta
import typing
from pathlib import Path
import platform
import subprocess
import shutil

# Excel
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from openpyxl.cell.cell import MergedCell
from openpyxl.worksheet.table import Table, TableStyleInfo

from reportlab.platypus import SimpleDocTemplate, Table as RLTable, TableStyle
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4

# Vues planning (export/communication)
from scheduler.planning_views import build_export_view
from scheduler.planning_schema import normalize_planning_df

# Scheduler
from scheduler.core_scheduler import Scheduler
import scheduler.config as engine_cfg
from utils.export_pdf import export_first_sheet_to_pdf

# Chemins
import scheduler.config_paths as cp
from scheduler.format_rules import format_vol_display
from utils.identifiers import normalize_be_number
from scheduler.config_paths import (
    TABLEAU_DE_BORD,
    VOLS,
    PLANNING_BENEVOLES,
    SHEET_PARAM_BENEV,
    SHEET_PARAM_DEST,
    SHEET_PARAM_EXP,
    SHEET_PARAM_BE,
    PLANNING_TEMPLATE,
    OUTPUT_PLANNING_DIR,
)

# Loaders
from loaders.load_shipments import get_shipments_df_cached
from loaders.load_vols import get_vols_df_cached
from loaders.load_benevoles import get_benevoles_cached
from loaders.universal_loader import load_and_normalize

# Column maps
from scheduler.column_map import (
    column_map_param_dest,
    column_map_param_expediteur,
    column_map_param_benev,
    column_map_param_be,
)

# State
from asf_app.ui.ui_planning.state_planning import get_planning_state
from asf_app.ui.ui_planning.formatting import build_preview


def show_mag_central_status():
    method = st.session_state.get("mag_central_write_method")
    if method == "excel":
        st.info("MAG CENTRAL mis à jour via Excel (validations préservées).")
    elif method == "openpyxl":
        st.warning("MAG CENTRAL mis à jour via openpyxl : validations de données possibles supprimées.")
    elif method == "no_updates":
        st.info("MAG CENTRAL : aucune cellule à mettre à jour.")
    elif method == "missing":
        st.warning("MAG CENTRAL non mis à jour : fichier introuvable.")
    elif method == "read_error":
        st.warning("MAG CENTRAL non mis à jour : erreur d’ouverture.")

# Communication DF
from asf_app.ui.ui_communication.clean_planning_df import build_df_comm

# ---------------------------------------------------------------------------
# MAQUETTE XLSX
# ---------------------------------------------------------------------------

ASFMM_TEMPLATE_DIR = cp.ASF_ONEDRIVE / "Planning MAB" / "ASFmm PLANNING 2025" / "aaSOURCE"
PLANNING_MAQUETTE = ASFMM_TEMPLATE_DIR / "Planning-maquette.xlsx"
# Fallback local si la maquette OneDrive est absente
if not PLANNING_MAQUETTE.exists():
    PLANNING_MAQUETTE = PLANNING_TEMPLATE


# ---------------------------------------------------------------------------
# PARAMETERS
# ---------------------------------------------------------------------------

def load_parameters():
    df_paramdest = load_and_normalize(
        path=TABLEAU_DE_BORD,
        sheet_name=SHEET_PARAM_DEST,
        mapping=column_map_param_dest
    )

    df_paramexp = load_and_normalize(
        path=TABLEAU_DE_BORD,
        sheet_name=SHEET_PARAM_EXP,
        mapping=column_map_param_expediteur
    )

    df_parambenev = load_and_normalize(
        path=PLANNING_BENEVOLES,
        sheet_name=SHEET_PARAM_BENEV,
        mapping=column_map_param_benev
    )

    df_parambe = load_and_normalize(
        path=TABLEAU_DE_BORD,
        sheet_name=SHEET_PARAM_BE,
        mapping=column_map_param_be
    )

    return df_paramdest, df_paramexp, df_parambenev, df_parambe


# ---------------------------------------------------------------------------
# WEEK DETECTION
# ---------------------------------------------------------------------------

def detect_week_year(df):
    if "Date_Vol" not in df.columns:
        return None, None
    dates = pd.to_datetime(df["Date_Vol"], errors="coerce").dropna()
    if dates.empty:
        return None, None
    dt = dates.min()
    return int(dt.isocalendar().week), int(dt.year)


# ---------------------------------------------------------------------------
# 7-DAY BLOCKS FOR EXCEL
# ---------------------------------------------------------------------------

def build_seven_day_blocks(df_planning):
    if df_planning.empty:
        return pd.DataFrame()

    week, year = detect_week_year(df_planning)
    if week is None:
        return df_planning

    monday = datetime.fromisocalendar(year, week, 1)
    blocks = []

    for i in range(7):
        day = monday + timedelta(days=i)
        sub = df_planning[
            pd.to_datetime(df_planning["DATE"], errors="coerce").dt.date == day.date()
        ]

        if sub.empty:
            blocks.append({
                "DATE": day.strftime("%Y-%m-%d"),
                "DATE_LONGUE": day.strftime("%A %d/%m/%Y"),
                "ITEM": day.strftime("%A"),
                "NOM": "",
                "OK": "",
                "DESTINATION": "",
                "IATA": "",
                "ROUTING": "",
                "NUMERO VOL": "",
                "HEURE VOL": "",
                "NUMERO BE": "",
                "NOMBRE COLIS": "",
                "TYPE": "",
                "LIVRAISON COLIS": "",
                "DATE TRANSFERT": "",
                "EXPEDITEUR": "",
                "DESTINATAIRE": "",
                "ID_BENEVOLE": "",
            })
        else:
            for _, r in sub.iterrows():
                blocks.append(r.to_dict())

    return pd.DataFrame(blocks)


# ---------------------------------------------------------------------------
# XLSX STYLING
# ---------------------------------------------------------------------------

def apply_excel_styles(wb):
    # Sélectionne une feuille si "Data" n'existe pas
    ws = wb["Data"] if "Data" in wb.sheetnames else wb.active

    font = Font(name="Aptos", size=12)
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    max_col = ws.max_column
    max_row = ws.max_row

    for row in ws.iter_rows(min_row=2, max_row=max_row, min_col=1, max_col=max_col):
        for cell in row:
            cell.font = font
            cell.border = border
            cell.alignment = align

    for col in ws.columns:
        col_letter = col[0].column_letter
        max_len = max((len(str(c.value)) for c in col if c.value), default=10)
        ws.column_dimensions[col_letter].width = max(10, min(max_len + 2, 40))


# ---------------------------------------------------------------------------
# EXPORT PDF VIA EXCEL (MAC/WINDOWS)
# ---------------------------------------------------------------------------

def _export_pdf_via_excel(filepath: Path, sheet_name: str = None) -> None:
    """
    Tente un export PDF fidèle en appelant Excel (macOS: AppleScript, Windows: PowerShell COM).
    Ne lève pas d'erreur bloquante : en cas d'échec, on ne génère pas de PDF.
    """
    try:
        sys_name = platform.system().lower()
        pdf_path = filepath.with_suffix(".pdf")
        if "darwin" in sys_name:
            # AppleScript : ouvre le classeur, copie la feuille cible dans un nouveau classeur, exporte en PDF
            ws_selector = 'worksheet 1 of wbSource'
            if sheet_name:
                ws_selector = f'worksheet "{sheet_name}" of wbSource'
            script = f'''
set theFile to POSIX file "{filepath}"
set pdfFile to POSIX file "{pdf_path}"
tell application "Microsoft Excel"
    activate
    set wbSource to open theFile
    try
        set wsTarget to {ws_selector}
        copy wsTarget to new workbook
        set wbTemp to active workbook
        save wbTemp in pdfFile as PDF file format
        close wbTemp saving no
    end try
    close wbSource saving no
end tell
'''
            res = subprocess.run(["osascript", "-e", script], capture_output=True, text=True)
            if res.returncode != 0 or not pdf_path.exists():
                raise RuntimeError(res.stderr or res.stdout)
        elif "windows" in sys_name:
            # PowerShell + COM Excel
            sheet_ps = "1"
            if sheet_name:
                sheet_ps = f'"{sheet_name}"'
            ps_cmd = r'''
$xl = New-Object -ComObject Excel.Application
$xl.Visible = $false
$wb = $xl.Workbooks.Open("{xlsx}")
try {
    $ws = $wb.Sheets.Item({sheet_sel})
    $ws.ExportAsFixedFormat(0, "{pdf}")
} catch {}
$wb.Close($false)
$xl.Quit()
'''.format(xlsx=str(filepath).replace("\\","\\\\"), pdf=str(pdf_path).replace("\\","\\\\"), sheet_sel=sheet_ps)
            subprocess.run(["powershell", "-NoProfile", "-Command", ps_cmd], check=False)
        # ne rien faire si Linux ou si échec
    except Exception:
        pass

def _export_pdf_via_soffice(filepath: Path, sheet_range: str = "1-1") -> Path | None:
    """
    Fallback LibreOffice headless : respecte la mise en forme Excel.
    Retourne le chemin du PDF si généré, sinon None. Ne lève pas d'erreur bloquante.
    """
    try:
        soffice = shutil.which("soffice")
        if not soffice:
            # chemins courants macOS
            candidates = [
                "/Applications/LibreOffice.app/Contents/MacOS/soffice",
                "/Applications/OpenOffice.app/Contents/MacOS/soffice",
            ]
            for c in candidates:
                if Path(c).exists():
                    soffice = c
                    break
        if not soffice:
            return None
        out_dir = filepath.parent
        result = subprocess.run(
            [
                soffice,
                "--headless",
                "--convert-to",
                "pdf:calc_pdf_Export",
                "--outdir",
                str(out_dir),
                str(filepath),
            ],
            check=False,
        )
        pdf_path = filepath.with_suffix(".pdf")
        return pdf_path if pdf_path.exists() else None
    except Exception:
        return None

# ---------------------------------------------------------------------------
# EXPORT XLSX
# ---------------------------------------------------------------------------

def export_excel_planning(
    df,
    week,
    year,
    df_vols=None,
    df_parambenev=None,
    df_dispos=None,
    df_paramdest=None,
    create_tables: bool = True,
    write_source_excel: bool = False,
):
    # Maquette : priorité OneDrive aaSOURCE/Planning-maquette.xlsx
    onedrive_maquette = (
        cp.ASF_ONEDRIVE
        / "Planning MAB"
        / "ASFmm PLANNING 2025"
        / "aaSOURCE"
        / "Planning-maquette.xlsx"
    )
    template = onedrive_maquette if onedrive_maquette.exists() else PLANNING_MAQUETTE
    if not template.exists():
        raise FileNotFoundError(f"❌ Maquette introuvable : {template}")

    # Nom et dossier de sortie par année (année mise à jour plus bas avec le lundi)
    filename = f"ASFmm - PLANNING SEMAINE N° {week:02d} - {year}.xlsx"
    planning_dir = (
        cp.OUTPUT_PLANNING_DIR
        if cp.is_graph_onedrive()
        else cp.ASF_ONEDRIVE / "Planning MAB" / f"ASFmm PLANNING {year}"
    )
    if not planning_dir.exists():
        planning_dir.mkdir(parents=True, exist_ok=True)
    out_path = planning_dir / filename

    shutil.copy2(template, out_path)
    try:
        wb = load_workbook(out_path)
    except Exception:
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Planning SXX"
        wb.save(out_path)
        wb = load_workbook(out_path)

    # Feuilles : priorité aux noms, fallback par position
    ws_plan = wb.worksheets[0]
    for sh in wb.sheetnames:
        if sh.lower().startswith("planning"):
            ws_plan = wb[sh]
            break
    ws_export = wb["Export planning"] if "Export planning" in wb.sheetnames else (wb.worksheets[1] if len(wb.worksheets) > 1 else wb.create_sheet("Export planning"))
    ws_vols = wb["Data Vols"] if "Data Vols" in wb.sheetnames else (wb.worksheets[2] if len(wb.worksheets) > 2 else wb.create_sheet("Data Vols"))
    ws_bene = wb["Data Benevoles"] if "Data Benevoles" in wb.sheetnames else (wb.worksheets[3] if len(wb.worksheets) > 3 else wb.create_sheet("Data Benevoles"))

    # Date lundi (utilisée pour l'année)
    monday = None
    try:
        monday = date.fromisocalendar(int(year), int(week), 1)
    except Exception:
        monday = None
    dfp = build_export_view(df, df_paramdest=df_paramdest, df_vols=df_vols).fillna("")

    def _clear_tables(ws):
        """Supprime tous les tableaux d'une feuille (compat dict/list) pour éviter les doublons corrompus."""
        try:
            tbls = ws._tables
            if isinstance(tbls, dict):
                tbls.clear()
            else:
                ws._tables = []
        except Exception:
            pass
    dfp["DATE"] = pd.to_datetime(dfp.get("Date_Vol", dfp.get("DATE", "")), errors="coerce")
    # Normaliser heures (remplacer h par :) avant conversion
    heures_raw = dfp.get("Heure_Vol", dfp.get("HEURE VOL", ""))
    dfp["HEURE_VOL_DT"] = (
        pd.to_datetime(
            heures_raw.astype(str).str.replace("h", ":", regex=False),
            errors="coerce",
        )
    )
    dfp["HEURE_VOL"] = dfp["HEURE_VOL_DT"]
    dfp["HEURE_MIN"] = (
        dfp["HEURE_VOL_DT"].dt.hour.fillna(99).astype(int) * 60
        + dfp["HEURE_VOL_DT"].dt.minute.fillna(59).astype(int)
    )
    dfp = dfp.sort_values(by=["DATE", "HEURE_MIN", "Destination", "Numero_Vol"], kind="mergesort")

    # Mapping prenom court NOM
    map_bene_to_display = {}
    if df_parambenev is not None and not getattr(df_parambenev, "empty", True):
        tmp = df_parambenev.copy()
        tmp["Benevole"] = tmp.get("Benevole", tmp.get("BENEVOLE", ""))
        tmp["Prenom_Court"] = tmp.get("Prenom_Court", tmp.get("Prenom court", ""))
        tmp["Nom"] = tmp.get("Nom", "")
        for _, r in tmp.iterrows():
            b = str(r.get("Benevole", "")).strip()
            pc = str(r.get("Prenom_Court", "")).strip()
            nom = str(r.get("Nom", "")).strip().upper()
            disp = f"{pc} {nom}".strip()
            map_bene_to_display[b] = disp

    def _bene_display(name):
        if name in map_bene_to_display:
            return map_bene_to_display[name]
        parts = str(name).strip().split()
        if len(parts) >= 2:
            return f"{parts[0][0].upper()}. {' '.join(parts[1:]).upper()}"
        return str(name).upper()

    dfp["BENEVOLE_DISP"] = dfp.get("Benevole", dfp.get("BENEVOLE", "")).apply(_bene_display)
    dfp["VILLE"] = dfp.get("Ville", dfp.get("Dest_Ville", dfp.get("Destination", ""))).astype(str).str.upper()
    dfp["IATA"] = dfp.get("IATA", dfp.get("Dest_IATA", dfp.get("Destination", ""))).astype(str).str.upper()
    # Recalage Routing depuis Data Vols (Date + Numero) avec fallback IATA
    if df_vols is not None and not getattr(df_vols, "empty", True):
        def _normalize_vol_key(value: object) -> str:
            s = str(value or "").strip().upper()
            if s.startswith("AF"):
                s = s.replace("AF", "").strip()
            digits = "".join(ch for ch in s if ch.isdigit())
            if digits:
                try:
                    return str(int(digits))
                except Exception:
                    return digits.lstrip("0") or digits
            return s

        vols_map = df_vols.copy()
        if "Date_Vol_dt" in vols_map.columns:
            vols_date = pd.to_datetime(vols_map["Date_Vol_dt"], errors="coerce")
        else:
            vols_date = pd.to_datetime(vols_map.get("Date_Vol", ""), errors="coerce", dayfirst=True)
        vols_map["_DATE_KEY"] = vols_date.dt.date
        vols_map["_VOL_KEY"] = vols_map.get("Numero_Vol", "").apply(_normalize_vol_key)
        routing_map = (
            vols_map.dropna(subset=["Routing"])
            .drop_duplicates(subset=["_DATE_KEY", "_VOL_KEY"])
            .set_index(["_DATE_KEY", "_VOL_KEY"])["Routing"]
            .to_dict()
        )
        dfp["_DATE_KEY"] = pd.to_datetime(dfp["DATE"], errors="coerce").dt.date
        dfp["_VOL_KEY"] = dfp.get("Numero_Vol", "").apply(_normalize_vol_key)
        routing_series = (
            dfp["Routing"]
            if "Routing" in dfp.columns
            else pd.Series([""] * len(dfp), index=dfp.index)
        )
        mask_routing_empty = (
            routing_series.fillna("")
            .astype(str)
            .str.strip()
            .str.lower()
            .replace({"nan": "", "none": ""})
            .eq("")
        )
        if mask_routing_empty.any():
            def _route_fallback(row):
                key = (row.get("_DATE_KEY"), row.get("_VOL_KEY"))
                if key in routing_map:
                    return routing_map[key]
                return ""

            dfp.loc[mask_routing_empty, "Routing"] = (
                dfp.loc[mask_routing_empty].apply(_route_fallback, axis=1)
            )
        dfp = dfp.drop(columns=["_DATE_KEY", "_VOL_KEY"], errors="ignore")

    dfp["VOL_AFF"] = dfp.get("Numero_Vol", dfp.get("Numero_Vol_Aff", "")).apply(format_vol_display)
    dfp["ROUTING"] = dfp.get("Routing", dfp.get("ROUTING", "")).astype(str).str.replace(",", "-").str.upper()
    dfp["HEURE_AFF"] = dfp["HEURE_VOL_DT"].dt.strftime("%Hh%M")
    dfp["BE_NUM"] = dfp.get("BE_Numero", dfp.get("NUMERO BE", dfp.get("BE_Num", ""))).apply(normalize_be_number)
    dfp["BE_COLIS"] = pd.to_numeric(dfp.get("BE_Nb_Colis", 0), errors="coerce").fillna(0).astype(int)
    dfp["BE_TYPE"] = dfp.get("BE_Type", "")
    dfp["BE_EXP"] = dfp.get("BE_Expediteur", "")
    dfp["BE_DEST"] = dfp.get("BE_Destinataire", "")
    dfp["BE_KEY"] = dfp["BE_NUM"]
    dfp["_STATUS"] = dfp.get("_STATUS", "normal")

    # ------------------------------------------------------------------
    # Maj MAG CENTRAL (col J/L) + récup de la date départ MAG utilisée
    # ------------------------------------------------------------------
    def _friday_previous_week(wk: int, yr: int):
        try:
            # vendredi de la semaine précédente : date(Semaine courante lundi) - 3 jours
            mon = date.fromisocalendar(yr, wk, 1)
            return mon - timedelta(days=3)
        except Exception:
            return None

    def update_mag_central_dates():
        nonlocal mag_write_method
        path = cp.TABLEAU_DE_BORD_SRC
        if not path.exists():
            mag_write_method = "missing"
            return {}
        try:
            from openpyxl import load_workbook
            wb_mag = load_workbook(path)
            ws_mag = wb_mag.active
            sheet_name = ws_mag.title
        except Exception:
            mag_write_method = "read_error"
            return {}

        # indexer MAG par clé BE (valeur brute + normalisée)
        mag_index = {}
        for row in ws_mag.iter_rows(min_row=1, max_row=ws_mag.max_row, min_col=1, max_col=20):
            val = row[0].value
            if val is None:
                continue
            sval = str(int(val)) if isinstance(val, (int, float)) else str(val).strip()
            key = normalize_be_number(sval)
            if not key:
                continue
            keys = {key, sval, key.lstrip("0")}
            for k in keys:
                mag_index[k] = row[0].row

        prev_friday = _friday_previous_week(week, year)
        used_dates = {}
        updates = {}

        def _safe_text(val: object) -> str:
            if val is None:
                return ""
            try:
                if pd.isna(val):
                    return ""
            except Exception:
                pass
            return str(val).strip()

        for _, r in dfp.iterrows():
            be_key = r.get("BE_KEY", "")
            if not be_key:
                continue
            row_idx = (
                mag_index.get(be_key)
                or mag_index.get(be_key.lstrip("0"))
                or mag_index.get(be_key[-3:] if len(be_key) > 3 else be_key)
            )
            if not row_idx:
                continue
            # Depart MAG col J (10) : si vide, écrire vendredi précédent
            if prev_friday:
                dm_cell = ws_mag.cell(row=row_idx, column=cp.MAG_CENTRAL_COL_DEPART_MAG)
                if dm_cell.value in (None, ""):
                    updates.setdefault(row_idx, {})[cp.MAG_CENTRAL_COL_DEPART_MAG] = prev_friday
                    used_dates[be_key] = prev_friday
                else:
                    used_dates[be_key] = dm_cell.value
            # Depart VOL col L (12) : date du vol du planning
            date_vol = r.get("DATE")
            if isinstance(date_vol, (pd.Timestamp, date)):
                dv_value = date_vol.date() if isinstance(date_vol, pd.Timestamp) else date_vol
                updates.setdefault(row_idx, {})[cp.MAG_CENTRAL_COL_DEPART_VOL] = dv_value

            # Colonnes W/X/Y/Z : ID bénév, Bénévole, Num Vol, Heure Vol
            bene_id = _safe_text(r.get("ID", ""))
            if bene_id.endswith(".0"):
                bene_id = bene_id[:-2]
            bene_disp = _safe_text(r.get("BENEVOLE_DISP", r.get("Benevole", "")))
            vol_aff = _safe_text(r.get("VOL_AFF", r.get("Numero_Vol", "")))
            heure_aff = _safe_text(r.get("HEURE_AFF", ""))
            if bene_id or bene_disp or vol_aff or heure_aff:
                row_updates = updates.setdefault(row_idx, {})
                row_updates[cp.MAG_CENTRAL_COL_ID_BENEV] = bene_id
                row_updates[cp.MAG_CENTRAL_COL_BENEV] = bene_disp
                row_updates[cp.MAG_CENTRAL_COL_VOL] = vol_aff
                row_updates[cp.MAG_CENTRAL_COL_HEURE] = heure_aff

        update_items = []
        for row_idx, cols in updates.items():
            for col_idx, val in cols.items():
                update_items.append((row_idx, col_idx, val))

        if not update_items:
            mag_write_method = "no_updates"
            return used_dates

        try:
            from utils.excel_automation import update_excel_cells
            if update_excel_cells(path, sheet_name, update_items):
                mag_write_method = "excel"
                cp.sync_local_file_to_onedrive(path)
                return used_dates
        except Exception:
            pass

        mag_write_method = "openpyxl"
        for row_idx, col_idx, val in update_items:
            ws_mag.cell(row=row_idx, column=col_idx).value = val

        try:
            wb_mag.save(path)
            cp.sync_local_file_to_onedrive(path)
        except Exception:
            pass
        return used_dates

    # Ecriture optionnelle vers MAG CENTRAL source
    mag_write_method = "disabled"
    map_depart_mag = update_mag_central_dates() if write_source_excel else {}
    if write_source_excel:
        st.session_state["mag_central_write_method"] = mag_write_method
    else:
        st.session_state.pop("mag_central_write_method", None)

    # Date départ MAG issue de MAG central ou fallback vendredi précédent
    def _depart_mag_for(be_key):
        if be_key in map_depart_mag:
            return map_depart_mag[be_key]
        prev_friday = None
        try:
            from datetime import timedelta
            mon = date.fromisocalendar(year, week, 1)
            prev_friday = mon - timedelta(days=3)
        except Exception:
            prev_friday = None
        return prev_friday
    dfp["DEPART_MAG"] = dfp["BE_KEY"].apply(_depart_mag_for)

    day_blocks = {
        0: (4, 32),
        1: (35, 63),
        2: (66, 94),
        3: (97, 125),
        4: (128, 156),
        5: (159, 187),
        6: (190, 218),
    }
    keep_rows = {3,4,5,6,31,32,33,34,35,36,37,62,63,64,65,66,67,68,93,94,95,96,97,98,99,124,125,126,127,128,129,130,155,156,157,158,159,160,161,186,187,188,189,190,191,192,217,218,219}

    # Efface contenu planning (colonnes 4..17)
    for row in ws_plan.iter_rows(min_row=3, max_row=ws_plan.max_row, min_col=4, max_col=17):
        for cell in row:
            # Ne pas toucher K219/L219 ni les cellules fusionnées
            if isinstance(cell, MergedCell):
                continue
            if cell.coordinate in ("K219", "L219"):
                continue
            cell.value = None
        ws_plan.row_dimensions[row[0].row].hidden = False

    for day_idx in range(7):
        start, end = day_blocks.get(day_idx, (None, None))
        if start is None:
            continue
        current_row = start
        df_day = dfp[dfp["DATE"].dt.dayofweek == day_idx]
        if df_day.empty:
            for r in range(start, end + 1):
                if r not in keep_rows:
                    ws_plan.row_dimensions[r].hidden = True
            continue
        df_day = df_day.sort_values(by=["DATE", "HEURE_MIN", "VOL_AFF", "Destination"], kind="mergesort")
        for (dt, hmin, vol), df_vol in df_day.groupby(["DATE", "HEURE_MIN", "VOL_AFF"], sort=False):
            df_vol = df_vol.reset_index(drop=True)
            if current_row > end:
                break
            bene_list = list(dict.fromkeys(df_vol["BENEVOLE_DISP"].tolist()))
            for idx, r in df_vol.iterrows():
                if current_row > end:
                    break
                is_first = idx == 0
                status = str(r.get("_STATUS", "normal")).lower()
                # Bénévole : une seule liste par vol, on ne répète pas au-delà du nb de bénévoles
                bene_val = bene_list[idx] if idx < len(bene_list) else ""
                ws_plan.cell(row=current_row, column=4).value = bene_val  # D
                if is_first:
                    ws_plan.cell(row=current_row, column=6).value = r["VILLE"]      # F
                    ws_plan.cell(row=current_row, column=7).value = r["IATA"]       # G
                    ws_plan.cell(row=current_row, column=8).value = r["ROUTING"]    # H
                    ws_plan.cell(row=current_row, column=9).value = r["VOL_AFF"]    # I
                    ws_plan.cell(row=current_row, column=10).value = r["HEURE_AFF"] # J
                be_num_val = r["BE_NUM"]
                ws_plan.cell(row=current_row, column=11).value = be_num_val        # K
                be_colis_val = "" if status.startswith("old") else r["BE_COLIS"]
                ws_plan.cell(row=current_row, column=12).value = be_colis_val      # L
                if status.startswith("old"):
                    ws_plan.cell(row=current_row, column=12).value = ""  # vider col L pour suppression
                ws_plan.cell(row=current_row, column=13).value = r["BE_TYPE"]       # M
                # Col O : Date départ MAG (dd/mm/yy)
                dep_mag = r.get("DEPART_MAG")
                if isinstance(dep_mag, pd.Timestamp):
                    dep_mag = dep_mag.date()
                if isinstance(dep_mag, date):
                    dep_mag_str = dep_mag.strftime("%d/%m/%y")
                else:
                    dep_mag_str = ""
                ws_plan.cell(row=current_row, column=15).value = dep_mag_str        # O
                ws_plan.cell(row=current_row, column=16).value = r["BE_EXP"]        # P
                ws_plan.cell(row=current_row, column=17).value = r["BE_DEST"]       # Q
                # Coloration si modif/supp/ajout (D à Q)
                fill_color = None
                if status.startswith("old"):
                    fill_color = PatternFill("solid", fgColor="F8CCCC")  # rouge clair
                elif status == "new":
                    fill_color = PatternFill("solid", fgColor="CFE2FF")  # bleu clair
                if fill_color:
                    for col_idx in range(4, 18):
                        ws_plan.cell(row=current_row, column=col_idx).fill = fill_color
                current_row += 1
            current_row += 2  # 2 lignes vides entre vols
        for r in range(current_row, end + 1):
            if r not in keep_rows:
                ws_plan.row_dimensions[r].hidden = True

    for col_letter in ["B", "C", "G"]:
        ws_plan.column_dimensions[col_letter].hidden = True
    # Conserver les largeurs source pour A-O, autosize sur P et Q uniquement
    for col_letter in ["P", "Q"]:
        col = ws_plan[col_letter]
        max_len = max((len(str(c.value)) for c in col if c.value), default=10)
        ws_plan.column_dimensions[col_letter].width = max(10, min(max_len + 2, 40))

    from openpyxl.utils.dataframe import dataframe_to_rows
    _clear_tables(ws_export)
    ws_export.delete_rows(1, ws_export.max_row)
    for row in dataframe_to_rows(dfp, index=False, header=True):
        ws_export.append(row)

    _clear_tables(ws_vols)
    if df_vols is not None and not getattr(df_vols, "empty", True):
        ws_vols.delete_rows(1, ws_vols.max_row)
        for row in dataframe_to_rows(df_vols, index=False, header=True):
            ws_vols.append(row)
        # Appliquer un tableau (filtres/tri)
        max_row = ws_vols.max_row
        max_col = ws_vols.max_column
        if create_tables and max_row >= 2 and max_col >= 1:
            _clear_tables(ws_vols)
            from openpyxl.utils import get_column_letter
            ref = f"A1:{get_column_letter(max_col)}{max_row}"
            tab = Table(displayName="Table_Vols", ref=ref)
            style = TableStyleInfo(
                name="TableStyleMedium9",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False,
            )
            tab.tableStyleInfo = style
            ws_vols.add_table(tab)

    # Export disponibilités bénévoles → feuille Data Benevoles
    df_dispo_export = df_dispos
    if df_dispo_export is None:
        try:
            df_dispo_export = get_benevoles_cached()
        except Exception:
            df_dispo_export = None

    _clear_tables(ws_bene)
    if df_dispo_export is not None and not getattr(df_dispo_export, "empty", True):
        df_dispo_export = df_dispo_export.copy()
        # garder les colonnes utiles (headers Disponibilités)
        cols_order = [
            "ID",
            "Benevole",
            "Nom",
            "Prenom",
            "Prenom_Court",
            "Date",
            "Heure_Arrivee",
            "Heure_Depart",
        ]
        cols_present = [c for c in cols_order if c in df_dispo_export.columns]
        df_dispo_export = df_dispo_export[cols_present]
        ws_bene.delete_rows(1, ws_bene.max_row)
        for row in dataframe_to_rows(df_dispo_export, index=False, header=True):
            ws_bene.append(row)
        max_row_b = ws_bene.max_row
        max_col_b = ws_bene.max_column
        if create_tables and max_row_b >= 2 and max_col_b >= 1:
            _clear_tables(ws_bene)
            from openpyxl.utils import get_column_letter
            ref_b = f"A1:{get_column_letter(max_col_b)}{max_row_b}"
            tab_b = Table(displayName="Table_Benevoles", ref=ref_b)
            style_b = TableStyleInfo(
                name="TableStyleMedium9",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False,
            )
            tab_b.tableStyleInfo = style_b
            ws_bene.add_table(tab_b)

    # Repositionner les cellules de colonne A (avec formules) sur une ligne visible centrée
    def move_to_middle(src_row, start, end, col_letter="A"):
        src = ws_plan[f"{col_letter}{src_row}"]
        if src.value is None:
            return
        def is_visible(r):
            rd = ws_plan.row_dimensions[r]
            return not bool(rd.hidden)
        visible_rows = [r for r in range(start, end + 1) if is_visible(r)]
        if not visible_rows:
            return
        dest_row = visible_rows[len(visible_rows)//2]
        if dest_row == src_row:
            return
        dst = ws_plan[f"{col_letter}{dest_row}"]
        dst.value = src.value
        dst._style = src._style
        src.value = None

    move_to_middle(17, 4, 32)
    move_to_middle(48, 35, 63)
    move_to_middle(79, 66, 94)
    move_to_middle(110, 97, 125)
    move_to_middle(141, 128, 156)
    move_to_middle(172, 159, 187)
    move_to_middle(203, 190, 218)

    wb.save(out_path)
    # Validation simple : recharger le fichier pour s'assurer qu'il n'est pas corrompu
    try:
        load_workbook(out_path)
    except Exception as e:
        raise RuntimeError(f"Export Excel invalide : {e}")
    # Renommer la première feuille
    try:
        ws_plan.title = f"Planning S{week:02d}-{year}"
    except Exception:
        pass

    # Déterminer semaine/année depuis A1 si possible
    def _week_year_from_a1():
        val = ws_plan["A1"].value
        wk, yr = week, year
        try:
            if isinstance(val, datetime):
                wk = val.isocalendar()[1]
                yr = val.isocalendar()[0]
            else:
                dt = pd.to_datetime(val, errors="coerce", dayfirst=True)
                if pd.notna(dt):
                    wk = dt.isocalendar()[1]
                    yr = dt.isocalendar()[0]
        except Exception:
            pass
        return int(wk), int(yr)

    week_final, year_final = _week_year_from_a1()

    # Construire nom final avec version Q1
    try:
        version_cell = ws_plan["Q1"].value
    except Exception:
        version_cell = ""
    version_str = str(version_cell).strip()
    try:
        if version_str.isdigit():
            version_str = f"{int(version_str):02d}"
    except Exception:
        pass
    if not version_str:
        version_str = "01"

    base_name = f"ASFmm - PLANNING SEMAINE N° {week_final:02d} - {year_final} - v{version_str}"
    planning_dir_final = (
        cp.OUTPUT_PLANNING_DIR
        if cp.is_graph_onedrive()
        else cp.ASF_ONEDRIVE / "Planning MAB" / f"ASFmm PLANNING {year_final}"
    )
    planning_dir_final.mkdir(parents=True, exist_ok=True)
    target_path = planning_dir_final / f"{base_name}.xlsx"
    counter = 2
    while target_path.exists():
        target_path = planning_dir_final / f"{base_name}-{counter}.xlsx"
        counter += 1

    # Sauvegarder dans le chemin final (déplacement si besoin)
    wb.save(out_path)
    if cp.is_graph_onedrive():
        remote_path = cp.get_output_remote_path(year_final, out_path.name)
        cp.sync_local_file_to_onedrive(out_path, remote_path=remote_path, conflict_behavior="rename")
    if out_path != target_path:
        try:
            shutil.move(out_path, target_path)
        except Exception:
            # fallback copie
            wb.save(target_path)
        out_path = target_path

    wb.save(out_path)

    # Export PDF (1ère feuille) via AppleScript helper
    if not cp.is_graph_onedrive():
        pdf_target = out_path.with_suffix(".pdf")
        try:
            pdf_path = export_first_sheet_to_pdf(out_path, pdf_target)
            if not pdf_path.exists():
                raise RuntimeError("PDF non généré.")
        except Exception:
            try:
                st.warning("PDF non généré : Excel non accessible pour l’export automatique.")
            except Exception:
                pass

    return out_path


# ---------------------------------------------------------------------------
# STREAMLIT UI – PLANNING TAB
# ---------------------------------------------------------------------------

def render_tab_planning():

    st.title("📊 Génération du Planning")

    planning_state = get_planning_state()

    st.subheader("📥 Chargement des données")

    with st.spinner("Chargement Paramètres…"):
        df_paramdest, df_paramexp, df_parambenev, df_parambe = load_parameters()

    with st.spinner("Chargement BE…"):
        _ = get_shipments_df_cached()

    with st.spinner("Chargement vols…"):
        df_vols = get_vols_df_cached()

    with st.spinner("Chargement disponibilités bénévoles…"):
        df_dispos = get_benevoles_cached()

    st.success("✔ Fichiers chargés")

    # -------------------------------------------------------------------
    # Génération planning
    # -------------------------------------------------------------------
    st.subheader("🚀 Génération du planning")

    col_gen1, col_gen2 = st.columns([1, 1])

    with col_gen1:
        if st.button("Lancer la génération du planning", type="primary"):

            with st.spinner("Exécution du moteur…"):
                scheduler = Scheduler(mode="real")
                df_planning, df_bilan = scheduler.run()

            if df_planning is None or df_planning.empty:
                st.error("❌ Aucun planning généré.")
                return

            df_planning = normalize_planning_df(df_planning)
            planning_state.set_planning(df_planning, df_bilan)
            st.success(f"✔ Planning généré ({len(df_planning)} lignes)")

    with col_gen2:
        if st.button("Rapport modes de planning", type="secondary"):
            st.info("Comparaison en cours : mode standard vs capacité anticipée…")
            old_flag = getattr(engine_cfg, "USE_REAL_CAPACITY_ESTIMATE", False)
            try:
                # Run standard
                engine_cfg.USE_REAL_CAPACITY_ESTIMATE = False
                sched_std = Scheduler(mode="real")
                df_plan_std, df_bilan_std = sched_std.run()

                # Run anticipée
                engine_cfg.USE_REAL_CAPACITY_ESTIMATE = True
                sched_real = Scheduler(mode="real")
                df_plan_real, df_bilan_real = sched_real.run()

                st.success("Comparaison terminée")
                c_std, c_real = st.columns(2)
                with c_std:
                    st.caption("Mode standard")
                    st.dataframe(
                        build_preview(df_plan_std, df_paramdest),
                        height=380,
                        width="stretch",
                        hide_index=True,
                    )
                with c_real:
                    st.caption("Mode capacité anticipée")
                    st.dataframe(
                        build_preview(df_plan_real, df_paramdest),
                        height=380,
                        width="stretch",
                        hide_index=True,
                    )
            finally:
                engine_cfg.USE_REAL_CAPACITY_ESTIMATE = old_flag

    # -------------------------------------------------------------------
    # Charger le planning depuis state
    # -------------------------------------------------------------------
    df_planning = normalize_planning_df(planning_state.planning)

    if df_planning is None or df_planning.empty:
        st.info("Aucun planning généré pour le moment.")
        return

    # -------------------------------------------------------------------
    # BILAN (expéditions / vols / bénévoles)
    # -------------------------------------------------------------------
    st.subheader("📊 Bilan")
    col_be, col_vols, col_benev = st.columns(3)

    # Bilan expéditions (basé sur df_planning)
    with col_be:
        st.caption("Bilan Expéditions")
        if df_planning.empty:
            st.info("Aucune expédition chargée.")
        else:
            from loaders.load_shipments import load_shipments_df

            df_plan = df_planning.copy()
            be_col_planning = (
                "BE_Numero"
                if "BE_Numero" in df_plan.columns
                else df_plan.columns[df_plan.columns.str.contains("BE", case=False)][0]
            )
            df_plan["_MANUEL"] = df_plan.get("_MANUEL", False)
            if isinstance(df_plan["_MANUEL"], bool):
                df_plan["_MANUEL"] = pd.Series([df_plan["_MANUEL"]] * len(df_plan))
            df_plan["_MANUEL"] = df_plan["_MANUEL"].fillna(False)
            df_plan["BE_KEY"] = df_plan[be_col_planning].apply(normalize_be_number)

            df_ship = load_shipments_df().copy()
            df_ship["BE_KEY"] = df_ship["BE_Numero"].apply(normalize_be_number)
            df_ship = df_ship.drop_duplicates(subset=["BE_KEY"])

            planned_set = set(df_plan["BE_KEY"].tolist())
            manual_set = set(df_plan.loc[df_plan["_MANUEL"], "BE_KEY"])

            df_be_view = df_ship[["Destination", "BE_Numero", "BE_Nb_Colis", "BE_KEY"]].copy()
            df_be_view["BE_Numero"] = df_be_view["BE_Numero"].apply(normalize_be_number)
            df_be_view["BE_Nb_Colis"] = (
                pd.to_numeric(df_be_view.get("BE_Nb_Colis", 0), errors="coerce").fillna(0).astype(int)
            )
            df_be_view["Planification"] = df_be_view["BE_KEY"].apply(
                lambda k: "OK planning" if k in planned_set else "Non planifié"
            )
            df_be_view = df_be_view.sort_values(["Destination", "BE_Numero"]).reset_index(drop=True)
            manual_mask = df_be_view["BE_KEY"].isin(manual_set)

            def _style_be(row):
                color = "background-color: #f2f2f2" if manual_mask.iloc[row.name] else ""
                return [color] * len(row)

            st.dataframe(
                df_be_view[["Destination", "BE_Numero", "BE_Nb_Colis", "Planification"]].style.apply(
                    _style_be, axis=1
                ),
                height=300,
                width="stretch",
                hide_index=True,
            )

    # Bilan vols
    with col_vols:
        st.caption("Bilan Vols")
        df_vols_bilan = df_planning.copy()
        df_vols_bilan["BE_Nb_Colis"] = pd.to_numeric(df_vols_bilan.get("BE_Nb_Colis", 0), errors="coerce").fillna(0).astype(int)
        df_vols_bilan = (
            df_vols_bilan.groupby(["Destination", "Date_Vol", "Numero_Vol"], dropna=False)["BE_Nb_Colis"]
            .sum()
            .reset_index()
            .rename(columns={"BE_Nb_Colis": "Nb_Colis"})
            .sort_values(["Destination", "Date_Vol"])
        )
        df_vols_view = df_vols_bilan.assign(
            Numero_Vol=lambda d: d["Numero_Vol"].apply(format_vol_display)
        )[["Destination", "Date_Vol", "Numero_Vol", "Nb_Colis"]].reset_index(drop=True)
        df_plan_manual = df_planning.copy()
        if "_MANUEL" in df_plan_manual.columns:
            df_plan_manual["_MANUEL"] = df_plan_manual["_MANUEL"].fillna(False)
        else:
            df_plan_manual["_MANUEL"] = False
        manual_keys = set(
            df_plan_manual[df_plan_manual["_MANUEL"]][["Destination", "Date_Vol", "Numero_Vol"]]
            .apply(tuple, axis=1)
            .tolist()
        )
        manual_mask_vols = df_vols_view.apply(
            lambda r: (r["Destination"], r["Date_Vol"], r["Numero_Vol"]) in manual_keys,
            axis=1,
        )
        def _style_vol(row):
            color = "background-color: #f2f2f2" if manual_mask_vols.iloc[row.name] else ""
            return [color]*len(row)
        st.dataframe(
            df_vols_view.style.apply(_style_vol, axis=1),
            height=300,
            width="stretch",
            hide_index=True,
        )

    # Bilan bénévoles
    with col_benev:
        st.caption("Bilan Bénévoles")
        df_benev_disp = df_dispos.copy()
        df_benev_disp["Date"] = pd.to_datetime(df_benev_disp.get("Date"), errors="coerce")

        def _slot(row):
            arrivee_raw = str(row.get("Heure_Arrivee", "")).strip()
            depart_raw = str(row.get("Heure_Depart", "")).strip()
            if not arrivee_raw and not depart_raw:
                return ""
            try:
                arr_dt = pd.to_datetime(arrivee_raw)
                dep_dt = pd.to_datetime(depart_raw)
                arr_display = (arr_dt + pd.Timedelta(hours=3)).strftime("%Hh%M")
                dep_display = dep_dt.strftime("%Hh%M")
            except Exception:
                arr_display = arrivee_raw
                dep_display = depart_raw
            day = row["Date"].day_name() if pd.notna(row["Date"]) else ""
            jours = {
                "Monday": "Lundi",
                "Tuesday": "Mardi",
                "Wednesday": "Mercredi",
                "Thursday": "Jeudi",
                "Friday": "Vendredi",
                "Saturday": "Samedi",
                "Sunday": "Dimanche",
            }
            day_fr = jours.get(day, day)
            return f"{day_fr} {arr_display} à {dep_display}".strip()

        df_benev_disp["Slot"] = df_benev_disp.apply(_slot, axis=1)
        df_benev_disp = df_benev_disp[df_benev_disp["Slot"].str.strip() != ""]

        vols_par_benev = (
            df_planning.groupby("Benevole", dropna=False)
            .size()
            .reset_index(name="Nb_Vols_Affectes")
        )
        df_benev_bilan = (
            df_benev_disp.groupby("Benevole")
            .agg(
                Disponibilites=("Slot", lambda s: " | ".join(sorted(set(s), key=lambda x: ["LUNDI","MARDI","MERCREDI","JEUDI","VENDREDI","SAMEDI","DIMANCHE"].index(x.split()[0].upper()) if x else 0))),
                Nb_Dispo=("Slot", "count"),
                Prenom_Court=("Prenom_Court", "first"),
                Nom=("Nom", "first"),
            )
            .reset_index()
        )
        df_benev_bilan = df_benev_bilan.merge(vols_par_benev, on="Benevole", how="left").fillna({"Nb_Vols_Affectes": 0})
        df_benev_bilan["Benevole_Aff"] = df_benev_bilan.apply(
            lambda r: f"{str(r['Prenom_Court']).strip()} {str(r['Nom']).strip().upper()}".strip(), axis=1
        )
        df_benev_bilan = df_benev_bilan.sort_values("Benevole_Aff")

        df_benev_view = df_benev_bilan[["Benevole_Aff", "Nb_Vols_Affectes", "Nb_Dispo", "Disponibilites"]].rename(columns={"Nb_Vols_Affectes": "Nb_Vols", "Benevole_Aff": "Benevole"}).reset_index(drop=True)
        # Nb_Vols en entier
        df_benev_view["Nb_Vols"] = pd.to_numeric(df_benev_view["Nb_Vols"], errors="coerce").fillna(0).astype(int)
        df_plan_manual = df_planning.copy()
        if "_MANUEL" in df_plan_manual.columns:
            df_plan_manual["_MANUEL"] = df_plan_manual["_MANUEL"].fillna(False)
        else:
            df_plan_manual["_MANUEL"] = False
        manual_bene = set(df_plan_manual[df_plan_manual["_MANUEL"]]["Benevole"].astype(str).str.strip())
        manual_mask_bene = df_benev_view["Benevole"].astype(str).str.strip().isin(manual_bene)
        def _style_bene(row):
            color = "background-color: #f2f2f2" if manual_mask_bene.iloc[row.name] else ""
            return [color]*len(row)
        st.dataframe(
            df_benev_view.style.apply(_style_bene, axis=1),
            height=300,
            width="stretch",
            hide_index=True,
        )

    # -------------------------------------------------------------------
    # Modification manuelle d’un BE (statut D)
    # -------------------------------------------------------------------
    st.subheader("🛠️ Modifier le planning")
    # (Style des boutons géré globalement dans app.py)

    # BE en statut D (inclut ceux déjà planifiés)
    df_be_mod = get_shipments_df_cached()
    if df_be_mod.empty:
        st.info("Aucun BE en statut D disponible.")
    else:
        def _norm_full(val):
            return normalize_be_number(val)

        def _norm_short(val):
            full = normalize_be_number(val)
            return full[-3:] if full else ""

        # BE présents dans le planning (toutes variantes de colonne)
        be_col_planning = None
        for cand in ["BE_Numero", "BE NUMERO", "BE_NUMERO", "BE_Num", "BE_numero"]:
            if cand in df_planning.columns:
                be_col_planning = cand
                break
        if be_col_planning is None:
            be_col_planning = "BE_Numero"
            df_planning[be_col_planning] = df_planning.get(be_col_planning, "")
        planning_be_raw = {str(x).strip() for x in df_planning[be_col_planning].tolist()}
        planning_be_full = {_norm_full(x) for x in df_planning[be_col_planning].tolist()}
        planning_be_short = {_norm_short(x) for x in df_planning[be_col_planning].tolist()}

        df_be_mod["BE_Numero_Str"] = df_be_mod["BE_Numero"].apply(_norm_full)
        # tri par destination
        df_be_mod = df_be_mod.sort_values(by=["Destination", "BE_Numero_Str"])
        be_lookup = df_be_mod.set_index("BE_Numero_Str")

        def _format_be(num_str: str) -> str:
            if num_str in be_lookup.index:
                r = be_lookup.loc[num_str]
                dest_iata = r.get("Destination", "")
                nb = pd.to_numeric(r.get("BE_Nb_Colis", 0), errors="coerce")
                nb_int = int(nb) if pd.notna(nb) else r.get("BE_Nb_Colis", "")
                in_planning = (
                    _norm_full(num_str) in planning_be_full
                    or _norm_short(num_str) in planning_be_short
                    or str(num_str).strip() in planning_be_raw
                )
                planned_label = "déjà au planning" if in_planning else "non planifié"
                return f"{dest_iata} | BE {num_str} — {nb_int} colis — {r.get('BE_Type','')} ({planned_label})"
            return str(num_str)

        # Préparer vols (date+vol+heure) pour affichage long
        def _fmt_time(val):
            t = pd.to_datetime(str(val), errors="coerce")
            if pd.isna(t):
                return str(val)
            return t.strftime("%Hh%M")

        def _fmt_vol(val):
            return format_vol_display(val) or str(val)

        def _fmt_date_long(val):
            try:
                d = pd.to_datetime(val, dayfirst=True)
                if pd.isna(d):
                    return str(val)
            except Exception:
                return str(val)
            jours = {
                "Monday": "Lundi",
                "Tuesday": "Mardi",
                "Wednesday": "Mercredi",
                "Thursday": "Jeudi",
                "Friday": "Vendredi",
                "Saturday": "Samedi",
                "Sunday": "Dimanche",
            }
            return f"{jours.get(d.day_name(), d.strftime('%A'))} {d.strftime('%d/%m/%y')}"

        # Disponibilités bénévoles : dictionnaire {Benevole -> liste (date, arrivée, départ)}
        df_dispo = df_dispos.copy()
        df_dispo["Date"] = pd.to_datetime(df_dispo.get("Date"), errors="coerce").dt.date
        def _time_only(val):
            t = pd.to_datetime(str(val), errors="coerce")
            return t.time() if pd.notna(t) else None
        df_dispo["Arr"] = df_dispo.get("Heure_Arrivee", "").apply(_time_only)
        df_dispo["Dep"] = df_dispo.get("Heure_Depart", "").apply(_time_only)

        # Sélection BE + Vol + Bénévole (3 blocs sur 1 ligne)
        # Colonnes : BE (un peu plus large que les 2 boutons) / Vol / Bénévole
        col1, col2, col3 = st.columns([1.5, 1.25, 1.25])

        with col1:
            selected_be = st.selectbox(
                "Sélectionner un BE",
                be_lookup.index.tolist(),
                format_func=_format_be,
            )

        if selected_be:
            be_info = be_lookup.loc[selected_be]
            # Colonne BE dans le planning (tolérance aux variantes)
            be_col_planning = None
            for cand in ["BE_Numero", "BE NUMERO", "BE_NUMERO", "BE_Num", "BE_numero"]:
                if cand in df_planning.columns:
                    be_col_planning = cand
                    break
            if be_col_planning is None:
                be_col_planning = "BE_Numero"
                df_planning[be_col_planning] = df_planning.get(be_col_planning, "")
            be_planning_norm = df_planning[be_col_planning].apply(_norm_full)
            mask_planned = (
                be_planning_norm == _norm_full(selected_be)
            ) | (
                df_planning[be_col_planning].apply(_norm_short) == _norm_short(selected_be)
            )
            planned_rows = df_planning[mask_planned]
            planned = not planned_rows.empty

            def _default(col, fallback=""):
                if planned and col in planned_rows.columns:
                    val = planned_rows.iloc[0].get(col, fallback)
                    if pd.notna(val):
                        return val
                return be_info.get(col, fallback)

            default_date = _default("Date_Vol", "")
            default_dest = _default("Destination", be_info.get("Destination", ""))
            default_vol = _default("Numero_Vol", "")
            default_heure = _default("Heure_Vol", "")
            default_bene = _default("Benevole", "")

            # Vols filtrés par destination du BE (IATA contenu dans routing)
            code_iata_be = str(be_info.get("Destination", default_dest)).strip().upper()
            # si Destination est une ville, essayer mapping ParamDest -> IATA
            try:
                map_ville_to_iata = (
                    df_paramdest.dropna(subset=["Dest_Ville"])
                    .assign(Dest_Ville_UP=lambda d: d["Dest_Ville"].astype(str).str.upper().str.strip())
                    .drop_duplicates(subset=["Dest_Ville_UP"])
                    .set_index("Dest_Ville_UP")["Dest_IATA"]
                    .astype(str)
                    .str.upper()
                    .to_dict()
                )
            except Exception:
                map_ville_to_iata = {}
            if len(code_iata_be) != 3:
                code_iata_be = map_ville_to_iata.get(code_iata_be.upper(), code_iata_be)
            default_dest = code_iata_be

            # df_vols peut être une liste (fallback) : sécuriser en DataFrame
            try:
                vol_source = df_vols.copy()
            except Exception:
                vol_source = pd.DataFrame(df_vols)
            if isinstance(vol_source, list):
                vol_source = pd.DataFrame(vol_source)

            def _normalize_vol_df(vdf):
                vdf = vdf.copy()
                if "Routing" not in vdf.columns:
                    vdf["Routing"] = ""
                if "Dest_IATA" not in vdf.columns:
                    vdf["Dest_IATA"] = vdf.get("Destination", "")
                if "Date_Vol" not in vdf.columns:
                    vdf["Date_Vol"] = vdf.get("Date", "")
                if "Numero_Vol" not in vdf.columns:
                    vdf["Numero_Vol"] = vdf.get("Vol", vdf.get("Numero_Vol", ""))
                if "Heure_Vol" not in vdf.columns:
                    vdf["Heure_Vol"] = vdf.get("Heure", vdf.get("Heure_Vol", ""))
                vdf["Routing_Str"] = vdf["Routing"].astype(str)
                vdf["Dest_IATA_UP"] = vdf["Dest_IATA"].astype(str).str.upper()
                return vdf

            vol_source = _normalize_vol_df(vol_source) if not vol_source.empty else pd.DataFrame(columns=["Date_Vol","Numero_Vol","Heure_Vol","Routing_Str","Dest_IATA_UP"])

            # Filtrage par destination IATA ; pas de fallback global pour éviter des vols hors destination
            # Filtrage par destination, sinon fallback à tous les vols dispo (pour afficher tous les vols de la destination)
            vol_filtered = pd.DataFrame(columns=vol_source.columns)
            if not vol_source.empty and code_iata_be:
                mask_dest = vol_source["Dest_IATA_UP"].str.contains(code_iata_be, na=False) | vol_source["Routing_Str"].str.upper().str.contains(code_iata_be, na=False)
                vol_filtered = vol_source[mask_dest]
                if vol_filtered.empty:
                    vol_filtered = vol_source

            # Fallback: si toujours vide, on tente depuis le planning existant (même destination)
            if vol_filtered.empty and not df_planning.empty:
                df_plan_vols = df_planning.copy()
                df_plan_vols["Destination_UP"] = df_plan_vols.get("Destination", "").astype(str).str.upper()
                df_plan_vols = df_plan_vols[
                    df_plan_vols["Destination_UP"].str.contains(code_iata_be, na=False)
                ]
                if not df_plan_vols.empty:
                    df_plan_vols = df_plan_vols.rename(columns={"Vol": "Numero_Vol"})
                    df_plan_vols["Routing_Str"] = ""
                    vol_filtered = _normalize_vol_df(df_plan_vols)

            vols_unique = (
                vol_filtered[["Date_Vol", "Numero_Vol", "Heure_Vol", "Routing_Str"]]
                .dropna(how="all")
                .drop_duplicates()
                .sort_values(by=["Date_Vol", "Heure_Vol"])
            ) if not vol_filtered.empty else pd.DataFrame(columns=["Date_Vol","Numero_Vol","Heure_Vol","Routing_Str"])

            # Fallback supplémentaire : vols présents dans le planning pour cette destination
            if vols_unique.empty and not df_planning.empty:
                df_plan_vols = df_planning.copy()
                df_plan_vols["Destination_UP"] = df_plan_vols.get("Destination", "").astype(str).str.upper()
                mask_plan = df_plan_vols["Destination_UP"].str.contains(code_iata_be, na=False)
                df_plan_vols = df_plan_vols[mask_plan]
                if not df_plan_vols.empty:
                    df_plan_vols = df_plan_vols.rename(columns={"Vol": "Numero_Vol"})
                    df_plan_vols["Routing_Str"] = df_plan_vols.get("Routing", "")
                    vols_unique = df_plan_vols[["Date_Vol", "Numero_Vol", "Heure_Vol", "Routing_Str"]].drop_duplicates()

            # Statut vol (déjà au planning ?) pour enrichir le libellé
            def _norm_volnum_cmp(v):
                try:
                    return str(int(float(v)))
                except Exception:
                    return str(v).strip()

            vol_options = []
            for _, r in vols_unique.iterrows():
                vol_num_raw = r.get("Numero_Vol","")
                date_raw = r.get("Date_Vol","")
                heure_raw = r.get("Heure_Vol","")

                # Si pas de date et pas de numéro, on ignore cette ligne
                if (str(vol_num_raw).strip() == "") and (str(date_raw).strip() == ""):
                    continue

                # statut planning
                dest_match = df_planning.get("Destination", pd.Series(dtype=str)).astype(str).str.upper().str.contains(code_iata_be, na=False)
                already_planned = (
                    dest_match
                    & (df_planning.get("Date_Vol", pd.Series(dtype=str)).astype(str) == str(date_raw))
                    & (df_planning.get("Numero_Vol", pd.Series(dtype=str)).apply(_norm_volnum_cmp) == _norm_volnum_cmp(vol_num_raw))
                ).any()
                statut = "déjà au planning" if already_planned else "disponible"

                date_disp = _fmt_date_long(date_raw) or "(date ?)"
                vol_disp = _fmt_vol(vol_num_raw) or "(vol ?)"
                heure_disp = _fmt_time(heure_raw) or ""

                label = f"{date_disp} — {vol_disp} — {heure_disp} — {statut}"
                value = (str(r.get("Date_Vol","")), str(r.get("Numero_Vol","")), str(r.get("Heure_Vol","")))
                vol_options.append((label, value))

            # Fallback final : si aucune option exploitable, proposer un placeholder
            if not vol_options:
                # Essayer vols du planning (toutes destinations) pour donner au moins un slot
                fallback_plan = pd.DataFrame()
                if not df_planning.empty:
                    fallback_plan = df_planning.rename(columns={"Vol": "Numero_Vol"}).copy()
                    fallback_plan["Routing_Str"] = ""
                    fallback_plan["Date_Vol"] = fallback_plan.get("Date_Vol", "")
                    fallback_plan["Heure_Vol"] = fallback_plan.get("Heure_Vol", "")
                if not fallback_plan.empty:
                    vols_unique = fallback_plan[["Date_Vol", "Numero_Vol", "Heure_Vol", "Routing_Str"]].drop_duplicates()
                    for _, r in vols_unique.iterrows():
                        vol_num_raw = r.get("Numero_Vol","")
                        date_raw = r.get("Date_Vol","")
                        heure_raw = r.get("Heure_Vol","")
                        if (str(vol_num_raw).strip() == "") and (str(date_raw).strip() == ""):
                            continue
                        label = f"{_fmt_date_long(date_raw)} — {_fmt_vol(vol_num_raw)} — {_fmt_time(heure_raw)} — disponible"
                        value = (str(date_raw), str(vol_num_raw), str(heure_raw))
                        vol_options.append((label, value))
                if not vol_options:
                    vol_options = [(f"Aucun vol disponible pour {code_iata_be or 'la destination'}", ("", "", ""))]

            with col2:
                # Sélection vol
                vol_labels = [v[0] for v in vol_options]
                vol_values = [v[1] for v in vol_options]
                default_vol_tuple = (str(default_date), str(default_vol), str(default_heure))
                if default_vol_tuple in vol_values:
                    default_idx = vol_values.index(default_vol_tuple)
                else:
                    default_idx = 0
                vol_choice = st.selectbox(
                    "Sélectionner un vol",
                    vol_labels,
                    index=default_idx if vol_labels else 0,
                )
                chosen_tuple = vol_values[vol_labels.index(vol_choice)] if vol_labels else ("","","")
                date_choice, vol_choice_val, heure_choice = chosen_tuple

            with col3:
                # Disponibilité bénévoles
                def _bene_status(name, date_str, heure_str):
                    try:
                        d = pd.to_datetime(date_str).date()
                        h = pd.to_datetime(heure_str).time()
                    except Exception:
                        return "indisponible"
                    rows = df_dispo[df_dispo["Benevole"] == name]
                    rows_same_day = rows[rows["Date"] == d]
                    # déjà affecté même créneau ?
                    already = (
                        (df_planning["Benevole"].astype(str) == str(name))
                        & (df_planning["Date_Vol"].astype(str) == str(date_choice))
                        & (df_planning["Numero_Vol"].astype(str) == str(vol_choice_val))
                    ).any()
                    if already:
                        return "déjà affecté sur ce créneau"

                    if rows_same_day.empty:
                        return "indisponible"

                    # si heure arrivée/depart vides pour ce jour → inconnu
                    has_info = False
                    ok_dispo = False
                    for _, r in rows_same_day.iterrows():
                        arr = r["Arr"]
                        dep = r["Dep"]
                        if arr is None and dep is None:
                            continue
                        has_info = True
                        arr = arr or h
                        dep = dep or h
                        if arr <= h <= dep:
                            ok_dispo = True
                            break

                    if not has_info:
                        return "inconnu"
                    return "disponible" if ok_dispo else "indisponible"

                bene_options = []
                for name in sorted(df_parambenev["Benevole"].dropna().unique()):
                    status = _bene_status(name, date_choice, heure_choice)
                    bene_options.append(f"{name} ({status})")
                default_bene_label = None
                if default_bene:
                    status_def = _bene_status(default_bene, date_choice, heure_choice)
                    default_bene_label = f"{default_bene} ({status_def})"
                bene_idx = 0
                if default_bene_label in bene_options:
                    bene_idx = bene_options.index(default_bene_label)
                bene_choice_label = st.selectbox(
                    "Bénévole affecté",
                    bene_options,
                    index=bene_idx if bene_options else 0,
                )
                bene_choice = bene_choice_label.split(" (")[0]

            # Boutons action (mise à jour / suppression si déjà planifié) sous le sélecteur BE
            with col1:
                # Boutons alignés sur une ligne sous le sélecteur BE
                btn_a, btn_b = st.columns([2, 1.4])
                with btn_a:
                    if st.button("Mettre à jour le planning", type="primary", key="btn_update_planning"):
                        df_new = df_planning.copy()
                        # Informations bénévole (ID, téléphone) depuis ParamBenev si dispo
                        bene_phone = ""
                        bene_id = ""
                        try:
                            row_b = df_parambenev[df_parambenev["Benevole"] == bene_choice]
                            if not row_b.empty:
                                bene_phone = row_b.get("Telephone", pd.Series([""])).iloc[0]
                                bene_id = row_b.get("ID", pd.Series([""])).iloc[0]
                        except Exception:
                            pass

                        # Numéro BE : essayer de retrouver un format long existant dans le planning
                        be_value_raw = be_info.get("BE_Numero", selected_be)
                        be_digits = _digits(be_value_raw)
                        # Chercher les candidats dans le planning dont le suffixe correspond
                        candidates = [n for n in planning_be_raw if str(n).endswith(be_digits)]
                        if candidates:
                            # prendre le plus long pour garder le préfixe (ex: 250714)
                            be_value_str = sorted(candidates, key=lambda x: len(str(x)), reverse=True)[0]
                        else:
                            be_value_str = str(be_value_raw)
                        if be_value_str.endswith(".0"):
                            be_value_str = be_value_str[:-2]

                        if mask_planned.any():
                            df_new.loc[mask_planned, "BE_Numero"] = be_value_str
                            df_new.loc[mask_planned, "Date_Vol"] = date_choice
                            df_new.loc[mask_planned, "Destination"] = default_dest
                            df_new.loc[mask_planned, "Numero_Vol"] = vol_choice_val
                            df_new.loc[mask_planned, "Heure_Vol"] = heure_choice
                            df_new.loc[mask_planned, "Benevole"] = bene_choice
                            df_new.loc[mask_planned, "ID"] = str(bene_id).replace(".0", "").strip()
                            df_new.loc[mask_planned, "Telephone"] = bene_phone
                            df_new.loc[mask_planned, "BE_Nb_Colis"] = be_info.get("BE_Nb_Colis", "")
                            df_new.loc[mask_planned, "BE_Nb_Equiv"] = be_info.get("BE_Nb_Equiv", be_info.get("Equiv_Colis", ""))
                            df_new.loc[mask_planned, "BE_Type"] = be_info.get("BE_Type", "")
                            df_new.loc[mask_planned, "BE_Expediteur"] = be_info.get("BE_Expediteur", "")
                            df_new.loc[mask_planned, "BE_Destinataire"] = be_info.get("BE_Destinataire", "")
                            df_new.loc[mask_planned, "_MANUEL"] = True
                        else:
                            new_row = {col: "" for col in df_new.columns}
                            id_bene = str(bene_id).replace(".0", "").strip()
                            be_num_fmt = be_info.get("BE_Numero_Str", be_value_str)
                            new_row.update({
                                "BE_Numero": be_num_fmt,
                                "Date_Vol": date_choice,
                                "Destination": default_dest,
                                "Numero_Vol": vol_choice_val,
                                "Heure_Vol": heure_choice,
                                "BE_Nb_Colis": be_info.get("BE_Nb_Colis", ""),
                                "BE_Nb_Equiv": be_info.get("BE_Nb_Equiv", be_info.get("Equiv_Colis", "")),
                                "BE_Type": be_info.get("BE_Type", ""),
                                "Benevole": bene_choice,
                                "ID": id_bene,
                                "BE_Expediteur": be_info.get("BE_Expediteur", ""),
                                "BE_Destinataire": be_info.get("BE_Destinataire", ""),
                                "Telephone": bene_phone,
                                "_MANUEL": True,
                            })
                            df_new = pd.concat([df_new, pd.DataFrame([new_row])], ignore_index=True)

                        # Trier par Date_Vol / Heure_Vol pour l'aperçu
                        try:
                            df_new["Date_Vol"] = pd.to_datetime(df_new["Date_Vol"], errors="coerce")
                            df_new["Heure_Vol"] = pd.to_datetime(df_new["Heure_Vol"], errors="coerce").dt.strftime("%H:%M")
                            df_new = df_new.sort_values(by=["Date_Vol", "Heure_Vol"], kind="mergesort")
                            df_new["Date_Vol"] = df_new["Date_Vol"].dt.date
                        except Exception:
                            pass

                        df_new = normalize_planning_df(df_new)
                        planning_state.set_planning(df_new, planning_state.bilan)
                        st.success("Planning mis à jour. L’aperçu ci-dessous est recalculé.")
                        df_planning = df_new

                with btn_b:
                    # Bouton suppression (fond rouge clair via CSS plus bas)
                    if st.button("Supprimer du planning", disabled=not planned, type="secondary"):
                        if planned:
                            df_new = df_planning[~mask_planned].copy()
                            df_new = normalize_planning_df(df_new)
                            planning_state.set_planning(df_new, planning_state.bilan)
                            st.success("BE supprimé du planning.")
                            df_planning = df_new

    st.subheader("📋 Aperçu du planning brut (moteur)")
    st.dataframe(
        build_preview(df_planning, df_paramdest),
        height=420,
        width="stretch",
        hide_index=True,
    )

    # -------------------------------------------------------------------
    # Export Excel
    # -------------------------------------------------------------------
    st.subheader("📤 Validation & Export Excel")

    if st.button("📊 Valider & Exporter le planning Excel", type="primary"):

        week, year = detect_week_year(df_planning)
        if week is None:
            st.error("Impossible de détecter la semaine.")
            return

        # FULL enrichissement
        out_path = export_excel_planning(
            df_planning,
            week,
            year,
            df_vols=df_vols,
            df_parambenev=df_parambenev,
            df_dispos=df_dispos,
            df_paramdest=df_paramdest,
        )
        planning_state.set_last_export_path(out_path)

        st.success(f"✔ Planning exporté : {out_path.name}")
        show_mag_central_status()

        st.download_button(
            "⬇ Télécharger le fichier",
            data=open(out_path, "rb").read(),
            file_name=out_path.name,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

        if st.button("📁 Ouvrir le dossier"):
            try:
                if platform.system() == "Darwin":
                    subprocess.Popen(["open", str(out_path.parent)])
                elif platform.system() == "Windows":
                    subprocess.Popen(["explorer", str(out_path.parent)])
                else:
                    subprocess.Popen(["xdg-open", str(out_path.parent)])
            except Exception as e:
                st.error(f"❌ Impossible d’ouvrir : {e}")

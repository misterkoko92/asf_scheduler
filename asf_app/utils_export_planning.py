# asf_app/utils_export_planning.py
# -*- coding: utf-8 -*-

import shutil
from pathlib import Path
from datetime import datetime, timedelta

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment

from scheduler.config_paths import (
    PLANNING_TEMPLATE,
    PLANNING_BENEVOLES
)

# ============================================================
#   FORMAT BÉNÉVOLE — Prénom Court + NOM
# ============================================================
def benevole_format(benevole, benev_df):
    if not isinstance(benevole, str):
        return ""
    row = benev_df.loc[
        benev_df["BENEVOLE"].str.strip().str.upper() ==
        benevole.strip().upper()
    ]
    if row.empty:
        return benevole.upper()

    prenom_court = str(row.iloc[0].get("PRENOM_COURT", "")).strip()
    nom = str(row.iloc[0].get("NOM", "")).strip().upper()

    if prenom_court and nom:
        return f"{prenom_court} {nom}"
    return benevole.upper()


# ============================================================
#   MASQUAGE DES JOURS SANS VOL / AVEC VOL
# ============================================================
def _mask_day_blocks(ws):
    """
    Masque les lignes inutiles, mais NE MASQUE PAS :
    - les lignes structurelles
    - les lignes contenant un BE
    - les lignes marquées "__EMPTY__" en colonne B (protection)
    """

    DAY1_START = 3
    BLOCK = 21
    DAYS = 7

    for i in range(DAYS):

        start = DAY1_START + i * BLOCK
        end = start + BLOCK - 1

        structural = {
            start,
            start + 1,
            start + 8,
            start + 9,
            start + 10,
            start + 19,
            start + 20,
        }

        # BE présent ?
        has_be = any(
            ws[f"K{r}"].value not in (None, "", " ")
            for r in range(start, end + 1)
        )

        if not has_be:
            # JOUR SANS VOL
            for r in range(start, end + 1):
                ws.row_dimensions[r].hidden = (r not in structural)
            continue

        # JOUR AVEC VOL
        for r in range(start, end + 1):

            # lignes structurelles visibles
            if r in structural:
                ws.row_dimensions[r].hidden = False
                continue

            # lignes vides protégées (insertions pour séparation vols)
            if ws[f"B{r}"].value == "__EMPTY__":
                ws.row_dimensions[r].hidden = False
                continue

            v = ws[f"K{r}"].value
            ws.row_dimensions[r].hidden = (v in (None, "", " "))


# ============================================================
#   EXPORT PDF VIA EXCEL / APPLESCRIPT (macOS)
# ============================================================
def _export_pdf_with_excel(xlsx_path: Path) -> Path | None:
    """
    Tente de générer un PDF du classeur Excel via AppleScript et Microsoft Excel (macOS).

    - PDF généré dans le même dossier, même nom, extension .pdf
    - En cas d'échec, affiche un warning en console et ne bloque pas le reste.
    """
    pdf_path = xlsx_path.with_suffix(".pdf")

    # Sécuriser les chemins (gérer les guillemets)
    xlsx_str = str(xlsx_path).replace('"', '\\"')
    pdf_str = str(pdf_path).replace('"', '\\"')

    # AppleScript approximatif pour Excel Mac
    script = f'''
    tell application "Microsoft Excel"
        activate
        set wb to open POSIX file "{xlsx_str}"
        tell wb
            save workbook as filename "{pdf_str}" file format PDF file format
        end tell
        close wb saving no
    end tell
    '''

    try:
        import subprocess
        subprocess.run(
            ["osascript", "-e", script],
            check=True
        )
        return pdf_path
    except Exception as e:
        print(f"⚠ Impossible de générer le PDF via Excel/AppleScript : {e}")
        return None


# ============================================================
#   EXPORT FINAL VIA MAQUETTE ASF
# ============================================================
def export_planning_from_template(df, week_num, year):

    # ---------------------------------------------------------
    # 1) Copier maquette
    # ---------------------------------------------------------
    filename = f"ASFmm - PLANNING SEMAINE N° {week_num:02d} - {year}.xlsx"
    export_path = PLANNING_TEMPLATE.parent / filename
    shutil.copyfile(PLANNING_TEMPLATE, export_path)

    benev_df = pd.read_excel(PLANNING_BENEVOLES, sheet_name="ParamBenev")

    # enrichissement routing
    try:
        from asf_app.utils_planning_xlsx.routing import enrich_with_routing
        df = enrich_with_routing(df)
    except:
        pass

    wb = load_workbook(export_path)
    ws = wb["Planning SXX"]
    ws.title = f"Planning S{week_num:02d}"

    # ---------------------------------------------------------
    # 2) A1 = prochain lundi
    # ---------------------------------------------------------
    today = datetime.today()
    days_until_monday = (7 - today.weekday()) % 7
    if days_until_monday == 0:
        days_until_monday = 7
    next_monday = today + timedelta(days=days_until_monday)
    ws["A1"].value = next_monday
    ws["A1"].number_format = "DD/MM/YY"

    friday_prev = next_monday - timedelta(days=3)

    # ---------------------------------------------------------
    # 3) Tri par vraie heure
    # ---------------------------------------------------------
    def to_time(x):
        try:
            return datetime.strptime(str(x).strip(), "%H:%M").time()
        except:
            return datetime.strptime("00:00", "%H:%M").time()

    df_sorted = df.copy()
    df_sorted["HeureSort"] = df_sorted["Heure_Vol"].apply(to_time)

    df_sorted = df_sorted.sort_values(
        ["Date_Vol", "HeureSort", "Destination", "Vol", "BE_Numero"],
        kind="stable"
    ).reset_index(drop=True)

    # ---------------------------------------------------------
    # 4) Insertion BE
    # ---------------------------------------------------------
    current_row = 4
    last_key = None

    for idx, r in df_sorted.iterrows():

        key = (r["Date_Vol"], r["Destination"], r["Vol"], r["HeureSort"])

        # changement de vol → ajouter 2 lignes vides PROTÉGÉES
        if last_key is not None and key != last_key:

            for i in range(2):
                ws.row_dimensions[current_row].hidden = False
                ws[f"K{current_row}"].value = ""              # visuellement vide
                ws[f"B{current_row}"].value = "__EMPTY__"     # protection invisible
                ws[f"D{current_row}"].value = ""
                current_row += 1

        last_key = key
        row = current_row

        # valeurs
        bene = benevole_format(r["Benevole"], benev_df)
        be_full = r.get("BE_Numero", "")
        dest_nom = str(r.get("Destination_Nom", "")).upper()
        iata = str(r.get("Destination", "")).upper()
        routing = r.get("ROUTING", "")
        vol = r.get("Vol", "")
        heure = r.get("Heure_Vol", "")
        nb_colis = r.get("BE_Nb_Colis", "")
        type_colis = r.get("BE_Type", "")
        exp = r.get("BE_Expediteur", "")
        dest = r.get("BE_Destinataire", "")

        # première ligne d'un vol ?
        show = True
        if idx > 0:
            p = df_sorted.iloc[idx - 1]
            show = not (
                p["Date_Vol"] == r["Date_Vol"] and
                p["Destination"] == r["Destination"] and
                p["Vol"] == r["Vol"] and
                p["HeureSort"] == r["HeureSort"]
            )

        # écriture
        ws[f"D{row}"].value = bene

        if show:
            ws[f"F{row}"].value = dest_nom
            ws[f"G{row}"].value = iata
            ws[f"H{row}"].value = routing
            ws[f"I{row}"].value = vol
            ws[f"J{row}"].value = heure
        else:
            ws[f"F{row}"].value = ""
            ws[f"G{row}"].value = ""
            ws[f"H{row}"].value = ""
            ws[f"I{row}"].value = ""
            ws[f"J{row}"].value = ""

        ws[f"K{row}"].value = be_full
        ws[f"L{row}"].value = nb_colis
        ws[f"M{row}"].value = type_colis
        ws[f"O{row}"].value = friday_prev
        ws[f"O{row}"].number_format = "DD/MM/YYYY"
        ws[f"P{row}"].value = exp
        ws[f"Q{row}"].value = dest

        for col in ["A","D","F","G","H","I","J","K","L","M","O","P","Q"]:
            ws[f"{col}{row}"].alignment = Alignment(horizontal="center", vertical="center")

        current_row += 1

    # ---------------------------------------------------------
    # 5) Masquage final
    # ---------------------------------------------------------
    _mask_day_blocks(ws)

    # Colonnes à masquer
    ws.column_dimensions["B"].hidden = True
    ws.column_dimensions["C"].hidden = True
    ws.column_dimensions["G"].hidden = True

    # ---------------------------------------------------------
    # 6) Mise en forme finale : largeurs colonnes dynamiques P / Q
    # ---------------------------------------------------------
    for col in ["P", "Q"]:
        max_len = 0
        for row in range(1, ws.max_row + 1):
            v = ws[f"{col}{row}"].value
            if v not in (None, "", " "):
                max_len = max(max_len, len(str(v)))
        ws.column_dimensions[col].width = max_len + 3

    # ---------------------------------------------------------
    # 7) Sauvegarde XLSX
    # ---------------------------------------------------------
    wb.save(export_path)

    # ---------------------------------------------------------
    # 8) Export PDF (best effort)
    # ---------------------------------------------------------
    _export_pdf_with_excel(export_path)

    # ---------------------------------------------------------
    # 9) Ouverture du classeur Excel
    # ---------------------------------------------------------
    try:
        import subprocess
        subprocess.run(["open", "-a", "Microsoft Excel", str(export_path)])
    except Exception:
        pass

    return export_path

# -*- coding: utf-8 -*-
from __future__ import annotations

import re
from pathlib import Path
from datetime import date, timedelta
from typing import Optional

import pandas as pd

import scheduler.config_paths as cp
from loaders.universal_loader import load_and_normalize
from scheduler.column_map import column_map_mag_central
from scheduler.config_paths import SHEET_MAG_CENTRAL, TABLEAU_DE_BORD
from asf_app.config.runtime import get_tableau_de_bord_src
from utils.datetime_utils import (
    parse_date_series,
    parse_time_series,
    coerce_datetime,
    normalize_hour_str,
    hour_min_from_series,
)
from utils.identifiers import normalize_be_number
from utils.logging_utils import get_logger

logger = get_logger("shipments_update_service", console=False)


def load_be_status(status_code: str, *, tdb_path: Path | None = None) -> pd.DataFrame:
    tdb_use = Path(tdb_path) if tdb_path is not None else TABLEAU_DE_BORD
    try:
        xls = pd.ExcelFile(tdb_use)
        sheets = [
            name
            for name in xls.sheet_names
            if str(name).strip().upper().startswith("MAG CENTRAL")
        ]
    except Exception:
        sheets = []

    if not sheets:
        sheets = [SHEET_MAG_CENTRAL]

    def _rank(name: str) -> tuple[int, str]:
        match = re.search(r"(20\d{2})", name)
        year = int(match.group(1)) if match else -1
        return (year, name)

    sheets = [name for name in sheets if _rank(name)[0] >= 2025]
    sheets = sorted(sheets, key=_rank)
    frames = []
    for sheet in sheets:
        df_sheet = load_and_normalize(
            path=tdb_use,
            sheet_name=sheet,
            mapping=column_map_mag_central,
            header=5,
        )
        if df_sheet is None or df_sheet.empty:
            continue
        df_sheet["_MAG_CENTRAL_SHEET"] = sheet
        frames.append(df_sheet)

    df = pd.concat(frames, ignore_index=True) if frames else pd.DataFrame()
    if df is None or df.empty:
        return pd.DataFrame(columns=["Week", "Year"])

    df = df.copy()
    if "BE_Statut" not in df.columns:
        df["BE_Statut"] = ""
    df["BE_Statut"] = df["BE_Statut"].astype(str).str.upper().str.strip()
    df = df[df["BE_Statut"] == status_code.upper()].copy()

    df["BE_Numero_Str"] = df.get("BE_Numero", "").astype(str).str.replace(r"\.0$", "", regex=True).str.strip()
    df["Date_Vol"] = coerce_datetime(df.get("BE_Date_Vol", pd.NaT), errors="coerce", dayfirst=False)
    iso = df["Date_Vol"].dt.isocalendar()
    df["Week"] = iso.week.astype("Int64")
    df["Year"] = iso.year.astype("Int64")
    return df


def load_be_status_d_for_week(week: int, year: int, *, tdb_path: Path | None = None) -> pd.DataFrame:
    df = load_be_status("D", tdb_path=tdb_path)
    if df.empty:
        return df
    mask_match = (df["Week"] == week) & (df["Year"] == year)
    mask_na = df["Week"].isna() | df["Year"].isna()
    df = df[mask_match | mask_na].copy()
    df["Source"] = "mag_central"
    return df


def _norm_be(value: str) -> str:
    return normalize_be_number(value) or str(value)


def _load_export_df(path: Path) -> pd.DataFrame:
    try:
        df_export = pd.read_excel(path, sheet_name="Export planning")
    except Exception:
        df_export = pd.DataFrame()
    if df_export.empty:
        try:
            df_export = pd.read_excel(path, sheet_name=0)
        except Exception:
            df_export = pd.DataFrame()
    df_export = df_export.copy()
    df_export.columns = [str(c) for c in df_export.columns]
    if "BE_Numero" not in df_export.columns:
        df_export["BE_Numero"] = df_export.get("BE_NUM", df_export.get("BE", ""))
    df_export["BE_Key"] = df_export["BE_Numero"].apply(_norm_be)
    if "Date_Vol" in df_export.columns:
        df_export["Date_Vol"] = parse_date_series(df_export["Date_Vol"]).dt.date
    if "Heure_Vol" in df_export.columns:
        df_export["Heure_Vol"] = parse_time_series(df_export["Heure_Vol"]).dt.time
        df_export["HEURE_MIN"] = hour_min_from_series(df_export["Heure_Vol"])
    if "_STATUS" not in df_export.columns:
        df_export["_STATUS"] = "normal"
    return df_export


def _apply_update_to_export_df(
    df_export: pd.DataFrame,
    *,
    action: str,
    be_num: str,
    dest_iata: str,
    date_new: str,
    vol_new: str,
    heure_new: str,
    bene_choice: str,
    be_info: dict | pd.Series | None,
    plan_row_full: dict | pd.Series | None,
    bene_meta: Optional[dict],
    bene_changed: bool,
) -> pd.DataFrame:
    be_info = be_info if isinstance(be_info, dict) else (be_info.to_dict() if be_info is not None else {})
    plan_row_full = plan_row_full if isinstance(plan_row_full, dict) else (
        plan_row_full.to_dict() if plan_row_full is not None else {}
    )

    def _is_missing(val: object) -> bool:
        if val is None:
            return True
        try:
            if pd.isna(val):
                return True
        except Exception:
            pass
        sval = str(val).strip().lower()
        return sval in ("", "nan", "none")

    def _first_not_missing(*values: object) -> object:
        for val in values:
            if not _is_missing(val):
                return val
        return ""

    def _from_plan_row(*keys: str) -> object:
        if not plan_row_full:
            return ""
        try:
            for k in keys:
                if k in plan_row_full and not _is_missing(plan_row_full.get(k)):
                    return plan_row_full.get(k)
        except Exception:
            pass
        return ""

    if action == "Annulation":
        df_export.loc[df_export["BE_Key"] == _norm_be(be_num), "_STATUS"] = "old"
        return df_export

    df_export.loc[df_export["BE_Key"] == _norm_be(be_num), "_STATUS"] = "old"
    existing_mask = df_export["BE_Key"] == _norm_be(be_num)
    base_row = df_export.loc[existing_mask].iloc[0].to_dict() if existing_mask.any() else {}

    date_val = parse_date_series(pd.Series([date_new])).iloc[0].date() if date_new else None
    heure_val = parse_time_series(pd.Series([heure_new])).iloc[0].time() if heure_new else None
    heure_norm = normalize_hour_str(pd.Series([heure_new])).iloc[0] if heure_new else ""
    heure_min = hour_min_from_series(pd.Series([heure_new])).iloc[0] if heure_new else None

    new_row = dict(base_row)
    new_row.update(
        {
            "BE_Numero": be_num,
            "BE_Key": _norm_be(be_num),
            "Destination": dest_iata,
            "IATA": dest_iata,
            "Date_Vol": date_val,
            "Heure_Vol": heure_val,
            "Heure": heure_norm,
            "HEURE_MIN": heure_min,
            "Numero_Vol": vol_new,
            "Benevole": bene_choice,
            "_STATUS": "new",
        }
    )

    new_row["BE_Nb_Colis"] = _first_not_missing(
        new_row.get("BE_Nb_Colis"),
        be_info.get("BE_Nb_Colis"),
        be_info.get("Nb_Colis"),
        _from_plan_row("BE_Nb_Colis", "Nb_Colis", "NB_COLIS", "BE_COLIS"),
    )
    new_row["BE_Nb_Equiv"] = _first_not_missing(
        new_row.get("BE_Nb_Equiv"),
        be_info.get("BE_Nb_Equiv"),
        be_info.get("Equiv_Colis"),
        _from_plan_row("BE_Nb_Equiv", "Equiv_Colis", "BE_Equiv", "BE_Equiv_Colis"),
        new_row.get("BE_Nb_Colis"),
    )
    new_row["BE_Type"] = _first_not_missing(
        new_row.get("BE_Type"),
        be_info.get("BE_Type"),
        be_info.get("Type"),
        _from_plan_row("BE_Type", "Type", "TYPE", "BE_TYPE"),
    )
    new_row["BE_Expediteur"] = _first_not_missing(
        new_row.get("BE_Expediteur"),
        be_info.get("BE_Expediteur"),
        _from_plan_row("BE_Expediteur", "EXPEDITEUR", "EXP", "BE_EXP"),
    )
    new_row["BE_Destinataire"] = _first_not_missing(
        new_row.get("BE_Destinataire"),
        be_info.get("BE_Destinataire"),
        _from_plan_row("BE_Destinataire", "DESTINATAIRE", "BE_DEST"),
    )

    if bene_changed:
        new_row["ID"] = _first_not_missing(bene_meta.get("ID") if bene_meta else "", "")
        new_row["Telephone"] = _first_not_missing(bene_meta.get("Telephone") if bene_meta else "", "")
    else:
        new_row["ID"] = _first_not_missing(new_row.get("ID"), bene_meta.get("ID") if bene_meta else "")
        new_row["Telephone"] = _first_not_missing(
            new_row.get("Telephone"),
            bene_meta.get("Telephone") if bene_meta else "",
        )

    return pd.concat([df_export, pd.DataFrame([new_row])], ignore_index=True)


def _sort_export_df(df_export: pd.DataFrame) -> pd.DataFrame:
    return df_export.sort_values(
        by=["Date_Vol", "Heure_Vol", "BE_Numero"],
        kind="mergesort",
    ).reset_index(drop=True)


def _sheet_year(name: str) -> int | None:
    match = re.search(r"(20\d{2})", str(name))
    return int(match.group(1)) if match else None


def _sheet_year_suffix(name: str) -> str | None:
    yr = _sheet_year(name)
    return str(yr)[-2:] if yr else None


def _mag_sheet_names(wb) -> list[str]:
    names = [name for name in wb.sheetnames if str(name).strip().upper().startswith("MAG CENTRAL")]
    if names:
        return names
    if cp.SHEET_MAG_CENTRAL in wb.sheetnames:
        return [cp.SHEET_MAG_CENTRAL]
    return [wb.active.title]


def _mag_lookup_keys(be_key: str) -> list[str]:
    keys: list[str] = []

    def _add(val: object) -> None:
        if val is None:
            return
        sval = str(val).strip()
        if sval and sval not in keys:
            keys.append(sval)

    base = str(be_key).strip()
    if not base:
        return keys
    _add(base)
    _add(base.lstrip("0"))
    if base.isdigit():
        if len(base) >= 4:
            suf4 = base[-4:]
            _add(suf4)
            _add(suf4.lstrip("0"))
            try:
                _add(str(int(suf4)))
            except Exception:
                pass
        if len(base) >= 3:
            suf3 = base[-3:]
            _add(suf3)
            _add(suf3.lstrip("0"))
            try:
                _add(str(int(suf3)))
            except Exception:
                pass
    return keys


def _alt_key_for_sheet(be_key: str, sheet_name: str) -> str | None:
    if not be_key or not be_key.isdigit() or len(be_key) < 4:
        return None
    if not be_key.startswith("00"):
        return None
    suffix = _sheet_year_suffix(sheet_name)
    if not suffix:
        return None
    return f"{suffix}{be_key[-4:]}"


def _sheet_order_for_be(be_key: str, mag_sheet_names: list[str]) -> list[str]:
    preferred = []
    if be_key:
        for name in mag_sheet_names:
            suffix = _sheet_year_suffix(name)
            if suffix and be_key.startswith(suffix):
                preferred.append(name)
    return preferred + [n for n in mag_sheet_names if n not in preferred]


def _build_mag_index(ws_mag) -> dict[str, int]:
    mag_index: dict[str, int] = {}
    for row in ws_mag.iter_rows(min_row=1, max_row=ws_mag.max_row, min_col=1, max_col=1):
        val = row[0].value
        if val is None:
            continue
        sval = str(int(val)) if isinstance(val, (int, float)) else str(val).strip()
        key = normalize_be_number(sval)
        if not key:
            continue
        keys = {key, sval, key.lstrip("0")}
        for k in keys:
            mag_index[k] = row[0].row
    return mag_index


def _find_mag_target_row(
    *,
    be_key: str,
    mag_sheet_names: list[str],
    mag_indexes: dict[str, dict[str, int]],
) -> tuple[str | None, int | None]:
    base_keys = _mag_lookup_keys(be_key)
    target_sheet = None
    target_row = None
    for sheet_name in _sheet_order_for_be(be_key, mag_sheet_names):
        idx = mag_indexes.get(sheet_name, {})
        if not idx:
            continue
        alt_key = _alt_key_for_sheet(be_key, sheet_name)
        keys = ([alt_key] if alt_key else []) + base_keys
        for key in keys:
            row_idx = idx.get(key)
            if row_idx:
                target_sheet = sheet_name
                target_row = row_idx
                break
        if target_row:
            break
    return target_sheet, target_row


def _parse_mag_departure_date(date_new: str) -> date | None:
    if not date_new:
        return None
    try:
        parsed = parse_date_series(pd.Series([date_new])).iloc[0]
        if pd.isna(parsed):
            return None
        return parsed.date()
    except Exception:
        return None


def _previous_iso_week_friday(date_obj: date | None) -> date | None:
    if date_obj is None:
        return None
    try:
        mon = date.fromisocalendar(date_obj.isocalendar()[0], date_obj.isocalendar()[1], 1)
        return mon - timedelta(days=3)
    except Exception:
        return None


def _clean_bene_id(value: object) -> str:
    bene_id = str(value or "").strip()
    if bene_id.endswith(".0"):
        bene_id = bene_id[:-2]
    return bene_id


def apply_planning_updates_batch(
    path: Path,
    updates: list[dict],
    *,
    week: int,
    year: int,
    df_vols: Optional[pd.DataFrame] = None,
    df_parambenev: Optional[pd.DataFrame] = None,
    df_dispos: Optional[pd.DataFrame] = None,
    df_paramdest: Optional[pd.DataFrame] = None,
    increment_version: bool = True,
    write_mag_central: bool = False,
    tdb_source_path: Optional[Path] = None,
) -> Path:
    from asf_app.services.export_service import export_planning_excel

    df_export = _load_export_df(path)
    for upd in updates:
        df_export = _apply_update_to_export_df(
            df_export,
            action=upd.get("action", ""),
            be_num=upd.get("be_num", ""),
            dest_iata=upd.get("dest_iata", ""),
            date_new=upd.get("date_new", ""),
            vol_new=upd.get("vol_new", ""),
            heure_new=upd.get("heure_new", ""),
            bene_choice=upd.get("bene_choice", ""),
            be_info=upd.get("be_info", {}),
            plan_row_full=upd.get("plan_row_full", {}),
            bene_meta=upd.get("bene_meta", {}),
            bene_changed=bool(upd.get("bene_changed", False)),
        )

    df_export = _sort_export_df(df_export)

    if increment_version:
        export_result = export_planning_excel(
            df_export,
            week,
            year,
            df_vols=df_vols,
            df_parambenev=df_parambenev,
            df_dispos=df_dispos,
            df_paramdest=df_paramdest,
            create_tables=False,
            write_source_excel=False,
            increment_version=True,
            output_dir=path.parent,
            generate_pdf=False,
        )
    else:
        export_result = export_planning_excel(
            df_export,
            week,
            year,
            df_vols=df_vols,
            df_parambenev=df_parambenev,
            df_dispos=df_dispos,
            df_paramdest=df_paramdest,
            create_tables=False,
            write_source_excel=False,
            increment_version=False,
            output_path=path,
            generate_pdf=False,
        )
    cp.sync_local_file_to_onedrive(export_result.output_path)

    if write_mag_central:
        for upd in updates:
            _update_mag_central_for_be(
                be_num=upd.get("be_num", ""),
                action=upd.get("action", ""),
                date_new=upd.get("date_new", ""),
                heure_new=upd.get("heure_new", ""),
                vol_new=upd.get("vol_new", ""),
                bene_choice=upd.get("bene_choice", ""),
                bene_meta=upd.get("bene_meta", {}),
                tdb_source_path=tdb_source_path,
            )

    return export_result.output_path


def apply_planning_update(
    path: Path,
    action: str,
    be_num: str,
    dest_iata: str,
    date_new: str,
    vol_new: str,
    heure_new: str,
    bene_choice: str,
    be_info: pd.Series,
    week: int,
    year: int,
    df_vols: Optional[pd.DataFrame] = None,
    df_parambenev: Optional[pd.DataFrame] = None,
    df_dispos: Optional[pd.DataFrame] = None,
    df_paramdest: Optional[pd.DataFrame] = None,
    plan_row: Optional[pd.Series] = None,
    plan_row_full: Optional[pd.Series] = None,
    bene_meta: Optional[dict] = None,
    bene_changed: bool = False,
    increment_version: bool = True,
    write_mag_central: bool = False,
    tdb_source_path: Optional[Path] = None,
):
    """
    Construit un DF consolidé, applique l'action, puis regénère Export/Planning.
    Retourne le chemin du fichier planning mis à jour.
    """
    from asf_app.services.export_service import export_planning_excel

    df_export = _load_export_df(path)
    df_export = _apply_update_to_export_df(
        df_export,
        action=action,
        be_num=be_num,
        dest_iata=dest_iata,
        date_new=date_new,
        vol_new=vol_new,
        heure_new=heure_new,
        bene_choice=bene_choice,
        be_info=be_info,
        plan_row_full=plan_row_full,
        bene_meta=bene_meta,
        bene_changed=bene_changed,
    )
    df_export = _sort_export_df(df_export)

    if increment_version:
        export_result = export_planning_excel(
            df_export,
            week,
            year,
            df_vols=df_vols,
            df_parambenev=df_parambenev,
            df_dispos=df_dispos,
            df_paramdest=df_paramdest,
            create_tables=False,
            write_source_excel=False,
            increment_version=True,
            output_dir=path.parent,
            generate_pdf=False,
        )
    else:
        export_result = export_planning_excel(
            df_export,
            week,
            year,
            df_vols=df_vols,
            df_parambenev=df_parambenev,
            df_dispos=df_dispos,
            df_paramdest=df_paramdest,
            create_tables=False,
            write_source_excel=False,
            increment_version=False,
            output_path=path,
            generate_pdf=False,
        )
    cp.sync_local_file_to_onedrive(export_result.output_path)

    if write_mag_central:
        _update_mag_central_for_be(
            be_num=be_num,
            action=action,
            date_new=date_new,
            heure_new=heure_new,
            vol_new=vol_new,
            bene_choice=bene_choice,
            bene_meta=bene_meta,
            tdb_source_path=tdb_source_path,
        )
    return export_result.output_path


def _update_mag_central_for_be(
    *,
    be_num: str,
    action: str,
    date_new: str,
    heure_new: str,
    vol_new: str,
    bene_choice: str,
    bene_meta: Optional[dict],
    tdb_source_path: Optional[Path],
) -> str:
    path = Path(tdb_source_path) if tdb_source_path is not None else get_tableau_de_bord_src()
    if not path.exists():
        return "missing"

    try:
        from openpyxl import load_workbook
    except Exception:
        return "openpyxl_missing"

    try:
        wb_mag = load_workbook(path)
    except Exception:
        return "read_error"

    mag_sheet_names = _mag_sheet_names(wb_mag)
    mag_sheet_names = sorted(
        mag_sheet_names,
        key=lambda n: (_sheet_year(n) is None, _sheet_year(n) or 0, str(n)),
    )

    mag_indexes = {}
    for sheet_name in mag_sheet_names:
        try:
            ws_mag = wb_mag[sheet_name]
        except Exception:
            continue
        mag_indexes[sheet_name] = _build_mag_index(ws_mag)

    be_key = normalize_be_number(be_num) or str(be_num)
    target_sheet, target_row = _find_mag_target_row(
        be_key=be_key,
        mag_sheet_names=mag_sheet_names,
        mag_indexes=mag_indexes,
    )

    if not target_sheet or not target_row:
        return "not_found"

    ws_mag = wb_mag[target_sheet]

    action_lower = str(action).strip().lower()
    if action_lower == "annulation":
        ws_mag.cell(row=target_row, column=cp.MAG_CENTRAL_COL_DEPART_VOL).value = None
        ws_mag.cell(row=target_row, column=cp.MAG_CENTRAL_COL_ID_BENEV).value = None
        ws_mag.cell(row=target_row, column=cp.MAG_CENTRAL_COL_BENEV).value = None
        ws_mag.cell(row=target_row, column=cp.MAG_CENTRAL_COL_VOL).value = None
        ws_mag.cell(row=target_row, column=cp.MAG_CENTRAL_COL_HEURE).value = None
    else:
        date_obj = _parse_mag_departure_date(date_new)

        # Date départ mag : vendredi précédent si cellule vide
        if date_obj:
            prev_friday = _previous_iso_week_friday(date_obj)
            try:
                dm_cell = ws_mag.cell(row=target_row, column=cp.MAG_CENTRAL_COL_DEPART_MAG)
                if dm_cell.value in (None, "") and prev_friday is not None:
                    dm_cell.value = prev_friday
            except Exception:
                pass

        if date_obj:
            ws_mag.cell(row=target_row, column=cp.MAG_CENTRAL_COL_DEPART_VOL).value = date_obj

        bene_id = _clean_bene_id(bene_meta.get("ID") if bene_meta else "")

        bene_name = str(bene_choice or "").strip()
        hour_val = normalize_hour_str(pd.Series([heure_new])).iloc[0] if heure_new else ""

        ws_mag.cell(row=target_row, column=cp.MAG_CENTRAL_COL_ID_BENEV).value = bene_id
        ws_mag.cell(row=target_row, column=cp.MAG_CENTRAL_COL_BENEV).value = bene_name
        ws_mag.cell(row=target_row, column=cp.MAG_CENTRAL_COL_VOL).value = str(vol_new or "")
        ws_mag.cell(row=target_row, column=cp.MAG_CENTRAL_COL_HEURE).value = hour_val

    try:
        wb_mag.save(path)
        cp.sync_local_file_to_onedrive(path)
    except Exception:
        return "write_error"
    return "updated"

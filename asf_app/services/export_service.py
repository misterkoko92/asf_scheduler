# -*- coding: utf-8 -*-
from __future__ import annotations

import re
import shutil
from dataclasses import dataclass, field
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Callable, Optional
from zipfile import BadZipFile

import pandas as pd
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import PatternFill
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.worksheet.table import Table, TableStyleInfo

import scheduler.config_paths as cp
from asf_app.config.runtime import (
    get_onedrive_root,
    get_output_planning_dir,
    get_output_remote_path,
    get_tableau_de_bord_src,
    is_graph_onedrive,
)
from loaders.load_benevoles import get_benevoles_cached
from scheduler.format_rules import format_vol_display
from scheduler.planning_views import build_export_view
from utils.datetime_utils import (
    coerce_datetime,
    format_date_value,
    format_time_series,
    parse_time_series,
)
from utils.excel_safety import sanitize_dataframe_for_excel, sanitize_excel_value
from utils.export_pdf import export_first_sheet_to_pdf
from utils.identifiers import normalize_be_number


@dataclass
class ExportResult:
    output_path: Path
    pdf_path: Optional[Path] = None
    mag_write_method: str = "disabled"
    warnings: list[str] = field(default_factory=list)


EXPORT_PDF_ERRORS = (
    FileNotFoundError,
    OSError,
    PermissionError,
    RuntimeError,
    ValueError,
    TypeError,
)

EXCEL_IO_ERRORS = (
    FileNotFoundError,
    OSError,
    PermissionError,
    RuntimeError,
    ValueError,
    TypeError,
    BadZipFile,
    InvalidFileException,
)


_PLAN_DAY_BLOCKS: dict[int, tuple[int, int]] = {
    0: (4, 32),
    1: (35, 63),
    2: (66, 94),
    3: (97, 125),
    4: (128, 156),
    5: (159, 187),
    6: (190, 218),
}

_PLAN_KEEP_ROWS: set[int] = {
    3,
    4,
    5,
    6,
    31,
    32,
    33,
    34,
    35,
    36,
    37,
    62,
    63,
    64,
    65,
    66,
    67,
    68,
    93,
    94,
    95,
    96,
    97,
    98,
    99,
    124,
    125,
    126,
    127,
    128,
    129,
    130,
    155,
    156,
    157,
    158,
    159,
    160,
    161,
    186,
    187,
    188,
    189,
    190,
    191,
    192,
    217,
    218,
    219,
}

_PLAN_MIDDLE_MOVES: tuple[tuple[int, int, int], ...] = (
    (17, 4, 32),
    (48, 35, 63),
    (79, 66, 94),
    (110, 97, 125),
    (141, 128, 156),
    (172, 159, 187),
    (203, 190, 218),
)


def _resolve_template() -> Path:
    onedrive_maquette = (
        get_onedrive_root()
        / "Planning MAB"
        / "ASFmm PLANNING 2025"
        / "aaSOURCE"
        / "Planning-maquette.xlsx"
    )
    template = onedrive_maquette if onedrive_maquette.exists() else cp.PLANNING_TEMPLATE
    return template


def _create_minimal_workbook(path: Path, *, week: int, year: int) -> None:
    from openpyxl import Workbook

    wb_min = Workbook()
    ws_min = wb_min.active
    ws_min.title = "Planning SXX"
    try:
        ws_min["A1"] = date.fromisocalendar(int(year), int(week), 1)
    except (TypeError, ValueError):
        ws_min["A1"] = None
    ws_min["Q1"] = 0
    wb_min.create_sheet("Export planning")
    wb_min.create_sheet("Data Vols")
    wb_min.create_sheet("Data Benevoles")
    wb_min.save(path)


def _load_workbook_with_minimal_fallback(path: Path, *, week: int, year: int):
    try:
        return load_workbook(path)
    except (FileNotFoundError, InvalidFileException, BadZipFile, OSError, ValueError):
        _create_minimal_workbook(path, week=week, year=year)
        return load_workbook(path)


def _resolve_target_sheets(wb):
    ws_plan = wb.worksheets[0]
    for sh in wb.sheetnames:
        if sh.lower().startswith("planning"):
            ws_plan = wb[sh]
            break
    ws_export = (
        wb["Export planning"]
        if "Export planning" in wb.sheetnames
        else (wb.worksheets[1] if len(wb.worksheets) > 1 else wb.create_sheet("Export planning"))
    )
    ws_vols = (
        wb["Data Vols"]
        if "Data Vols" in wb.sheetnames
        else (wb.worksheets[2] if len(wb.worksheets) > 2 else wb.create_sheet("Data Vols"))
    )
    ws_bene = (
        wb["Data Benevoles"]
        if "Data Benevoles" in wb.sheetnames
        else (wb.worksheets[3] if len(wb.worksheets) > 3 else wb.create_sheet("Data Benevoles"))
    )
    return ws_plan, ws_export, ws_vols, ws_bene


def _week_year_from_ws_plan(ws_plan, *, fallback_week: int, fallback_year: int) -> tuple[int, int]:
    val = ws_plan["A1"].value
    wk, yr = fallback_week, fallback_year
    try:
        if isinstance(val, datetime):
            wk = val.isocalendar()[1]
            yr = val.isocalendar()[0]
        else:
            dt = coerce_datetime(val, errors="coerce", dayfirst=True)
            if pd.notna(dt):
                wk = dt.isocalendar()[1]
                yr = dt.isocalendar()[0]
    except (TypeError, ValueError, AttributeError):
        pass
    return int(wk), int(yr)


def _extract_version_from_planning_name(name: str, *, week: int, year: int) -> int | None:
    stem = Path(name).stem
    m = re.search(r"SEMAINE\s*(20\d{2})\D+(\d{1,2})\D+(\d+)", stem, re.IGNORECASE)
    if m:
        try:
            y, w, v = int(m.group(1)), int(m.group(2)), int(m.group(3))
            if y == year and w == week:
                return v
        except (TypeError, ValueError):
            pass
    m = re.search(r"N[°o]?\s*(\d{1,2}).*?(20\d{2}).*?v(\d+)", stem, re.IGNORECASE)
    if m:
        try:
            w, y, v = int(m.group(1)), int(m.group(2)), int(m.group(3))
            if y == year and w == week:
                return v
        except (TypeError, ValueError):
            pass
    m = re.search(r"N[°o]?\s*(\d{1,2}).*?(20\d{2})", stem, re.IGNORECASE)
    if m:
        try:
            w, y = int(m.group(1)), int(m.group(2))
            if y == year and w == week:
                return 1
        except (TypeError, ValueError):
            pass
    return None


def _build_bene_display_map(df_parambenev: pd.DataFrame | None) -> dict[str, str]:
    display_map: dict[str, str] = {}
    if df_parambenev is None or getattr(df_parambenev, "empty", True):
        return display_map

    tmp = df_parambenev.copy()
    tmp["Benevole"] = tmp.get("Benevole", tmp.get("BENEVOLE", ""))
    tmp["Prenom_Court"] = tmp.get("Prenom_Court", tmp.get("Prenom court", ""))
    tmp["Nom"] = tmp.get("Nom", "")
    for _, row in tmp.iterrows():
        benevole = str(row.get("Benevole", "")).strip()
        prenom_court = str(row.get("Prenom_Court", "")).strip()
        nom = str(row.get("Nom", "")).strip().upper()
        display_map[benevole] = f"{prenom_court} {nom}".strip()
    return display_map


def _format_bene_display(name: object, *, display_map: dict[str, str]) -> str:
    name_key = str(name).strip()
    if name_key in display_map:
        return display_map[name_key]
    parts = name_key.split()
    if len(parts) >= 2:
        return f"{parts[0][0].upper()}. {' '.join(parts[1:]).upper()}"
    return name_key.upper()


def _prepare_export_dataframe(
    df: pd.DataFrame,
    *,
    df_paramdest: pd.DataFrame | None,
    df_vols: pd.DataFrame | None,
    df_parambenev: pd.DataFrame | None,
) -> pd.DataFrame:
    dfp = build_export_view(df, df_paramdest=df_paramdest, df_vols=df_vols).fillna("")
    dfp["DATE"] = coerce_datetime(dfp.get("Date_Vol", dfp.get("DATE", "")), errors="coerce")
    heures_raw = dfp.get("Heure_Vol", dfp.get("HEURE VOL", pd.Series([""] * len(dfp), index=dfp.index)))
    if not isinstance(heures_raw, pd.Series):
        heures_raw = pd.Series([heures_raw] * len(dfp), index=dfp.index)
    dfp["HEURE_VOL_DT"] = parse_time_series(
        heures_raw,
        allow_hour_only=True,
        allow_general_fallback=True,
        strip_spaces=True,
        lowercase=True,
    )
    dfp["HEURE_VOL"] = dfp["HEURE_VOL_DT"]
    dfp["HEURE_MIN"] = (
        dfp["HEURE_VOL_DT"].dt.hour.fillna(99).astype(int) * 60
        + dfp["HEURE_VOL_DT"].dt.minute.fillna(59).astype(int)
    )
    dfp = dfp.sort_values(by=["DATE", "HEURE_MIN", "Destination", "Numero_Vol"], kind="mergesort")

    bene_display_map = _build_bene_display_map(df_parambenev)
    dfp["BENEVOLE_DISP"] = dfp.get("Benevole", dfp.get("BENEVOLE", "")).apply(
        lambda value: _format_bene_display(value, display_map=bene_display_map)
    )
    dfp["VILLE"] = dfp.get("Ville", dfp.get("Dest_Ville", dfp.get("Destination", ""))).astype(str).str.upper()
    dfp["IATA"] = dfp.get("IATA", dfp.get("Dest_IATA", dfp.get("Destination", ""))).astype(str).str.upper()

    if df_vols is not None and not getattr(df_vols, "empty", True):
        dfp = _apply_routing_fallback_from_vols(dfp, df_vols=df_vols)

    dfp["VOL_AFF"] = dfp.get("Numero_Vol", dfp.get("Numero_Vol_Aff", "")).apply(format_vol_display)
    dfp["ROUTING"] = dfp.get("Routing", dfp.get("ROUTING", "")).astype(str).str.replace(",", "-").str.upper()
    dfp["HEURE_AFF"] = format_time_series(
        heures_raw,
        fmt="%Hh%M",
        allow_hour_only=True,
        allow_general_fallback=True,
        strip_spaces=True,
        lowercase=True,
    )
    dfp["BE_NUM"] = dfp.get("BE_Numero", dfp.get("NUMERO BE", dfp.get("BE_Num", ""))).apply(normalize_be_number)
    dfp["BE_COLIS"] = pd.to_numeric(dfp.get("BE_Nb_Colis", 0), errors="coerce").fillna(0).astype(int)
    dfp["BE_TYPE"] = dfp.get("BE_Type", "")
    dfp["BE_EXP"] = dfp.get("BE_Expediteur", "")
    dfp["BE_DEST"] = dfp.get("BE_Destinataire", "")
    dfp["BE_KEY"] = dfp["BE_NUM"]
    dfp["_STATUS"] = dfp.get("_STATUS", "normal")
    return dfp


def _normalize_vol_key(value: object) -> str:
    normalized = str(value or "").strip().upper()
    if normalized.startswith("AF"):
        normalized = normalized.replace("AF", "").strip()
    digits = "".join(ch for ch in normalized if ch.isdigit())
    if digits:
        try:
            return str(int(digits))
        except ValueError:
            return digits.lstrip("0") or digits
    return normalized


def _apply_routing_fallback_from_vols(
    df_export: pd.DataFrame,
    *,
    df_vols: pd.DataFrame,
) -> pd.DataFrame:
    vols_map = df_vols.copy()
    if "Date_Vol_dt" in vols_map.columns:
        vols_date = coerce_datetime(vols_map["Date_Vol_dt"], errors="coerce")
    else:
        vols_date = coerce_datetime(vols_map.get("Date_Vol", ""), errors="coerce", dayfirst=True)
    vols_map["_DATE_KEY"] = vols_date.dt.date
    vols_map["_VOL_KEY"] = vols_map.get("Numero_Vol", "").apply(_normalize_vol_key)
    routing_map = (
        vols_map.dropna(subset=["Routing"])
        .drop_duplicates(subset=["_DATE_KEY", "_VOL_KEY"])
        .set_index(["_DATE_KEY", "_VOL_KEY"])["Routing"]
        .to_dict()
    )

    out = df_export.copy()
    out["_DATE_KEY"] = coerce_datetime(out["DATE"], errors="coerce").dt.date
    out["_VOL_KEY"] = out.get("Numero_Vol", "").apply(_normalize_vol_key)
    routing_series = out["Routing"] if "Routing" in out.columns else pd.Series([""] * len(out), index=out.index)
    mask_routing_empty = (
        routing_series.fillna("")
        .astype(str)
        .str.strip()
        .str.lower()
        .replace({"nan": "", "none": ""})
        .eq("")
    )
    if mask_routing_empty.any():

        def _route_fallback(row: pd.Series) -> object:
            key = (row.get("_DATE_KEY"), row.get("_VOL_KEY"))
            if key in routing_map:
                return routing_map[key]
            return ""

        out.loc[mask_routing_empty, "Routing"] = out.loc[mask_routing_empty].apply(_route_fallback, axis=1)
    return out.drop(columns=["_DATE_KEY", "_VOL_KEY"], errors="ignore")


def _friday_previous_week(wk: int, yr: int):
    try:
        mon = date.fromisocalendar(yr, wk, 1)
        return mon - timedelta(days=3)
    except (TypeError, ValueError):
        return None


def _sheet_year(name: str) -> int | None:
    match = re.search(r"(20\d{2})", str(name))
    return int(match.group(1)) if match else None


def _sheet_year_suffix(name: str) -> str | None:
    yr = _sheet_year(name)
    return str(yr)[-2:] if yr else None


def _mag_sheet_names(wb) -> list[str]:
    names = [name for name in wb.sheetnames if str(name).strip().upper().startswith("MAG CENTRAL")]
    if names:
        return names
    if cp.SHEET_MAG_CENTRAL in wb.sheetnames:
        return [cp.SHEET_MAG_CENTRAL]
    return [wb.active.title]


def _build_mag_index(ws_mag) -> dict[str, int]:
    mag_index: dict[str, int] = {}
    for row in ws_mag.iter_rows(min_row=1, max_row=ws_mag.max_row, min_col=1, max_col=20):
        val = row[0].value
        if val is None:
            continue
        sval = str(int(val)) if isinstance(val, (int, float)) else str(val).strip()
        key = normalize_be_number(sval)
        if not key:
            continue
        keys = {key, sval, key.lstrip("0")}
        for lookup_key in keys:
            mag_index[lookup_key] = row[0].row
    return mag_index


def _sheet_order_for_be(be_key: str, mag_sheet_names: list[str], *, preferred_year: int | None) -> list[str]:
    preferred = []
    if be_key:
        for name in mag_sheet_names:
            suffix = _sheet_year_suffix(name)
            if suffix and be_key.startswith(suffix):
                preferred.append(name)
    if isinstance(preferred_year, int):
        for name in mag_sheet_names:
            if _sheet_year(name) == int(preferred_year) and name not in preferred:
                preferred.append(name)
    return preferred + [name for name in mag_sheet_names if name not in preferred]


def _mag_lookup_keys(be_key: str) -> list[str]:
    keys: list[str] = []

    def _add(val: object) -> None:
        if val is None:
            return
        sval = str(val).strip()
        if sval and sval not in keys:
            keys.append(sval)

    base = str(be_key).strip()
    if not base:
        return keys
    _add(base)
    _add(base.lstrip("0"))
    if base.isdigit():
        if len(base) >= 4:
            suf4 = base[-4:]
            _add(suf4)
            _add(suf4.lstrip("0"))
            try:
                _add(str(int(suf4)))
            except ValueError:
                pass
        if len(base) >= 3:
            suf3 = base[-3:]
            _add(suf3)
            _add(suf3.lstrip("0"))
            try:
                _add(str(int(suf3)))
            except ValueError:
                pass
    return keys


def _alt_key_for_sheet(be_key: str, sheet_name: str) -> str | None:
    if not be_key or not be_key.isdigit() or len(be_key) < 4:
        return None
    if not be_key.startswith("00"):
        return None
    suffix = _sheet_year_suffix(sheet_name)
    if not suffix:
        return None
    return f"{suffix}{be_key[-4:]}"


def _find_mag_row(
    *,
    be_key: str,
    mag_sheet_names: list[str],
    mag_indexes: dict[str, dict[str, int]],
    preferred_year: int | None,
) -> tuple[str | None, int | None]:
    if not be_key:
        return None, None
    base_keys = _mag_lookup_keys(be_key)
    for sheet_name in _sheet_order_for_be(be_key, mag_sheet_names, preferred_year=preferred_year):
        idx = mag_indexes.get(sheet_name, {})
        if not idx:
            continue
        alt_key = _alt_key_for_sheet(be_key, sheet_name)
        keys = ([alt_key] if alt_key else []) + base_keys
        for lookup_key in keys:
            row_idx = idx.get(lookup_key)
            if row_idx:
                return sheet_name, row_idx
    return None, None


def _safe_text(val: object) -> str:
    if val is None:
        return ""
    try:
        if pd.isna(val):
            return ""
    except (TypeError, ValueError):
        pass
    return str(val).strip()


def _build_update_items_by_sheet(updates_by_sheet: dict[str, dict[int, dict[int, object]]]) -> dict[str, list[tuple[int, int, object]]]:
    update_items_by_sheet: dict[str, list[tuple[int, int, object]]] = {}
    for sheet_name, rows in updates_by_sheet.items():
        sheet_items: list[tuple[int, int, object]] = []
        for row_idx, cols in rows.items():
            for col_idx, val in cols.items():
                sheet_items.append((row_idx, col_idx, val))
        if sheet_items:
            update_items_by_sheet[sheet_name] = sheet_items
    return update_items_by_sheet


def _resolve_depart_mag_date(
    be_key: object,
    *,
    map_depart_mag: dict,
    week: int,
    year: int,
) -> date | None | object:
    if be_key in map_depart_mag:
        return map_depart_mag[be_key]
    try:
        mon = date.fromisocalendar(int(year), int(week), 1)
        return mon - timedelta(days=3)
    except (TypeError, ValueError):
        return None


def _clear_worksheet_tables(ws) -> None:
    try:
        tbls = ws._tables
        if isinstance(tbls, dict):
            tbls.clear()
        else:
            ws._tables = []
    except (AttributeError, TypeError):
        pass


def _add_excel_table_if_needed(
    ws,
    *,
    create_tables: bool,
    table_name: str,
) -> None:
    max_row = ws.max_row
    max_col = ws.max_column
    if not create_tables or max_row < 2 or max_col < 1:
        return
    from openpyxl.utils import get_column_letter

    ref = f"A1:{get_column_letter(max_col)}{max_row}"
    tab = Table(displayName=table_name, ref=ref)
    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    tab.tableStyleInfo = style
    ws.add_table(tab)


def _write_dataframe_to_sheet(
    ws,
    df_to_write: pd.DataFrame | None,
    *,
    create_tables: bool,
    table_name: str | None = None,
) -> None:
    from openpyxl.utils.dataframe import dataframe_to_rows

    _clear_worksheet_tables(ws)
    ws.delete_rows(1, ws.max_row)
    if df_to_write is None or getattr(df_to_write, "empty", True):
        return
    for row in dataframe_to_rows(df_to_write, index=False, header=True):
        ws.append(row)
    if table_name:
        _add_excel_table_if_needed(
            ws,
            create_tables=create_tables,
            table_name=table_name,
        )


def _export_pdf_with_warning(
    out_path: Path,
    *,
    enabled: bool,
    pdf_exporter: Callable[[Path, Path], Path],
    warnings: list[str],
) -> Path | None:
    if not enabled:
        return None
    pdf_target = out_path.with_suffix(".pdf")
    try:
        generated = pdf_exporter(out_path, pdf_target)
        pdf_path = Path(generated) if generated else None
        if pdf_path is None or not pdf_path.exists():
            raise RuntimeError("PDF non généré.")
        return pdf_path
    except EXPORT_PDF_ERRORS:
        warnings.append("PDF non généré : Excel non accessible pour l’export automatique.")
        return None


def _reset_planning_grid(ws_plan) -> None:
    for row in ws_plan.iter_rows(min_row=3, max_row=ws_plan.max_row, min_col=4, max_col=17):
        for cell in row:
            if isinstance(cell, MergedCell):
                continue
            if cell.coordinate in ("K219", "L219"):
                continue
            cell.value = None
        ws_plan.row_dimensions[row[0].row].hidden = False


def _apply_status_fill(ws_plan, *, row_idx: int, status: str) -> None:
    fill_color = None
    if status.startswith("old"):
        fill_color = PatternFill("solid", fgColor="F8CCCC")
    elif status == "new":
        fill_color = PatternFill("solid", fgColor="CFE2FF")
    if fill_color:
        for col_idx in range(4, 18):
            ws_plan.cell(row=row_idx, column=col_idx).fill = fill_color


def _write_planning_row(
    ws_plan,
    *,
    row_idx: int,
    row_data: pd.Series,
    bene_val: object,
    is_first: bool,
    safe_excel: Callable[[object], object],
) -> None:
    status = str(row_data.get("_STATUS", "normal")).lower()
    ws_plan.cell(row=row_idx, column=4).value = safe_excel(bene_val)
    if is_first:
        ws_plan.cell(row=row_idx, column=6).value = safe_excel(row_data["VILLE"])
        ws_plan.cell(row=row_idx, column=7).value = safe_excel(row_data["IATA"])
        ws_plan.cell(row=row_idx, column=8).value = safe_excel(row_data["ROUTING"])
        ws_plan.cell(row=row_idx, column=9).value = safe_excel(row_data["VOL_AFF"])
        ws_plan.cell(row=row_idx, column=10).value = safe_excel(row_data["HEURE_AFF"])

    ws_plan.cell(row=row_idx, column=11).value = safe_excel(row_data["BE_NUM"])
    be_colis_val = "" if status.startswith("old") else row_data["BE_COLIS"]
    ws_plan.cell(row=row_idx, column=12).value = safe_excel(be_colis_val)
    ws_plan.cell(row=row_idx, column=13).value = safe_excel(row_data["BE_TYPE"])

    dep_mag = row_data.get("DEPART_MAG")
    if isinstance(dep_mag, pd.Timestamp):
        dep_mag = dep_mag.date()
    dep_mag_str = format_date_value(dep_mag, fmt="%d/%m/%y", default="")
    ws_plan.cell(row=row_idx, column=15).value = safe_excel(dep_mag_str)
    ws_plan.cell(row=row_idx, column=16).value = safe_excel(row_data["BE_EXP"])
    ws_plan.cell(row=row_idx, column=17).value = safe_excel(row_data["BE_DEST"])

    _apply_status_fill(ws_plan, row_idx=row_idx, status=status)


def _hide_non_keep_rows(
    ws_plan,
    *,
    start: int,
    end: int,
    keep_rows: set[int],
) -> None:
    for row_idx in range(start, end + 1):
        if row_idx not in keep_rows:
            ws_plan.row_dimensions[row_idx].hidden = True


def _populate_planning_sheet(
    ws_plan,
    *,
    dfp: pd.DataFrame,
    day_blocks: dict[int, tuple[int, int]],
    keep_rows: set[int],
    safe_excel: Callable[[object], object],
) -> None:
    for day_idx in range(7):
        block = day_blocks.get(day_idx)
        if block is None:
            continue
        start, end = block
        current_row = start
        df_day = dfp[dfp["DATE"].dt.dayofweek == day_idx]
        if df_day.empty:
            _hide_non_keep_rows(ws_plan, start=start, end=end, keep_rows=keep_rows)
            continue

        df_day = df_day.sort_values(by=["DATE", "HEURE_MIN", "VOL_AFF", "Destination"], kind="mergesort")
        for (_, _, _), df_vol in df_day.groupby(["DATE", "HEURE_MIN", "VOL_AFF"], sort=False):
            df_vol = df_vol.reset_index(drop=True)
            if current_row > end:
                break
            bene_list = list(dict.fromkeys(df_vol["BENEVOLE_DISP"].tolist()))
            for idx, row_data in df_vol.iterrows():
                if current_row > end:
                    break
                bene_val = bene_list[idx] if idx < len(bene_list) else ""
                _write_planning_row(
                    ws_plan,
                    row_idx=current_row,
                    row_data=row_data,
                    bene_val=bene_val,
                    is_first=(idx == 0),
                    safe_excel=safe_excel,
                )
                current_row += 1
            current_row += 2

        _hide_non_keep_rows(ws_plan, start=current_row, end=end, keep_rows=keep_rows)


def _move_cell_value_to_visible_middle(
    ws_plan,
    *,
    src_row: int,
    start: int,
    end: int,
    col_letter: str = "A",
) -> None:
    src = ws_plan[f"{col_letter}{src_row}"]
    if src.value is None:
        return

    def _is_visible(row_idx: int) -> bool:
        return not bool(ws_plan.row_dimensions[row_idx].hidden)

    visible_rows = [row_idx for row_idx in range(start, end + 1) if _is_visible(row_idx)]
    if not visible_rows:
        return
    dest_row = visible_rows[len(visible_rows) // 2]
    if dest_row == src_row:
        return
    dst = ws_plan[f"{col_letter}{dest_row}"]
    dst.value = src.value
    dst._style = src._style
    src.value = None


def _apply_planning_layout(
    ws_plan,
    *,
    hidden_columns: tuple[str, ...] = ("B", "C", "E", "G", "N"),
    auto_width_columns: tuple[str, ...] = ("P", "Q"),
    middle_moves: tuple[tuple[int, int, int], ...] = _PLAN_MIDDLE_MOVES,
) -> None:
    for col_letter in hidden_columns:
        ws_plan.column_dimensions[col_letter].hidden = True
    for col_letter in auto_width_columns:
        col = ws_plan[col_letter]
        max_len = max((len(str(cell.value)) for cell in col if cell.value), default=10)
        ws_plan.column_dimensions[col_letter].width = max(10, min(max_len + 2, 40))
    for src_row, start, end in middle_moves:
        _move_cell_value_to_visible_middle(
            ws_plan,
            src_row=src_row,
            start=start,
            end=end,
        )


def _prepare_output_workbook_path(
    *,
    week: int,
    year: int,
    template: Path,
    has_template: bool,
    output_path: Path | None,
    output_dir: Path | None,
    skip_versioning: bool,
) -> tuple[Path, bool]:
    if output_path is not None:
        out_path = Path(output_path)
        if not out_path.exists():
            out_path.parent.mkdir(parents=True, exist_ok=True)
            if has_template:
                shutil.copy2(template, out_path)
            else:
                _create_minimal_workbook(out_path, week=week, year=year)
        return out_path, True

    filename = f"ASFmm - PLANNING SEMAINE {year}-{week:02d}-TMP.xlsx"
    planning_dir = (
        output_dir
        if output_dir is not None
        else (
            get_output_planning_dir()
            if is_graph_onedrive()
            else get_onedrive_root() / "Planning MAB" / f"ASFmm PLANNING {year}"
        )
    )
    planning_dir.mkdir(parents=True, exist_ok=True)
    out_path = planning_dir / filename
    if has_template:
        shutil.copy2(template, out_path)
    else:
        _create_minimal_workbook(out_path, week=week, year=year)
    return out_path, bool(skip_versioning)


def _resolve_benevoles_export_dataframe(
    df_dispos_safe: pd.DataFrame | None,
    *,
    benev_path: Path | None,
) -> pd.DataFrame | None:
    df_dispo_export = df_dispos_safe
    if df_dispo_export is None:
        try:
            df_dispo_export = get_benevoles_cached(planning_path=benev_path)
        except (RuntimeError, ValueError, OSError):
            df_dispo_export = None

    if df_dispo_export is None or getattr(df_dispo_export, "empty", True):
        return None

    df_dispo_export = sanitize_dataframe_for_excel(df_dispo_export.copy())
    cols_order = [
        "ID",
        "Benevole",
        "Nom",
        "Prenom",
        "Prenom_Court",
        "Date",
        "Heure_Arrivee",
        "Heure_Depart",
    ]
    cols_present = [col for col in cols_order if col in df_dispo_export.columns]
    return df_dispo_export[cols_present]


def _collect_existing_planning_versions(
    planning_dir_final: Path,
    *,
    week_final: int,
    year_final: int,
) -> tuple[list[Path], int]:
    existing_files = [
        path for path in planning_dir_final.glob("ASFmm - PLANNING SEMAINE *.xls*") if path.is_file()
    ]
    versions: list[int] = []
    for path in existing_files:
        version = _extract_version_from_planning_name(path.name, week=week_final, year=year_final)
        if version is not None:
            versions.append(version)
    max_version = max(versions) if versions else 0
    return existing_files, max_version


def _archive_latest_planning_if_needed(
    *,
    increment_version: bool,
    max_version: int,
    existing_files: list[Path],
    planning_dir_final: Path,
    week_final: int,
    year_final: int,
) -> None:
    if increment_version or max_version <= 0:
        return
    history_dir = planning_dir_final / "Historique"
    history_dir.mkdir(parents=True, exist_ok=True)
    candidates = [
        path
        for path in existing_files
        if _extract_version_from_planning_name(path.name, week=week_final, year=year_final) == max_version
    ]
    if not candidates:
        return

    def _mtime(path: Path) -> float:
        try:
            return path.stat().st_mtime
        except OSError:
            return 0.0

    current = sorted(candidates, key=_mtime, reverse=True)[0]
    destination = history_dir / current.name
    if destination.exists():
        base = destination.stem
        suffix = destination.suffix
        counter = 2
        while destination.exists():
            destination = history_dir / f"{base}-{counter}{suffix}"
            counter += 1
    try:
        shutil.move(current, destination)
    except (OSError, shutil.Error):
        pass


def _resolve_target_planning_path(
    *,
    planning_dir_final: Path,
    week_final: int,
    year_final: int,
    max_version: int,
    increment_version: bool,
) -> tuple[int, Path]:
    if increment_version:
        version_num = max_version + 1 if max_version else 1
    else:
        version_num = max_version if max_version else 1

    version_str = f"{int(version_num):02d}"
    base_name = f"ASFmm - PLANNING SEMAINE {year_final}-{week_final:02d}-{version_str}"
    target_path = planning_dir_final / f"{base_name}.xlsx"

    if increment_version:
        while target_path.exists():
            version_num += 1
            version_str = f"{int(version_num):02d}"
            base_name = f"ASFmm - PLANNING SEMAINE {year_final}-{week_final:02d}-{version_str}"
            target_path = planning_dir_final / f"{base_name}.xlsx"

    return int(version_num), target_path


def _set_q1_version(ws_plan, version_num: int) -> None:
    try:
        ws_plan["Q1"].value = int(version_num)
    except (TypeError, ValueError):
        pass


def _increment_q1_if_requested(ws_plan, *, increment_version: bool) -> None:
    if not increment_version:
        return
    try:
        q1_val = ws_plan["Q1"].value
        q1_num = int(q1_val) if q1_val not in (None, "") else 0
        ws_plan["Q1"].value = q1_num + 1
    except (TypeError, ValueError):
        pass


def _save_sync_and_move_planning_output(
    wb,
    *,
    out_path: Path,
    target_path: Path,
    year_final: int,
) -> Path:
    wb.save(out_path)
    if is_graph_onedrive():
        remote_path = get_output_remote_path(year_final, out_path.name)
        cp.sync_local_file_to_onedrive(out_path, remote_path=remote_path, conflict_behavior="rename")
    if out_path != target_path:
        try:
            shutil.move(out_path, target_path)
        except (OSError, shutil.Error):
            wb.save(target_path)
        out_path = target_path

    wb.save(out_path)
    return out_path


def _update_mag_central_dates_for_export(
    *,
    df_export: pd.DataFrame,
    week: int,
    year: int,
    tdb_source_path: Path | None,
) -> tuple[dict, str]:
    path = Path(tdb_source_path) if tdb_source_path is not None else get_tableau_de_bord_src()
    if not path.exists():
        return {}, "missing"
    try:
        wb_mag = load_workbook(path)
    except (InvalidFileException, BadZipFile, OSError, ValueError):
        return {}, "read_error"

    mag_sheet_names = _mag_sheet_names(wb_mag)
    mag_sheet_names = sorted(
        mag_sheet_names,
        key=lambda name: (_sheet_year(name) is None, _sheet_year(name) or 0, str(name)),
    )

    mag_sheets: dict[str, object] = {}
    mag_indexes: dict[str, dict[str, int]] = {}
    for sheet_name in mag_sheet_names:
        try:
            ws_mag = wb_mag[sheet_name]
        except KeyError:
            continue
        mag_sheets[sheet_name] = ws_mag
        mag_indexes[sheet_name] = _build_mag_index(ws_mag)

    if not mag_indexes:
        return {}, "read_error"

    prev_friday = _friday_previous_week(week, year)
    used_dates: dict = {}
    updates_by_sheet: dict[str, dict[int, dict[int, object]]] = {}

    preferred_year = year if isinstance(year, int) else None
    for _, row in df_export.iterrows():
        be_key = row.get("BE_KEY", "")
        if not be_key:
            continue
        found_sheet_name, found_row_idx = _find_mag_row(
            be_key=be_key,
            mag_sheet_names=mag_sheet_names,
            mag_indexes=mag_indexes,
            preferred_year=preferred_year,
        )
        if found_sheet_name is None or found_row_idx is None:
            continue
        ws_mag = mag_sheets.get(found_sheet_name)
        if ws_mag is None:
            continue
        if prev_friday:
            dm_cell = ws_mag.cell(row=found_row_idx, column=cp.MAG_CENTRAL_COL_DEPART_MAG)
            if dm_cell.value in (None, ""):
                updates_by_sheet.setdefault(found_sheet_name, {}).setdefault(found_row_idx, {})[
                    cp.MAG_CENTRAL_COL_DEPART_MAG
                ] = prev_friday
                used_dates[be_key] = prev_friday
            else:
                used_dates[be_key] = dm_cell.value

        date_vol = row.get("DATE")
        if isinstance(date_vol, (pd.Timestamp, date)):
            dv_value = date_vol.date() if isinstance(date_vol, pd.Timestamp) else date_vol
            updates_by_sheet.setdefault(found_sheet_name, {}).setdefault(found_row_idx, {})[
                cp.MAG_CENTRAL_COL_DEPART_VOL
            ] = dv_value

        bene_id = _safe_text(row.get("ID", ""))
        if bene_id.endswith(".0"):
            bene_id = bene_id[:-2]
        bene_disp = _safe_text(row.get("BENEVOLE_DISP", row.get("Benevole", "")))
        vol_aff = _safe_text(row.get("VOL_AFF", row.get("Numero_Vol", "")))
        heure_aff = _safe_text(row.get("HEURE_AFF", ""))
        if bene_id or bene_disp or vol_aff or heure_aff:
            row_updates = updates_by_sheet.setdefault(found_sheet_name, {}).setdefault(found_row_idx, {})
            row_updates[cp.MAG_CENTRAL_COL_ID_BENEV] = bene_id
            row_updates[cp.MAG_CENTRAL_COL_BENEV] = bene_disp
            row_updates[cp.MAG_CENTRAL_COL_VOL] = vol_aff
            row_updates[cp.MAG_CENTRAL_COL_HEURE] = heure_aff

    update_items_by_sheet = _build_update_items_by_sheet(updates_by_sheet)
    if not update_items_by_sheet:
        return used_dates, "no_updates"

    try:
        from utils.excel_automation import update_excel_cells

        all_ok = True
        for sheet_name, items in update_items_by_sheet.items():
            if not update_excel_cells(path, sheet_name, items):
                all_ok = False
                break
        if all_ok:
            cp.sync_local_file_to_onedrive(path)
            return used_dates, "excel"
    except (ImportError, RuntimeError, OSError, TypeError, ValueError):
        pass

    for sheet_name, items in update_items_by_sheet.items():
        ws_mag = mag_sheets.get(sheet_name)
        if ws_mag is None:
            continue
        for row_idx, col_idx, val in items:
            ws_mag.cell(row=row_idx, column=col_idx).value = val

    try:
        wb_mag.save(path)
        cp.sync_local_file_to_onedrive(path)
    except EXCEL_IO_ERRORS:
        pass
    return used_dates, "openpyxl"


def export_planning_excel(
    df,
    week,
    year,
    *,
    df_vols=None,
    df_parambenev=None,
    df_dispos=None,
    df_paramdest=None,
    create_tables: bool = True,
    write_source_excel: bool = False,
    increment_version: bool = True,
    benev_path: Path | None = None,
    tdb_source_path: Path | None = None,
    pdf_exporter: Callable[[Path, Path], Path] | None = None,
    output_path: Path | None = None,
    output_dir: Path | None = None,
    skip_versioning: bool = False,
    generate_pdf: bool = True,
) -> ExportResult:
    warnings: list[str] = []
    pdf_exporter = pdf_exporter or export_first_sheet_to_pdf

    template = _resolve_template()
    has_template = template.exists()
    if not has_template:
        warnings.append(f"Maquette introuvable ({template}); génération avec classeur minimal.")

    if output_dir is not None:
        output_dir = Path(output_dir)

    out_path, skip_versioning = _prepare_output_workbook_path(
        week=week,
        year=year,
        template=template,
        has_template=has_template,
        output_path=Path(output_path) if output_path is not None else None,
        output_dir=output_dir,
        skip_versioning=skip_versioning,
    )
    wb = _load_workbook_with_minimal_fallback(out_path, week=week, year=year)
    ws_plan, ws_export, ws_vols, ws_bene = _resolve_target_sheets(wb)

    dfp = _prepare_export_dataframe(
        df,
        df_paramdest=df_paramdest,
        df_vols=df_vols,
        df_parambenev=df_parambenev,
    )
    safe_excel = sanitize_excel_value

    mag_write_method = "disabled"
    map_depart_mag: dict = {}
    if write_source_excel:
        map_depart_mag, mag_write_method = _update_mag_central_dates_for_export(
            df_export=dfp,
            week=week,
            year=year,
            tdb_source_path=tdb_source_path,
        )

    dfp["DEPART_MAG"] = dfp["BE_KEY"].apply(
        lambda be_key: _resolve_depart_mag_date(
            be_key,
            map_depart_mag=map_depart_mag,
            week=week,
            year=year,
        )
    )

    _reset_planning_grid(ws_plan)
    _populate_planning_sheet(
        ws_plan,
        dfp=dfp,
        day_blocks=_PLAN_DAY_BLOCKS,
        keep_rows=_PLAN_KEEP_ROWS,
        safe_excel=safe_excel,
    )
    _apply_planning_layout(ws_plan)

    dfp_safe = sanitize_dataframe_for_excel(dfp)
    df_vols_safe = sanitize_dataframe_for_excel(df_vols)
    df_dispos_safe = sanitize_dataframe_for_excel(df_dispos)

    _write_dataframe_to_sheet(
        ws_export,
        dfp_safe,
        create_tables=False,
        table_name=None,
    )
    _write_dataframe_to_sheet(
        ws_vols,
        df_vols_safe,
        create_tables=create_tables,
        table_name="Table_Vols",
    )

    df_dispo_export = _resolve_benevoles_export_dataframe(
        df_dispos_safe,
        benev_path=benev_path,
    )
    if df_dispo_export is not None and not getattr(df_dispo_export, "empty", True):
        _write_dataframe_to_sheet(
            ws_bene,
            df_dispo_export,
            create_tables=create_tables,
            table_name="Table_Benevoles",
        )
    else:
        _write_dataframe_to_sheet(
            ws_bene,
            None,
            create_tables=False,
            table_name=None,
        )

    wb.save(out_path)
    try:
        load_workbook(out_path)
    except EXCEL_IO_ERRORS as e:
        raise RuntimeError(f"Export Excel invalide : {e}")

    try:
        ws_plan.title = f"Planning S{week:02d}-{year}"
    except ValueError:
        pass

    week_final, year_final = _week_year_from_ws_plan(
        ws_plan,
        fallback_week=week,
        fallback_year=year,
    )

    if skip_versioning:
        _increment_q1_if_requested(ws_plan, increment_version=increment_version)
        wb.save(out_path)
        pdf_path = _export_pdf_with_warning(
            out_path,
            enabled=bool(generate_pdf and (not is_graph_onedrive())),
            pdf_exporter=pdf_exporter,
            warnings=warnings,
        )
        return ExportResult(
            output_path=out_path,
            pdf_path=pdf_path,
            mag_write_method=mag_write_method,
            warnings=warnings,
        )

    planning_dir_final = (
        output_dir
        if output_dir is not None
        else (
            get_output_planning_dir()
            if is_graph_onedrive()
            else get_onedrive_root() / "Planning MAB" / f"ASFmm PLANNING {year_final}"
        )
    )
    planning_dir_final.mkdir(parents=True, exist_ok=True)

    existing_files, max_version = _collect_existing_planning_versions(
        planning_dir_final,
        week_final=week_final,
        year_final=year_final,
    )
    _archive_latest_planning_if_needed(
        increment_version=increment_version,
        max_version=max_version,
        existing_files=existing_files,
        planning_dir_final=planning_dir_final,
        week_final=week_final,
        year_final=year_final,
    )
    version_num, target_path = _resolve_target_planning_path(
        planning_dir_final=planning_dir_final,
        week_final=week_final,
        year_final=year_final,
        max_version=max_version,
        increment_version=increment_version,
    )

    _set_q1_version(ws_plan, version_num)
    out_path = _save_sync_and_move_planning_output(
        wb,
        out_path=out_path,
        target_path=target_path,
        year_final=year_final,
    )

    pdf_path = _export_pdf_with_warning(
        out_path,
        enabled=bool(not is_graph_onedrive()),
        pdf_exporter=pdf_exporter,
        warnings=warnings,
    )

    return ExportResult(
        output_path=out_path,
        pdf_path=pdf_path,
        mag_write_method=mag_write_method,
        warnings=warnings,
    )

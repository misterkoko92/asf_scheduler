# -*- coding: utf-8 -*-
from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Optional

import pandas as pd

from utils.datetime_utils import format_date_value, format_time_value

from loaders.universal_loader import load_and_normalize
from loaders.load_vols import load_vols_df
from loaders.load_params import load_param_dest_from_path
from scheduler.column_map import (
    column_map_mag_central,
    column_map_param_be,
    column_map_param_dest,
    column_map_param_benev,
    column_map_benev_dispo,
    column_map_vols,
)
from scheduler.config_paths import (
    SHEET_MAG_CENTRAL,
    SHEET_PARAM_BE,
    SHEET_PARAM_DEST,
    SHEET_PARAM_BENEV,
    SHEET_BENEV_DISPO,
    SHEET_VOLS,
)
from utils.logging_utils import get_logger


class InputLoadError(RuntimeError):
    pass


@dataclass(frozen=True)
class TdbData:
    df_be: pd.DataFrame
    df_param_be: pd.DataFrame
    df_param_dest: pd.DataFrame


@dataclass(frozen=True)
class BenevData:
    df_param_benev: pd.DataFrame
    df_benev: pd.DataFrame


logger = get_logger("input_service", console=False)


def load_tdb(path: Path) -> TdbData:
    if path is None or not Path(path).exists():
        raise FileNotFoundError("TABLEAU DE BORD introuvable")
    df_mag = load_and_normalize(path, SHEET_MAG_CENTRAL, column_map_mag_central, header=5)
    df_param_be = load_and_normalize(path, SHEET_PARAM_BE, column_map_param_be, header=0)
    df_param_dest = load_and_normalize(path, SHEET_PARAM_DEST, column_map_param_dest, header=0)
    return TdbData(df_be=df_mag, df_param_be=df_param_be, df_param_dest=df_param_dest)


def load_benev(path: Path) -> BenevData:
    if path is None or not Path(path).exists():
        raise FileNotFoundError("Planning Bénévoles introuvable")
    df_param_benev = load_and_normalize(path, SHEET_PARAM_BENEV, column_map_param_benev, header=0)
    df_dispo = load_and_normalize(path, SHEET_BENEV_DISPO, column_map_benev_dispo, header=0)
    return BenevData(df_param_benev=df_param_benev, df_benev=df_dispo)


def load_vols(
    vols_path: Path,
    *,
    param_dest_df: Optional[pd.DataFrame] = None,
    tdb_path: Optional[Path] = None,
) -> pd.DataFrame:
    if vols_path is None or not Path(vols_path).exists():
        raise FileNotFoundError("Vols introuvable")

    if param_dest_df is None and tdb_path is not None and Path(tdb_path).exists():
        try:
            param_dest_df = load_param_dest_from_path(Path(tdb_path))
        except Exception:
            param_dest_df = None

    try:
        return load_vols_df(vols_path=Path(vols_path), param_dest_df=param_dest_df)
    except Exception as exc:
        logger.error("Erreur load_vols_df: %s", exc)
        try:
            return load_and_normalize(vols_path, SHEET_VOLS, column_map_vols, header=0)
        except Exception as exc2:
            raise InputLoadError("Erreur chargement Vols") from exc2


def load_normalized_sheet(
    path: Path,
    sheet_name: str,
    mapping: dict,
    header: int = 0,
) -> pd.DataFrame:
    return load_and_normalize(path=path, sheet_name=sheet_name, mapping=mapping, header=header)


def get_benev_source_message(path: Path) -> str:
    """
    Lit D2 (date) et E2 (heure) dans la feuille 'Source' du planning bénévoles.
    Retourne une chaîne "DD/MM/YY à HHhMM" ou "N/A" si non disponible.
    """
    try:
        from openpyxl import load_workbook

        if path is None or not Path(path).exists():
            return "N/A"
        wb = load_workbook(path, data_only=True)
        if "Source" not in wb.sheetnames:
            return "N/A"
        ws = wb["Source"]
        d = ws["D2"].value
        h = ws["E2"].value

        d_str = format_date_value(d, fmt="%d/%m/%y", default="")
        h_str = format_time_value(h, fmt="%Hh%M", default="")
        if d_str or h_str:
            return f"{d_str} à {h_str}".strip(" à ")
        return "N/A"
    except Exception:
        return "N/A"

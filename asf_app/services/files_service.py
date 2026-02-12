# asf_app/services/files_service.py
# -*- coding: utf-8 -*-

import os
import datetime
from pathlib import Path
from typing import Dict

import pandas as pd
import scheduler.config_paths as cp
from utils.datetime_utils import format_date_value
from utils.excel_safety import sanitize_excel_value


# ============================================================
# Horodatage fichier
# ============================================================

def pretty_mtime(path_str: str) -> str:
    try:
        ts = os.path.getmtime(path_str)
        dt = datetime.datetime.fromtimestamp(ts)
        return format_date_value(dt, fmt="%d/%m/%Y à %H:%M", default="N/A")
    except Exception:
        return "N/A"


# ============================================================
# Lecture Excel robuste
# ============================================================

def read_excel_sheet(path: str | Path, sheet_name: str, dtype=str) -> pd.DataFrame:
    """
    Lecture robuste d’un onglet Excel.
    - dtype=str → homogénéité
    - NaN → ""
    - support xlsx uniquement
    """
    return (
        pd.read_excel(path, sheet_name=sheet_name, dtype=dtype, engine="openpyxl")
        .fillna("")
    )


# ============================================================
# Sauvegarde Excel SAFE : écrase proprement l’onglet
# ============================================================

def save_excel_sheet(path: str | Path, sheet_name: str, df: pd.DataFrame) -> None:
    """
    Sauvegarde un onglet en remplaçant son contenu.
    - Excel automation si dispo pour préserver validations + mises en forme
    - Fallback openpyxl in-place (sans supprimer la feuille)
    - Ne détruit pas les autres onglets
    """
    path = Path(path)

    # Si le fichier n'existe pas → on crée un workbook
    if not path.exists():
        with pd.ExcelWriter(path, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)
        cp.sync_local_file_to_onedrive(path)
        return

    df_clean = df.where(pd.notna(df), "")
    table = [[sanitize_excel_value(c) for c in df_clean.columns]]
    table.extend(
        [
            [sanitize_excel_value(v) for v in row]
            for row in df_clean.itertuples(index=False, name=None)
        ]
    )

    # Sinon → tentative Excel automation pour préserver validations + mises en forme
    try:
        from utils.excel_automation import write_sheet_table

        if write_sheet_table(path, sheet_name, table):
            cp.sync_local_file_to_onedrive(path)
            return
    except Exception:
        pass

    # Fallback openpyxl : réécriture in-place (sans supprimer la feuille)
    from openpyxl import load_workbook

    wb = load_workbook(path)
    ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.create_sheet(sheet_name)

    max_row = max(ws.max_row, len(table))
    max_col = max(ws.max_column, len(table[0]) if table else 0)
    if max_row and max_col:
        for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
            for cell in row:
                cell.value = None

    for r_idx, row in enumerate(table, start=1):
        for c_idx, val in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=val)

    wb.save(path)
    cp.sync_local_file_to_onedrive(path)


# ============================================================
# Ajouter une ligne à un onglet (respect colonnes)
# ============================================================

def append_row_to_sheet(path: str | Path, sheet_name: str, new_row: Dict[str, str]) -> None:
    df = read_excel_sheet(path, sheet_name)
    cols = df.columns.tolist()

    row = {c: "" for c in cols}
    row.update({k: v for k, v in new_row.items() if k in row})

    df2 = pd.concat([df, pd.DataFrame([row])], ignore_index=True)
    save_excel_sheet(path, sheet_name, df2)

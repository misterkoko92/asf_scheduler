# -*- coding: utf-8 -*-
"""
generate_from_template.py — Génération du planning via maquette Excel.
"""

import shutil
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from scheduler.config_paths import (
    PLANNING_TEMPLATE,
    OUTPUT_PLANNING_DIR,
)


# Colonnes attendues dans la feuille "Data"
DATA_HEADERS = [
    "DATE",
    "ITEM",
    "DATE LONGUE",
    "NOM",
    "OK",
    "DESTINATION",
    "IATA",
    "ROUTING",
    "NUMERO VOL",
    "HEURE VOL",
    "NUMERO BE",
    "NOMBRE COLIS",
    "TYPE",
    "LIVRAISON COLIS",
    "DATE TRANSFERT",
    "EXPEDITEUR",
    "DESTINATAIRE",
]


def generate_planning_from_template(df, week:int, year:int):
    """
    Génère un planning en dupliquant la maquette Excel.

    Paramètres
    ----------
    df   : DataFrame préparé contenant toutes les colonnes nécessaires.
    week : numéro de semaine (int)
    year : année concernée (ex: 2025)
    """

    # -------------------------------------------------------------------------
    # 1) Fichier de sortie
    # -------------------------------------------------------------------------
    filename = f"ASFmm - PLANNING SEMAINE N° {week:02d} - {year}.xlsx"
    output_path = OUTPUT_PLANNING_DIR / filename

    # -------------------------------------------------------------------------
    # 2) Copier la maquette dans le fichier final
    # -------------------------------------------------------------------------
    shutil.copy(PLANNING_TEMPLATE, output_path)

    # -------------------------------------------------------------------------
    # 3) Charger le fichier final
    # -------------------------------------------------------------------------
    wb = load_workbook(output_path)
    if "Data" not in wb.sheetnames:
        raise RuntimeError("La feuille 'Data' est absente de la maquette Excel.")

    ws_data = wb["Data"]

    # -------------------------------------------------------------------------
    # 4) Nettoyage de l'onglet Data (sauf headers ligne 1)
    # -------------------------------------------------------------------------
    for row in ws_data.iter_rows(min_row=2, max_row=ws_data.max_row, min_col=1, max_col=len(DATA_HEADERS)):
        for cell in row:
            cell.value = None

    # -------------------------------------------------------------------------
    # 5) Remplissage
    # -------------------------------------------------------------------------
    row_idx = 2
    for _, line in df.iterrows():

        values = [
            line.get("DATE"),              # A
            line.get("ITEM"),
            line.get("DATE_LONGUE"),
            line.get("NOM"),
            "",                             # OK (vide)
            line.get("DESTINATION"),
            line.get("IATA"),
            line.get("ROUTING"),
            line.get("NUMERO_VOL"),
            line.get("HEURE_VOL"),
            line.get("NUMERO_BE"),
            line.get("NOMBRE_COLIS"),
            line.get("TYPE"),
            "MAG EXPORT",                   # LIVRAISON COLIS
            "",                             # DATE TRANSFERT (vide)
            line.get("EXPEDITEUR"),
            line.get("DESTINATAIRE"),
        ]

        for col_idx, val in enumerate(values, 1):
            ws_data.cell(row=row_idx, column=col_idx).value = val

        row_idx += 1

    # -------------------------------------------------------------------------
    # 6) Renommer la feuille "Planning SXX"
    # -------------------------------------------------------------------------
    target_name = f"Planning S{week:02d}"

    if "Planning SXX" in wb.sheetnames:
        ws = wb["Planning SXX"]
        ws.title = target_name
    else:
        # Si tu changes le nom dans la maquette, il faut adapter ici
        raise RuntimeError("La feuille 'Planning SXX' est absente de la maquette.")

    # -------------------------------------------------------------------------
    # 7) Sauvegarde
    # -------------------------------------------------------------------------
    wb.save(output_path)

    return output_path

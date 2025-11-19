# asf_app/ui_inputs.py
# -*- coding: utf-8 -*-

import os
import re
import tempfile
import pandas as pd
import streamlit as st
from datetime import datetime
from openpyxl import load_workbook

from scheduler.config_paths import (
    TABLEAU_DE_BORD,
    PLANNING_BENEVOLES,
    VOLS,
    SHEET_MAG_CENTRAL,
)

from loaders.load_shipments import load_shipments
from scheduler import be_manager


# =====================================================================
# UTILITAIRES
# =====================================================================

def pretty_mtime(path_str: str) -> str:
    try:
        ts = os.path.getmtime(path_str)
        dt = datetime.fromtimestamp(ts)
        return dt.strftime("%d/%m/%Y à %H:%M")
    except Exception:
        return "N/A"


def extract_integer(value):
    try:
        return str(int(float(value)))
    except:
        return str(value)


def find_column(df, possible):
    def norm(s):
        return re.sub(r"[^a-z0-9]", "", str(s).lower())

    norm_cols = {norm(c): c for c in df.columns}

    for name in possible:
        target = norm(name)
        for nc, original in norm_cols.items():
            if target in nc or nc in target:
                return original
    return None


def read_excel_cell(path, sheet, cell):
    try:
        wb = load_workbook(path, data_only=True)
        ws = wb[sheet]
        return ws[cell].value
    except Exception:
        return None


def excel_time_to_string(x):
    if isinstance(x, datetime):
        return x.strftime("%Hh%M")

    if hasattr(x, "hour"):
        try:
            return f"{x.hour:02d}h{x.minute:02d}"
        except:
            pass

    if isinstance(x, str):
        m = re.match(r"^\s*(\d{1,2}):(\d{2})", x)
        if m:
            h = int(m.group(1))
            mn = int(m.group(2))
            return f"{h:02d}h{mn:02d}"

    try:
        val = float(x)
        if 0 <= val < 1:
            total_minutes = int(round(val * 24 * 60))
            h = total_minutes // 60
            mn = total_minutes % 60
            return f"{h:02d}h{mn:02d}"
    except:
        pass

    return "00h00"


# =============================================================================
# ONGLET FICHIERS D’ENTRÉE (NOUVELLE MISE EN FORME)
# =============================================================================

def render_tab_inputs():

    st.header("📁 Fichiers d’entrée")

    # ==========================================================================
    # 1) TABLEAU DE BORD
    # ==========================================================================
    st.subheader("📘 Tableau de bord")

    path_tdb = st.session_state.paths["tdb"]

    # ---- Fichier utilisé
    st.write(f"**Fichier utilisé :** `{os.path.basename(path_tdb)}`")

    # ---- Feuilles réellement utilisées (et non toutes les feuilles)
    st.write("**📄 Feuilles utilisées :** MAG CENTRAL, ParamBE, ParamDest")

    # ---- Dernière modification
    st.write(f"🕒 Dernière modification : {pretty_mtime(path_tdb)}")



    # ---------------------------
    # 📦 Analyse via moteur ASF
    # ---------------------------
    try:
        be_raw = load_shipments()
        be_filtered = be_manager.filter_shipments(be_raw)
        be_sorted   = be_manager.sort_shipments(be_filtered)

        # ---- BE prêts
        nb_ready = len(be_sorted)

        if nb_ready > 0:
            dest_counts = {}
            for s in be_sorted:
                dest_counts[s.dest] = dest_counts.get(s.dest, 0) + 1

            dest_str = ", ".join([f"{d} ({c})" for d, c in dest_counts.items()])
            st.write(f"📦 **{nb_ready} BE à planifier : {dest_str}**")
        else:
            st.write("📦 Aucun BE prêt à être planifié.")

        # ---- Destinations uniques
        unique_dest = sorted({s.dest for s in be_raw if s.dest})
        st.write(f"🌍 **{len(unique_dest)} destinations trouvées : {', '.join(unique_dest)}**")

        # ---- Dernier BE modifié
        if be_raw and hasattr(be_raw[0], "_origin_df"):
            df_src = be_raw[0]._origin_df
        else:
            df_src = None

        if df_src is not None:
            possible_cols_mod = [
                "Dernière modification",
                "DATE_MODIF",
                "DATE MISE A JOUR",
                "LAST_UPDATE",
                "MODIFICATION",
            ]
            col_mod = find_column(df_src, possible_cols_mod)

            if col_mod:
                df_mod = df_src[df_src[col_mod] != ""]
                if not df_mod.empty:
                    last = df_mod.sort_values(col_mod, ascending=False).iloc[0]
                    be_clean = extract_integer(last.get("BE_NUMERO", ""))
                    st.write(f"🔧 **Dernier BE modifié : {be_clean}**")

    except Exception as e:
        st.error(f"❌ Erreur lors de l’analyse via moteur ASF : {e}")

        # ---- Upload fichier
    tdb_file = st.file_uploader(
        "Importer un nouveau TABLEAU DE BORD.xlsx",
        type=["xlsx"],
        key="upload_tdb"
    )
    if tdb_file is not None:
        tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx").name
        with open(tmp, "wb") as f:
            f.write(tdb_file.read())
        st.session_state.paths["tdb"] = tmp
        path_tdb = tmp
        st.success("TABLEAU DE BORD mis à jour pour la session.")

    st.markdown("---")

    # ==========================================================================
    # 2) PLANNING BÉNÉVOLES
    # ==========================================================================
    st.subheader("👥 Planning bénévoles")

    path_benev = st.session_state.paths["benev"]

    # ---- Fichier utilisé
    st.write(f"**Fichier utilisé :** `{os.path.basename(path_benev)}`")

    # ---- Feuilles utilisées
    st.write("**📄 Feuilles utilisées :** Source, ParamBenev, Disponibilités")

    # ---- Dernière modification
    st.write(f"🕒 Dernière modification : {pretty_mtime(path_benev)}")

    # ---- Dernier message (D2/E2)
    try:
        df_src = pd.read_excel(path_benev, sheet_name="Source", header=None)

        # Lecture directe de D2 et E2
        date_raw = df_src.iloc[1, 3]  # D2
        heure_raw = df_src.iloc[1, 4] # E2

        # --- DATE ---
        date_norm = pd.to_datetime(str(date_raw), errors="coerce")
        date_str = (
            date_norm.strftime("%Y-%m-%d")
            if date_norm is not pd.NaT
            else str(date_raw)
        )

        # --- HEURE (ancienne logique) ---
        heure_norm = pd.to_datetime(str(heure_raw), errors="coerce")
        heure_str = (
            heure_norm.strftime("%Hh%M")
            if heure_norm is not pd.NaT
            else "00h00"
        )

        st.write(f"📩 **Dernier message pris en compte : {date_str} à {heure_str}**")

    except Exception as e:
        st.error(f"❌ Erreur lecture Source bénévoles : {e}")





    # ---- Upload
    benev_file = st.file_uploader(
        "Importer Planning Bénévoles.xlsx",
        type=["xlsx"],
        key="upload_benev"
    )
    if benev_file is not None:
        tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx").name
        with open(tmp, "wb") as f:
            f.write(benev_file.read())
        st.session_state.paths["benev"] = tmp
        path_benev = tmp
        st.success("Planning bénévoles mis à jour.")

    st.markdown("---")

    # ==========================================================================
    # 3) FICHIER VOLS
    # ==========================================================================
    st.subheader("✈️ Vols")

    path_vols = st.session_state.paths["vols"]

    # ---- Fichier utilisé
    st.write(f"**Fichier utilisé :** `{os.path.basename(path_vols)}`")

    # ---- Feuilles réellement utilisées
    st.write("**📄 Feuille utilisée :** Feuille principale des vols")

    # ---- Dernière modification
    st.write(f"🕒 Dernière modification : {pretty_mtime(path_vols)}")

    # ---- Semaine & période
    try:
        df_vols = pd.read_excel(path_vols, dtype=str).fillna("")

        if "PVOL_DATE" in df_vols.columns:
            dates = pd.to_datetime(df_vols["PVOL_DATE"], errors="coerce").dropna()
            if not dates.empty:
                d_min, d_max = dates.min(), dates.max()
                week = d_min.isocalendar()[1]

                st.write(
                    f"🗓️ **Vols de la Semaine {week} / "
                    f"du {d_min.strftime('%d/%m/%Y')} au {d_max.strftime('%d/%m/%Y')}**"
                )

    except Exception as e:
        st.error(f"❌ Erreur lecture période des vols : {e}")

    # ---- Upload
    vols_file = st.file_uploader(
        "Importer Vols.xlsx",
        type=["xlsx"],
        key="upload_vols"
    )
    if vols_file is not None:
        tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx").name
        with open(tmp, "wb") as f:
            f.write(vols_file.read())
        st.session_state.paths["vols"] = tmp
        path_vols = tmp
        st.success("Fichier vols mis à jour.")

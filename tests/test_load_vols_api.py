# -*- coding: utf-8 -*-
from __future__ import annotations

import logging
from datetime import date
from types import SimpleNamespace

import pandas as pd
from openpyxl import load_workbook

import loaders.load_vols_api as load_vols_api_mod


def _paramdest_df() -> pd.DataFrame:
    return pd.DataFrame(
        [
            {"Dest_IATA": "DLA", "Dest_Ville": "Douala", "Max_Colis_Par_Vol": 12},
        ]
    )


def test_load_vols_api_passes_time_origin_type(monkeypatch):
    captured: dict = {}
    monkeypatch.setattr(load_vols_api_mod, "load_paramdest_codes", lambda: _paramdest_df())
    monkeypatch.setattr(load_vols_api_mod, "load_be_dest_codes", lambda: [])

    def _fake_fetch_multiple(**kwargs):
        captured.update(kwargs)
        return []

    monkeypatch.setattr(load_vols_api_mod, "fetch_multiple", _fake_fetch_multiple)

    df = load_vols_api_mod.load_vols_api(
        date(2026, 1, 23),
        date(2026, 1, 23),
        time_origin_type="S",
    )

    assert captured["time_origin_type"] == "S"
    assert isinstance(df, pd.DataFrame)
    assert df.empty


def test_load_paramdest_codes_delegates_to_loader(monkeypatch):
    captured: dict[str, object] = {}

    def _fake_load(*args, **kwargs):
        captured["args"] = args
        captured["kwargs"] = kwargs
        return pd.DataFrame([{"Dest_IATA": "DLA"}])

    monkeypatch.setattr(load_vols_api_mod, "load_and_normalize", _fake_load)

    out = load_vols_api_mod.load_paramdest_codes()
    assert len(out) == 1
    assert captured["kwargs"]["header"] == 0


def test_load_vols_api_maps_flights_to_dataframe(monkeypatch, caplog):
    monkeypatch.setattr(load_vols_api_mod, "load_paramdest_codes", lambda: _paramdest_df())
    monkeypatch.setattr(load_vols_api_mod, "load_be_dest_codes", lambda: ["DLA"])
    monkeypatch.setattr(
        load_vols_api_mod,
        "fetch_multiple",
        lambda **kwargs: [
            SimpleNamespace(
                route="CDG-DLA",
                date_depart="23/01/26",
                heure_depart="21h01",
                numero_vol="AF 652",
            )
        ],
    )

    with caplog.at_level(logging.INFO, logger="ASF-SCHEDULER"):
        df = load_vols_api_mod.load_vols_api(date(2026, 1, 23), date(2026, 1, 23))

    assert len(df) == 1
    assert df.loc[0, "IATA"] == "DLA"
    assert df.loc[0, "Destination"] == "DOUALA"
    assert df.loc[0, "Numero_Vol"] == "AF 652"
    assert df.loc[0, "Heure_Vol"] == "21h01"
    assert int(df.loc[0, "HEURE_MIN"]) == 21 * 60 + 1
    assert any("[AF API] Vols charges: 1" in msg for msg in caplog.messages)


def test_load_be_dest_codes_returns_empty_on_loader_error(monkeypatch):
    monkeypatch.setattr(
        load_vols_api_mod,
        "load_shipments_df",
        lambda *args, **kwargs: (_ for _ in ()).throw(RuntimeError("boom")),
    )
    assert load_vols_api_mod.load_be_dest_codes() == []


def test_load_be_dest_codes_reads_both_dest_iata_and_destination(monkeypatch):
    monkeypatch.setattr(
        load_vols_api_mod,
        "load_shipments_df",
        lambda *args, **kwargs: pd.DataFrame(
            [{"Dest_IATA": "dla"}, {"Dest_IATA": "RUN"}, {"Dest_IATA": "XX"}]
        ),
    )
    assert load_vols_api_mod.load_be_dest_codes() == ["DLA", "RUN"]

    monkeypatch.setattr(
        load_vols_api_mod,
        "load_shipments_df",
        lambda *args, **kwargs: pd.DataFrame(
            [{"Destination": "dla"}, {"Destination": "run"}, {"Destination": "ABCD"}]
        ),
    )
    assert load_vols_api_mod.load_be_dest_codes() == ["DLA", "RUN"]


def test_load_vols_api_splits_routes_and_deduplicates(monkeypatch):
    monkeypatch.setattr(
        load_vols_api_mod,
        "load_paramdest_codes",
        lambda: pd.DataFrame(
            [
                {"Dest_IATA": "DLA", "Dest_Ville": "Douala", "Max_Colis_Par_Vol": 12},
                {"Dest_IATA": "RUN", "Dest_Ville": "Saint-Denis", "Max_Colis_Par_Vol": 24},
            ]
        ),
    )
    monkeypatch.setattr(load_vols_api_mod, "load_be_dest_codes", lambda: [])
    monkeypatch.setattr(
        load_vols_api_mod,
        "fetch_multiple",
        lambda **kwargs: [
            SimpleNamespace(
                route="CDG-DLA-RUN",
                date_depart="2026-01-23",
                heure_depart="21h01",
                numero_vol="AF0652",
            ),
            SimpleNamespace(
                route="CDG-DLA-RUN",
                date_depart="2026-01-23",
                heure_depart="21h01",
                numero_vol="AF0652",
            ),
        ],
    )

    df = load_vols_api_mod.load_vols_api(date(2026, 1, 23), date(2026, 1, 23))

    assert sorted(df["IATA"].tolist()) == ["DLA", "RUN"]
    assert set(df["Origine"]) == {"CDG"}
    assert set(df["Max_Colis"].astype("Int64").tolist()) == {12, 24}


def test_store_vols_api_sheet_creates_workbook_when_missing(tmp_path, monkeypatch):
    target = tmp_path / "Vols.xlsx"
    df = pd.DataFrame(
        [{"Date_Vol": "23/01/26", "Heure_Vol": "21h01", "Numero_Vol": "AF0652", "Destination": "DOUALA"}]
    )
    synced: list[str] = []
    monkeypatch.setattr(load_vols_api_mod.cp, "sync_local_file_to_onedrive", lambda p: synced.append(str(p)))

    sheet_name = load_vols_api_mod.store_vols_api_sheet(df, date(2026, 1, 23), path=target)

    assert sheet_name == "API-S04-2026"
    assert target.exists()
    wb = load_workbook(target)
    assert sheet_name in wb.sheetnames
    assert wb[sheet_name].cell(1, 1).value == "Date_Vol"
    assert synced == [str(target)]


def test_store_vols_api_sheet_uses_write_sheet_table_when_available(tmp_path, monkeypatch):
    target = tmp_path / "Vols.xlsx"
    pd.DataFrame([{"x": 1}]).to_excel(target, index=False)
    df = pd.DataFrame([{"Date_Vol": "23/01/26"}])
    called: dict[str, object] = {}
    synced: list[str] = []

    import utils.excel_automation as excel_auto

    monkeypatch.setattr(
        excel_auto,
        "write_sheet_table",
        lambda path, sheet_name, table_rows: called.update(
            {"path": path, "sheet_name": sheet_name, "rows": table_rows}
        )
        or True,
    )
    monkeypatch.setattr(load_vols_api_mod.cp, "sync_local_file_to_onedrive", lambda p: synced.append(str(p)))

    sheet_name = load_vols_api_mod.store_vols_api_sheet(df, date(2026, 1, 23), path=target)

    assert sheet_name == "API-S04-2026"
    assert called["path"] == target
    assert called["sheet_name"] == "API-S04-2026"
    assert isinstance(called["rows"], list)
    assert synced == [str(target)]


def test_copy_api_sheet_to_tmp_copies_values_and_syncs(tmp_path, monkeypatch):
    src = tmp_path / "src.xlsx"
    dst = tmp_path / "dst.xlsx"
    sheet_name = "API-S04-2026"
    pd.DataFrame([{"A": 1, "B": "x"}]).to_excel(src, sheet_name=sheet_name, index=False)
    synced: list[str] = []
    monkeypatch.setattr(load_vols_api_mod.cp, "sync_local_file_to_onedrive", lambda p: synced.append(str(p)))

    load_vols_api_mod.copy_api_sheet_to_tmp(sheet_name=sheet_name, src_path=src, dst_path=dst)

    assert dst.exists()
    wb = load_workbook(dst)
    assert sheet_name in wb.sheetnames
    assert wb[sheet_name].cell(1, 1).value == "A"
    assert wb[sheet_name].cell(2, 1).value == 1
    assert synced == [str(dst)]


def test_fmt_date_handles_invalid_and_valid_values():
    assert load_vols_api_mod._fmt_date("23/01/2026") == "23/01/26"
    assert load_vols_api_mod._fmt_date(None) == ""


def test_load_vols_api_empty_paramdest_returns_empty(monkeypatch):
    monkeypatch.setattr(load_vols_api_mod, "load_paramdest_codes", lambda: pd.DataFrame())
    monkeypatch.setattr(load_vols_api_mod, "load_be_dest_codes", lambda: ["DLA"])

    df = load_vols_api_mod.load_vols_api(date(2026, 1, 23), date(2026, 1, 23))

    assert isinstance(df, pd.DataFrame)
    assert df.empty


def test_load_vols_api_ignores_invalid_routes(monkeypatch):
    monkeypatch.setattr(load_vols_api_mod, "load_paramdest_codes", lambda: _paramdest_df())
    monkeypatch.setattr(load_vols_api_mod, "load_be_dest_codes", lambda: [])
    monkeypatch.setattr(
        load_vols_api_mod,
        "fetch_multiple",
        lambda **kwargs: [
            SimpleNamespace(route="", date_depart="23/01/26", heure_depart="21h01", numero_vol="AF 652"),
            SimpleNamespace(route="CDG", date_depart="23/01/26", heure_depart="21h01", numero_vol="AF 653"),
        ],
    )

    df = load_vols_api_mod.load_vols_api(date(2026, 1, 23), date(2026, 1, 23))

    assert df.empty


def test_load_vols_api_swallow_logging_errors(monkeypatch):
    monkeypatch.setattr(load_vols_api_mod, "load_paramdest_codes", lambda: _paramdest_df())
    monkeypatch.setattr(load_vols_api_mod, "load_be_dest_codes", lambda: [])
    monkeypatch.setattr(
        load_vols_api_mod,
        "fetch_multiple",
        lambda **kwargs: [
            SimpleNamespace(route="CDG-DLA", date_depart="23/01/26", heure_depart="21h01", numero_vol="AF 652"),
        ],
    )
    monkeypatch.setattr(load_vols_api_mod.logger, "info", lambda *_a, **_k: (_ for _ in ()).throw(TypeError("boom")))

    out = load_vols_api_mod.load_vols_api(date(2026, 1, 23), date(2026, 1, 23))
    assert len(out) == 1


def test_store_vols_api_sheet_returns_empty_for_none_dataframe(tmp_path):
    out = load_vols_api_mod.store_vols_api_sheet(None, date(2026, 1, 23), path=tmp_path / "Vols.xlsx")
    assert out == ""


def test_store_vols_api_sheet_with_astype_failure_keeps_working(tmp_path, monkeypatch):
    target = tmp_path / "Vols.xlsx"
    pd.DataFrame([{"x": 1}]).to_excel(target, index=False)
    df = pd.DataFrame([{"Date_Vol": "23/01/26"}])
    monkeypatch.setattr(load_vols_api_mod.cp, "sync_local_file_to_onedrive", lambda _p: None)
    monkeypatch.setattr(
        pd.DataFrame,
        "astype",
        lambda *_a, **_k: (_ for _ in ()).throw(TypeError("astype-fail")),
    )

    sheet_name = load_vols_api_mod.store_vols_api_sheet(df, date(2026, 1, 23), path=target)
    assert sheet_name == "API-S04-2026"


def test_store_vols_api_sheet_falls_back_to_openpyxl_when_write_sheet_table_returns_false(tmp_path, monkeypatch):
    target = tmp_path / "Vols.xlsx"
    pd.DataFrame([{"x": 1}]).to_excel(target, index=False)
    df = pd.DataFrame([{"Date_Vol": "23/01/26", "Heure_Vol": "21h01", "Numero_Vol": "AF0652"}])
    synced: list[str] = []

    import utils.excel_automation as excel_auto

    monkeypatch.setattr(excel_auto, "write_sheet_table", lambda *args, **kwargs: False)
    monkeypatch.setattr(load_vols_api_mod.cp, "sync_local_file_to_onedrive", lambda p: synced.append(str(p)))

    sheet_name = load_vols_api_mod.store_vols_api_sheet(df, date(2026, 1, 23), path=target)

    wb = load_workbook(target)
    assert sheet_name == "API-S04-2026"
    assert sheet_name in wb.sheetnames
    assert wb[sheet_name].cell(1, 1).value == "Date_Vol"
    assert synced == [str(target)]


def test_copy_api_sheet_to_tmp_skips_when_source_missing(tmp_path, monkeypatch):
    dst = tmp_path / "dst.xlsx"
    synced: list[str] = []
    monkeypatch.setattr(load_vols_api_mod.cp, "sync_local_file_to_onedrive", lambda p: synced.append(str(p)))

    load_vols_api_mod.copy_api_sheet_to_tmp(
        sheet_name="API-S04-2026",
        src_path=tmp_path / "missing.xlsx",
        dst_path=dst,
    )

    assert not dst.exists()
    assert synced == []


def test_copy_api_sheet_to_tmp_returns_when_empty_sheet_name_or_missing_sheet(tmp_path, monkeypatch):
    src = tmp_path / "src.xlsx"
    pd.DataFrame([{"A": 1}]).to_excel(src, sheet_name="Other", index=False)
    dst = tmp_path / "dst.xlsx"
    synced: list[str] = []
    monkeypatch.setattr(load_vols_api_mod.cp, "sync_local_file_to_onedrive", lambda p: synced.append(str(p)))

    load_vols_api_mod.copy_api_sheet_to_tmp("", src_path=src, dst_path=dst)
    load_vols_api_mod.copy_api_sheet_to_tmp("API-S04-2026", src_path=src, dst_path=dst)

    assert not dst.exists()
    assert synced == []


def test_copy_api_sheet_to_tmp_swallow_exception(monkeypatch, tmp_path):
    monkeypatch.setattr(load_vols_api_mod, "load_workbook", lambda *_a, **_k: (_ for _ in ()).throw(OSError("boom")))
    load_vols_api_mod.copy_api_sheet_to_tmp(
        sheet_name="API-S04-2026",
        src_path=tmp_path / "src.xlsx",
        dst_path=tmp_path / "dst.xlsx",
    )

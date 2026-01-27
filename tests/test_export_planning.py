# -*- coding: utf-8 -*-
from __future__ import annotations

from datetime import date
from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook

import scheduler.config_paths as cp
from asf_app.services.export_service import export_planning_excel


def _write_template(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Planning SXX"
    wb.save(path)


def _write_mag_central(path: Path, be_value: str, sheet_name: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.cell(row=1, column=1, value=be_value)
    wb.save(path)


def _planning_df(be_num: str, vol_num: str, dest: str, when: date) -> pd.DataFrame:
    return pd.DataFrame(
        [
            {
                "Date_Vol": when,
                "Heure_Vol": "10:00",
                "Numero_Vol": vol_num,
                "Destination": dest,
                "BE_Numero": be_num,
                "BE_Nb_Colis": 2,
                "BE_Nb_Equiv": 2,
                "BE_Type": "MM",
                "BE_Expediteur": "ASF",
                "BE_Destinataire": "Hopital",
                "Benevole": "DUPONT",
                "ID": "1",
            }
        ]
    )


def _param_dest_df(dest: str, city: str) -> pd.DataFrame:
    return pd.DataFrame(
        [
            {"Dest_IATA": dest, "Dest_Ville": city},
        ]
    )


def _vols_df(dest: str, vol_num: str, when: date) -> pd.DataFrame:
    return pd.DataFrame(
        [
            {
                "Date_Vol": when.strftime("%d/%m/%y"),
                "Numero_Vol": vol_num,
                "Destination": dest,
                "IATA": dest,
                "Routing": f"CDG-{dest}",
            }
        ]
    )


def test_export_planning_increments_and_updates_mag(tmp_path, monkeypatch):
    monkeypatch.setattr(cp, "USE_GRAPH_ONEDRIVE", False, raising=False)
    onedrive_root = tmp_path / "onedrive"
    monkeypatch.setattr(cp, "ASF_ONEDRIVE", onedrive_root)

    maquette_dir = onedrive_root / "Planning MAB" / "ASFmm PLANNING 2025" / "aaSOURCE"
    maquette_dir.mkdir(parents=True, exist_ok=True)
    template_path = maquette_dir / "Planning-maquette.xlsx"
    _write_template(template_path)

    tdb_path = tmp_path / "TABLEAU_DE_BORD.xlsx"
    _write_mag_central(tdb_path, "260001", "MAG CENTRAL 2026")

    monkeypatch.setattr(cp, "sync_local_file_to_onedrive", lambda *args, **kwargs: True)

    week = 1
    year = 2026
    monday = date.fromisocalendar(year, week, 1)
    df_plan = _planning_df("260001", "1234", "DLA", monday)

    result = export_planning_excel(
        df_plan,
        week,
        year,
        df_vols=_vols_df("DLA", "1234", monday),
        df_paramdest=_param_dest_df("DLA", "DOUALA"),
        df_dispos=pd.DataFrame(),
        write_source_excel=True,
        increment_version=True,
        tdb_source_path=tdb_path,
        pdf_exporter=lambda src, dst: dst,
    )
    out_path = result.output_path

    assert out_path.exists()
    assert out_path.name.startswith("ASFmm - PLANNING SEMAINE 2026-01-01")

    wb = load_workbook(tdb_path)
    ws = wb["MAG CENTRAL 2026"]
    assert ws.cell(row=1, column=cp.MAG_CENTRAL_COL_DEPART_VOL).value is not None
    assert ws.cell(row=1, column=cp.MAG_CENTRAL_COL_DEPART_MAG).value is not None
    assert ws.cell(row=1, column=cp.MAG_CENTRAL_COL_VOL).value is not None
    assert ws.cell(row=1, column=cp.MAG_CENTRAL_COL_BENEV).value is not None


def test_export_planning_archives_when_no_increment(tmp_path, monkeypatch):
    monkeypatch.setattr(cp, "USE_GRAPH_ONEDRIVE", False, raising=False)
    onedrive_root = tmp_path / "onedrive"
    monkeypatch.setattr(cp, "ASF_ONEDRIVE", onedrive_root)

    maquette_dir = onedrive_root / "Planning MAB" / "ASFmm PLANNING 2025" / "aaSOURCE"
    maquette_dir.mkdir(parents=True, exist_ok=True)
    template_path = maquette_dir / "Planning-maquette.xlsx"
    _write_template(template_path)

    monkeypatch.setattr(cp, "sync_local_file_to_onedrive", lambda *args, **kwargs: True)

    week = 2
    year = 2026
    planning_dir = onedrive_root / "Planning MAB" / f"ASFmm PLANNING {year}"
    planning_dir.mkdir(parents=True, exist_ok=True)
    existing = planning_dir / "ASFmm - PLANNING SEMAINE 2026-02-01.xlsx"
    _write_template(existing)

    df_plan = _planning_df("260002", "5678", "DLA", date.fromisocalendar(year, week, 1))
    result = export_planning_excel(
        df_plan,
        week,
        year,
        df_dispos=pd.DataFrame(),
        increment_version=False,
        pdf_exporter=lambda src, dst: dst,
    )
    out_path = result.output_path

    history_dir = planning_dir / "Historique"
    archived = history_dir / existing.name
    assert archived.exists()
    assert out_path.exists()
    assert out_path.name == existing.name

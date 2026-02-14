# -*- coding: utf-8 -*-
from __future__ import annotations

from pathlib import Path

import pandas as pd

from asf_app.services.shipments_update_service import (
    apply_planning_update,
    apply_planning_updates_batch,
    load_be_status,
    load_be_status_d_for_week,
)


def _build_mag_central(path: Path):
    df_mag = pd.DataFrame(
        [
            {
                "N° BE": "250001",
                "NB": 2,
                "DEST": "DLA",
                "TYPE": "MM",
                "EXP": "ASF",
                "DESTINATAIRE": "Hopital",
                "DATE DE DEPART VOL": "2025-01-06",
                "Statut BE": "D",
            },
            {
                "N° BE": "250002",
                "NB": 1,
                "DEST": "DLA",
                "TYPE": "MM",
                "EXP": "ASF",
                "DESTINATAIRE": "Hopital",
                "DATE DE DEPART VOL": "2025-01-20",
                "Statut BE": "X",
            },
        ]
    )
    with pd.ExcelWriter(path) as writer:
        df_mag.to_excel(writer, sheet_name="MAG CENTRAL 2025", index=False, startrow=5)


def test_load_be_status_filters_and_weeks(tmp_path):
    path = tmp_path / "tdb.xlsx"
    _build_mag_central(path)

    df = load_be_status("D", tdb_path=path)
    assert len(df) == 1
    assert df.iloc[0]["BE_Numero_Str"] == "250001"
    assert int(df.iloc[0]["Year"]) == 2025
    assert int(df.iloc[0]["Week"]) == pd.Timestamp("2025-01-06").isocalendar().week


def test_load_be_status_d_for_week(tmp_path):
    path = tmp_path / "tdb.xlsx"
    _build_mag_central(path)

    week = pd.Timestamp("2025-01-06").isocalendar().week
    df = load_be_status_d_for_week(week, 2025, tdb_path=path)
    assert len(df) == 1
    assert df.iloc[0]["BE_Numero_Str"] == "250001"


def test_apply_planning_update(tmp_path):
    import datetime as dt

    from openpyxl import Workbook, load_workbook

    import scheduler.config_paths as cp
    from utils.datetime_utils import coerce_datetime

    path = tmp_path / "planning.xlsx"
    wb = Workbook()
    ws_plan = wb.active
    ws_plan.title = "Planning"
    ws_plan["Q1"] = 1

    ws_exp = wb.create_sheet("Export planning")
    headers = [
        "BE_Numero",
        "Date_Vol",
        "Heure_Vol",
        "BE_Nb_Colis",
        "BE_Type",
        "BE_Expediteur",
        "BE_Destinataire",
    ]
    ws_exp.append(headers)
    ws_exp.append(["250001", "01/01/2025", "10:00", 2, "MM", "ASF", "Hopital"])
    wb.save(path)

    be_info = pd.Series({"BE_Nb_Colis": 3, "BE_Type": "MM"})
    # Align with export_planning_excel versioning logic based on template A1
    wk, yr = 1, 2025
    try:
        template = cp.PLANNING_TEMPLATE
        if template.exists():
            wb_tpl = load_workbook(template)
            val = wb_tpl.worksheets[0]["A1"].value
            if isinstance(val, dt.datetime):
                wk = val.isocalendar()[1]
                yr = val.isocalendar()[0]
            else:
                dt_val = coerce_datetime(val, errors="coerce", dayfirst=True)
                if pd.notna(dt_val):
                    wk = dt_val.isocalendar()[1]
                    yr = dt_val.isocalendar()[0]
    except Exception:
        pass

    # Fake existing version to force increment to v02
    existing = tmp_path / f"ASFmm - PLANNING SEMAINE {yr}-{wk:02d}-01.xlsx"
    existing.touch()

    out_path = apply_planning_update(
        path=path,
        action="Replanification",
        be_num="250001",
        dest_iata="DLA",
        date_new="02/01/2025",
        vol_new="AF123",
        heure_new="11:00",
        bene_choice="DUPONT",
        be_info=be_info,
        week=wk,
        year=yr,
    )

    out_path = Path(out_path)
    assert out_path.exists()
    wb2 = load_workbook(out_path)
    assert wb2.worksheets[0]["Q1"].value == 2

    ws_exp2 = wb2["Export planning"]
    rows = list(ws_exp2.iter_rows(values_only=True))
    header = list(rows[0])
    status_idx = header.index("_STATUS")
    be_idx = header.index("BE_Numero")

    statuses = [r[status_idx] for r in rows[1:]]
    bes = [str(r[be_idx]) for r in rows[1:]]

    assert "250001" in bes
    assert "old" in statuses
    assert "new" in statuses
    assert any(str(name).startswith("Planning") for name in wb2.sheetnames)


def test_apply_planning_updates_batch(tmp_path):
    import datetime as dt

    from openpyxl import Workbook, load_workbook

    import scheduler.config_paths as cp
    from utils.datetime_utils import coerce_datetime

    path = tmp_path / "planning.xlsx"
    wb = Workbook()
    ws_plan = wb.active
    ws_plan.title = "Planning"
    ws_plan["Q1"] = 1

    ws_exp = wb.create_sheet("Export planning")
    headers = [
        "BE_Numero",
        "Date_Vol",
        "Heure_Vol",
        "BE_Nb_Colis",
        "BE_Type",
        "BE_Expediteur",
        "BE_Destinataire",
    ]
    ws_exp.append(headers)
    ws_exp.append(["250001", "01/01/2025", "10:00", 2, "MM", "ASF", "Hopital"])
    wb.save(path)

    wk, yr = 1, 2025
    try:
        template = cp.PLANNING_TEMPLATE
        if template.exists():
            wb_tpl = load_workbook(template)
            val = wb_tpl.worksheets[0]["A1"].value
            if isinstance(val, dt.datetime):
                wk = val.isocalendar()[1]
                yr = val.isocalendar()[0]
            else:
                dt_val = coerce_datetime(val, errors="coerce", dayfirst=True)
                if pd.notna(dt_val):
                    wk = dt_val.isocalendar()[1]
                    yr = dt_val.isocalendar()[0]
    except Exception:
        pass

    existing = tmp_path / f"ASFmm - PLANNING SEMAINE {yr}-{wk:02d}-01.xlsx"
    existing.touch()

    updates = [
        {
            "action": "Changement de date ou bénévole",
            "be_num": "250001",
            "dest_iata": "DLA",
            "date_new": "02/01/2025",
            "vol_new": "AF123",
            "heure_new": "11:00",
            "bene_choice": "DUPONT",
            "be_info": {"BE_Nb_Colis": 3, "BE_Type": "MM"},
            "plan_row_full": {},
            "bene_meta": {},
            "bene_changed": False,
        },
        {
            "action": "Ajouter au planning",
            "be_num": "250002",
            "dest_iata": "DLA",
            "date_new": "03/01/2025",
            "vol_new": "AF456",
            "heure_new": "12:00",
            "bene_choice": "MARTIN",
            "be_info": {"BE_Nb_Colis": 1, "BE_Type": "MM"},
            "plan_row_full": {},
            "bene_meta": {},
            "bene_changed": False,
        },
    ]

    out_path = apply_planning_updates_batch(
        path,
        updates,
        week=wk,
        year=yr,
        increment_version=True,
        write_mag_central=False,
    )

    out_path = Path(out_path)
    assert out_path.exists()
    wb2 = load_workbook(out_path)
    assert wb2.worksheets[0]["Q1"].value == 2

    ws_exp2 = wb2["Export planning"]
    rows = list(ws_exp2.iter_rows(values_only=True))
    header = list(rows[0])
    status_idx = header.index("_STATUS")
    be_idx = header.index("BE_Numero")

    statuses = [r[status_idx] for r in rows[1:]]
    bes = [str(r[be_idx]) for r in rows[1:]]

    assert "250001" in bes
    assert "250002" in bes
    assert "old" in statuses
    assert "new" in statuses


def test_apply_planning_updates_batch_without_increment_and_with_mag_updates(monkeypatch, tmp_path):
    path = tmp_path / "planning.xlsx"
    path.write_text("x", encoding="utf-8")

    base_df = pd.DataFrame([{"BE_Key": "250001", "BE_Numero": "250001"}])
    monkeypatch.setattr("asf_app.services.shipments_update_service._load_export_df", lambda _p: base_df.copy())
    monkeypatch.setattr("asf_app.services.shipments_update_service._sort_export_df", lambda df: df)
    monkeypatch.setattr(
        "asf_app.services.shipments_update_service._apply_update_to_export_df",
        lambda df, **_kwargs: pd.concat([df, pd.DataFrame([{"BE_Key": "250002", "BE_Numero": "250002"}])], ignore_index=True),
    )

    class _Result:
        output_path = tmp_path / "out.xlsx"

    called_kwargs: list[dict] = []

    def _fake_export(*_args, **kwargs):
        called_kwargs.append(kwargs)
        return _Result()

    updated: list[str] = []
    monkeypatch.setattr("asf_app.services.export_service.export_planning_excel", _fake_export)
    monkeypatch.setattr("asf_app.services.shipments_update_service.cp.sync_local_file_to_onedrive", lambda *_a, **_k: None)
    monkeypatch.setattr(
        "asf_app.services.shipments_update_service._update_mag_central_for_be",
        lambda **kwargs: updated.append(str(kwargs.get("be_num"))) or "updated",
    )

    out = apply_planning_updates_batch(
        path=path,
        updates=[
            {"be_num": "250001", "action": "Replanification"},
            {"be_num": "250002", "action": "Replanification"},
        ],
        week=4,
        year=2026,
        increment_version=False,
        write_mag_central=True,
    )

    assert out == _Result.output_path
    assert called_kwargs and called_kwargs[0]["increment_version"] is False
    assert called_kwargs[0]["output_path"] == path
    assert updated == ["250001", "250002"]


def test_apply_planning_update_without_increment_and_with_mag_updates(monkeypatch, tmp_path):
    path = tmp_path / "planning.xlsx"
    path.write_text("x", encoding="utf-8")
    base_df = pd.DataFrame([{"BE_Key": "250001", "BE_Numero": "250001"}])
    monkeypatch.setattr("asf_app.services.shipments_update_service._load_export_df", lambda _p: base_df.copy())
    monkeypatch.setattr("asf_app.services.shipments_update_service._sort_export_df", lambda df: df)
    monkeypatch.setattr(
        "asf_app.services.shipments_update_service._apply_update_to_export_df",
        lambda df, **_kwargs: pd.concat([df, pd.DataFrame([{"BE_Key": "250001", "_STATUS": "new"}])], ignore_index=True),
    )

    class _Result:
        output_path = tmp_path / "out_single.xlsx"

    called_kwargs: list[dict] = []

    def _fake_export(*_args, **kwargs):
        called_kwargs.append(kwargs)
        return _Result()

    mag_calls: list[str] = []
    monkeypatch.setattr("asf_app.services.export_service.export_planning_excel", _fake_export)
    monkeypatch.setattr("asf_app.services.shipments_update_service.cp.sync_local_file_to_onedrive", lambda *_a, **_k: None)
    monkeypatch.setattr(
        "asf_app.services.shipments_update_service._update_mag_central_for_be",
        lambda **kwargs: mag_calls.append(str(kwargs.get("be_num"))) or "updated",
    )

    out = apply_planning_update(
        path=path,
        action="Replanification",
        be_num="250001",
        dest_iata="DLA",
        date_new="06/01/2025",
        vol_new="AF123",
        heure_new="11:30",
        bene_choice="DUPONT",
        be_info=pd.Series({"BE_Nb_Colis": 2}),
        week=4,
        year=2026,
        increment_version=False,
        write_mag_central=True,
    )

    assert out == _Result.output_path
    assert called_kwargs and called_kwargs[0]["increment_version"] is False
    assert called_kwargs[0]["output_path"] == path
    assert mag_calls == ["250001"]

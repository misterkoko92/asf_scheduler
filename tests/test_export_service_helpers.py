# -*- coding: utf-8 -*-
from __future__ import annotations

from datetime import date, datetime

import pandas as pd
from openpyxl import Workbook, load_workbook

import scheduler.config_paths as cp
import utils.excel_automation as excel_automation
from asf_app.services.export_service import (
    _alt_key_for_sheet,
    _apply_planning_layout,
    _apply_routing_fallback_from_vols,
    _archive_latest_planning_if_needed,
    _build_bene_display_map,
    _collect_existing_planning_versions,
    _create_minimal_workbook,
    _export_pdf_with_warning,
    _extract_version_from_planning_name,
    _find_mag_row,
    _format_bene_display,
    _friday_previous_week,
    _increment_q1_if_requested,
    _mag_lookup_keys,
    _normalize_vol_key,
    _populate_planning_sheet,
    _prepare_export_dataframe,
    _prepare_output_workbook_path,
    _reset_planning_grid,
    _resolve_benevoles_export_dataframe,
    _resolve_depart_mag_date,
    _resolve_target_planning_path,
    _safe_text,
    _save_sync_and_move_planning_output,
    _set_q1_version,
    _sheet_order_for_be,
    _update_mag_central_dates_for_export,
    _week_year_from_ws_plan,
    _write_dataframe_to_sheet,
    _write_planning_row,
)


def test_create_minimal_workbook_has_expected_sheets(tmp_path):
    path = tmp_path / "minimal.xlsx"
    _create_minimal_workbook(path, week=3, year=2026)

    wb = load_workbook(path)
    assert wb.worksheets[0].title == "Planning SXX"
    assert wb.worksheets[0]["Q1"].value == 0
    assert "Export planning" in wb.sheetnames
    assert "Data Vols" in wb.sheetnames
    assert "Data Benevoles" in wb.sheetnames


def test_week_year_from_ws_plan_datetime_value():
    wb = Workbook()
    ws = wb.active
    ws["A1"] = datetime(2026, 1, 14, 10, 30)

    week, year = _week_year_from_ws_plan(ws, fallback_week=1, fallback_year=2025)
    assert (week, year) == (3, 2026)


def test_week_year_from_ws_plan_string_value():
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "14/01/2026"

    week, year = _week_year_from_ws_plan(ws, fallback_week=1, fallback_year=2025)
    assert (week, year) == (3, 2026)


def test_week_year_from_ws_plan_invalid_fallback():
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "not-a-date"

    week, year = _week_year_from_ws_plan(ws, fallback_week=5, fallback_year=2027)
    assert (week, year) == (5, 2027)


def test_extract_version_from_planning_name_new_format():
    name = "ASFmm - PLANNING SEMAINE 2026-05-03.xlsx"
    assert _extract_version_from_planning_name(name, week=5, year=2026) == 3


def test_extract_version_from_planning_name_legacy_with_v():
    name = "ASFmm - PLANNING SEMAINE No 05 - 2026 v7.xlsm"
    assert _extract_version_from_planning_name(name, week=5, year=2026) == 7


def test_extract_version_from_planning_name_legacy_without_v_defaults_to_1():
    name = "ASFmm - PLANNING SEMAINE No 05 - 2026.xlsx"
    assert _extract_version_from_planning_name(name, week=5, year=2026) == 1


def test_extract_version_from_planning_name_returns_none_for_other_week():
    name = "ASFmm - PLANNING SEMAINE 2026-05-03.xlsx"
    assert _extract_version_from_planning_name(name, week=6, year=2026) is None


def test_build_bene_display_map_and_format_bene_display():
    df_parambenev = pd.DataFrame(
        {
            "BENEVOLE": ["Alice Martin"],
            "Prenom court": ["A"],
            "Nom": ["Martin"],
        }
    )
    display_map = _build_bene_display_map(df_parambenev)

    assert display_map["Alice Martin"] == "A MARTIN"
    assert _format_bene_display("Alice Martin", display_map=display_map) == "A MARTIN"
    assert _format_bene_display("Bob Dupont", display_map=display_map) == "B. DUPONT"
    assert _format_bene_display("mono", display_map=display_map) == "MONO"


def test_normalize_vol_key_handles_af_prefix_and_digits():
    assert _normalize_vol_key("AF0652") == "652"
    assert _normalize_vol_key(" af 0652A ") == "652"
    assert _normalize_vol_key("KLM") == "KLM"
    assert _normalize_vol_key(None) == ""


def test_apply_routing_fallback_from_vols_fills_only_missing_routing():
    df_export = pd.DataFrame(
        {
            "DATE": [pd.Timestamp("2026-01-23"), pd.Timestamp("2026-01-23")],
            "Numero_Vol": ["AF0652", "AF0653"],
            "Routing": ["", "KEEP-ME"],
        }
    )
    df_vols = pd.DataFrame(
        {
            "Date_Vol": ["23/01/2026", "23/01/2026"],
            "Numero_Vol": ["652", "653"],
            "Routing": ["CDG-RUN", "CDG-FRA-RUN"],
        }
    )

    out = _apply_routing_fallback_from_vols(df_export, df_vols=df_vols)

    assert out.loc[0, "Routing"] == "CDG-RUN"
    assert out.loc[1, "Routing"] == "KEEP-ME"
    assert "_DATE_KEY" not in out.columns
    assert "_VOL_KEY" not in out.columns


def test_apply_routing_fallback_from_vols_creates_routing_column_when_absent():
    df_export = pd.DataFrame(
        {
            "DATE": [pd.Timestamp("2026-01-23")],
            "Numero_Vol": ["AF0652"],
        }
    )
    df_vols = pd.DataFrame(
        {
            "Date_Vol_dt": [pd.Timestamp("2026-01-23")],
            "Numero_Vol": ["652"],
            "Routing": ["CDG-RUN"],
        }
    )

    out = _apply_routing_fallback_from_vols(df_export, df_vols=df_vols)

    assert "Routing" in out.columns
    assert out.loc[0, "Routing"] == "CDG-RUN"


def test_prepare_export_dataframe_enriches_and_sorts(monkeypatch):
    source = pd.DataFrame([{"dummy": 1}, {"dummy": 2}])
    built = pd.DataFrame(
        [
            {
                "Date_Vol": "2026-01-23",
                "Heure_Vol": "12:00",
                "Numero_Vol": "AF653",
                "Destination": "RUN",
                "Ville": "Saint-Denis",
                "IATA": "run",
                "Routing": "",
                "BE_Numero": "260002",
                "BE_Nb_Colis": "1",
                "BE_Type": "MM",
                "BE_Expediteur": "ASF",
                "BE_Destinataire": "HOP",
                "Benevole": "Alice Martin",
            },
            {
                "Date_Vol": "2026-01-23",
                "Heure_Vol": "10:00",
                "Numero_Vol": "AF652",
                "Destination": "RUN",
                "Ville": "Saint-Denis",
                "IATA": "run",
                "Routing": "",
                "BE_Numero": "260001",
                "BE_Nb_Colis": "2",
                "BE_Type": "FRET",
                "BE_Expediteur": "ASF",
                "BE_Destinataire": "HOP",
                "Benevole": "Alice Martin",
            },
        ]
    )
    monkeypatch.setattr(
        "asf_app.services.export_service.build_export_view",
        lambda *_args, **_kwargs: built.copy(),
    )
    df_parambenev = pd.DataFrame(
        [{"Benevole": "Alice Martin", "Prenom_Court": "A", "Nom": "Martin"}]
    )
    df_vols = pd.DataFrame(
        [
            {"Date_Vol": "23/01/2026", "Numero_Vol": "652", "Routing": "CDG-RUN"},
            {"Date_Vol": "23/01/2026", "Numero_Vol": "653", "Routing": "CDG-RUN"},
        ]
    )

    out = _prepare_export_dataframe(
        source,
        df_paramdest=None,
        df_vols=df_vols,
        df_parambenev=df_parambenev,
    )

    assert list(out["VOL_AFF"]) == ["AF 652", "AF 653"]
    assert list(out["BE_NUM"]) == ["260001", "260002"]
    assert list(out["BENEVOLE_DISP"]) == ["A MARTIN", "A MARTIN"]
    assert list(out["HEURE_AFF"]) == ["10h00", "12h00"]
    assert list(out["ROUTING"]) == ["CDG-RUN", "CDG-RUN"]
    assert list(out["BE_COLIS"]) == [2, 1]
    assert out["_STATUS"].tolist() == ["normal", "normal"]


def test_resolve_depart_mag_date_prefers_map_and_fallback():
    mapped = _resolve_depart_mag_date(
        "260001",
        map_depart_mag={"260001": date(2026, 1, 2)},
        week=1,
        year=2026,
    )
    assert mapped == date(2026, 1, 2)

    fallback = _resolve_depart_mag_date(
        "260002",
        map_depart_mag={},
        week=1,
        year=2026,
    )
    assert fallback == _friday_previous_week(1, 2026)

    invalid = _resolve_depart_mag_date(
        "260003",
        map_depart_mag={},
        week=99,
        year=2026,
    )
    assert invalid is None


def test_write_dataframe_to_sheet_writes_rows_and_table():
    wb = Workbook()
    ws = wb.active
    df = pd.DataFrame([{"A": 1, "B": "x"}, {"A": 2, "B": "y"}])

    _write_dataframe_to_sheet(
        ws,
        df,
        create_tables=True,
        table_name="Table_Test",
    )
    assert ws.max_row == 3
    assert ws.max_column == 2
    assert ws["A2"].value == 1
    assert ws["B3"].value == "y"
    assert "Table_Test" in ws.tables

    _write_dataframe_to_sheet(
        ws,
        None,
        create_tables=False,
        table_name=None,
    )
    assert len(ws.tables) == 0


def test_export_pdf_with_warning_success_and_failure(tmp_path):
    out_path = tmp_path / "planning.xlsx"
    out_path.write_text("xlsx")
    warnings: list[str] = []

    def ok_exporter(_src, dst):
        dst.write_text("pdf")
        return dst

    pdf_ok = _export_pdf_with_warning(
        out_path,
        enabled=True,
        pdf_exporter=ok_exporter,
        warnings=warnings,
    )
    assert pdf_ok == out_path.with_suffix(".pdf")
    assert warnings == []

    def failing_exporter(_src, _dst):
        raise RuntimeError("boom")

    pdf_ko = _export_pdf_with_warning(
        out_path,
        enabled=True,
        pdf_exporter=failing_exporter,
        warnings=warnings,
    )
    assert pdf_ko is None
    assert any("PDF non généré" in message for message in warnings)

    pdf_disabled = _export_pdf_with_warning(
        out_path,
        enabled=False,
        pdf_exporter=ok_exporter,
        warnings=warnings,
    )
    assert pdf_disabled is None


def test_prepare_output_workbook_path_explicit_and_default(tmp_path):
    template = tmp_path / "template.xlsx"
    _create_minimal_workbook(template, week=4, year=2026)

    explicit_out = tmp_path / "custom.xlsx"
    out_path, skip = _prepare_output_workbook_path(
        week=4,
        year=2026,
        template=template,
        has_template=True,
        output_path=explicit_out,
        output_dir=None,
        skip_versioning=False,
    )
    assert out_path == explicit_out
    assert out_path.exists()
    assert skip is True

    out_dir = tmp_path / "out"
    auto_out, auto_skip = _prepare_output_workbook_path(
        week=4,
        year=2026,
        template=tmp_path / "missing-template.xlsx",
        has_template=False,
        output_path=None,
        output_dir=out_dir,
        skip_versioning=False,
    )
    assert auto_out.exists()
    assert auto_out.name == "ASFmm - PLANNING SEMAINE 2026-04-TMP.xlsx"
    assert auto_skip is False


def test_resolve_benevoles_export_dataframe_fallback_and_columns(monkeypatch):
    fallback_df = pd.DataFrame(
        [
            {
                "ID": "1",
                "Benevole": "ALICE",
                "Nom": "Dupont",
                "Prenom": "Alice",
                "Prenom_Court": "A",
                "Date": "2026-01-23",
                "Heure_Arrivee": "10:00",
                "Heure_Depart": "12:00",
                "Unused": "x",
            }
        ]
    )
    monkeypatch.setattr(
        "asf_app.services.export_service.get_benevoles_cached",
        lambda **_kwargs: fallback_df,
    )

    out = _resolve_benevoles_export_dataframe(None, benev_path=None)
    assert out is not None
    assert list(out.columns) == [
        "ID",
        "Benevole",
        "Nom",
        "Prenom",
        "Prenom_Court",
        "Date",
        "Heure_Arrivee",
        "Heure_Depart",
    ]

    monkeypatch.setattr(
        "asf_app.services.export_service.get_benevoles_cached",
        lambda **_kwargs: (_ for _ in ()).throw(RuntimeError("boom")),
    )
    out_none = _resolve_benevoles_export_dataframe(None, benev_path=None)
    assert out_none is None


def test_collect_versions_archive_and_resolve_target(tmp_path):
    planning_dir = tmp_path / "planning"
    planning_dir.mkdir(parents=True, exist_ok=True)
    v1 = planning_dir / "ASFmm - PLANNING SEMAINE 2026-04-01.xlsx"
    v2 = planning_dir / "ASFmm - PLANNING SEMAINE 2026-04-02.xlsx"
    v1.touch()
    v2.touch()

    existing, max_version = _collect_existing_planning_versions(
        planning_dir,
        week_final=4,
        year_final=2026,
    )
    assert max_version == 2
    assert set(existing) == {v1, v2}

    _archive_latest_planning_if_needed(
        increment_version=False,
        max_version=max_version,
        existing_files=existing,
        planning_dir_final=planning_dir,
        week_final=4,
        year_final=2026,
    )
    history_dir = planning_dir / "Historique"
    assert history_dir.exists()
    assert (history_dir / v2.name).exists()

    version_no_inc, target_no_inc = _resolve_target_planning_path(
        planning_dir_final=planning_dir,
        week_final=4,
        year_final=2026,
        max_version=max_version,
        increment_version=False,
    )
    assert version_no_inc == 2
    assert target_no_inc.name == "ASFmm - PLANNING SEMAINE 2026-04-02.xlsx"

    occupied = planning_dir / "ASFmm - PLANNING SEMAINE 2026-04-03.xlsx"
    occupied.touch()
    version_inc, target_inc = _resolve_target_planning_path(
        planning_dir_final=planning_dir,
        week_final=4,
        year_final=2026,
        max_version=max_version,
        increment_version=True,
    )
    assert version_inc == 4
    assert target_inc.name == "ASFmm - PLANNING SEMAINE 2026-04-04.xlsx"


def test_set_and_increment_q1_helpers():
    wb = Workbook()
    ws = wb.active
    ws["Q1"] = None
    _increment_q1_if_requested(ws, increment_version=True)
    assert ws["Q1"].value == 1
    _increment_q1_if_requested(ws, increment_version=True)
    assert ws["Q1"].value == 2
    _increment_q1_if_requested(ws, increment_version=False)
    assert ws["Q1"].value == 2
    _set_q1_version(ws, 7)
    assert ws["Q1"].value == 7


def test_save_sync_and_move_output_handles_move_and_graph_sync(tmp_path, monkeypatch):
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "x"

    out_path = tmp_path / "tmp.xlsx"
    target_path = tmp_path / "final.xlsx"
    wb.save(out_path)

    monkeypatch.setattr("asf_app.services.export_service.is_graph_onedrive", lambda: False)
    moved = _save_sync_and_move_planning_output(
        wb,
        out_path=out_path,
        target_path=target_path,
        year_final=2026,
    )
    assert moved == target_path
    assert target_path.exists()

    sync_calls: list[tuple] = []
    monkeypatch.setattr("asf_app.services.export_service.is_graph_onedrive", lambda: True)
    monkeypatch.setattr(cp, "sync_local_file_to_onedrive", lambda *args, **kwargs: sync_calls.append((args, kwargs)))

    same_path = _save_sync_and_move_planning_output(
        wb,
        out_path=target_path,
        target_path=target_path,
        year_final=2026,
    )
    assert same_path == target_path
    assert len(sync_calls) == 1


def test_reset_planning_grid_clears_cells_and_preserves_totals():
    wb = Workbook()
    ws = wb.active
    ws["D3"] = "x"
    ws["K219"] = "KEEP-K"
    ws["L219"] = "KEEP-L"
    ws.row_dimensions[3].hidden = True
    ws.row_dimensions[4].hidden = True

    _reset_planning_grid(ws)

    assert ws["D3"].value is None
    assert ws["K219"].value == "KEEP-K"
    assert ws["L219"].value == "KEEP-L"
    assert ws.row_dimensions[3].hidden is False


def test_write_planning_row_writes_cells_and_status_fill():
    wb = Workbook()
    ws = wb.active
    row_data_old = pd.Series(
        {
            "VILLE": "DOUALA",
            "IATA": "DLA",
            "ROUTING": "CDG-DLA",
            "VOL_AFF": "AF 654",
            "HEURE_AFF": "10h00",
            "BE_NUM": "260001",
            "BE_COLIS": 3,
            "BE_TYPE": "MM",
            "DEPART_MAG": pd.Timestamp("2026-01-02"),
            "BE_EXP": "ASF",
            "BE_DEST": "HOP",
            "_STATUS": "old",
        }
    )
    _write_planning_row(
        ws,
        row_idx=5,
        row_data=row_data_old,
        bene_val="A MARTIN",
        is_first=True,
        safe_excel=lambda value: value,
    )
    assert ws.cell(row=5, column=4).value == "A MARTIN"
    assert ws.cell(row=5, column=6).value == "DOUALA"
    assert ws.cell(row=5, column=12).value == ""
    fill_rgb = str(ws.cell(row=5, column=4).fill.fgColor.rgb or "")
    assert fill_rgb.endswith("F8CCCC")

    row_data_new = row_data_old.copy()
    row_data_new["_STATUS"] = "new"
    row_data_new["BE_COLIS"] = 4
    _write_planning_row(
        ws,
        row_idx=6,
        row_data=row_data_new,
        bene_val="B DUPONT",
        is_first=False,
        safe_excel=lambda value: value,
    )
    assert ws.cell(row=6, column=6).value is None
    assert ws.cell(row=6, column=12).value == 4
    fill_rgb_new = str(ws.cell(row=6, column=4).fill.fgColor.rgb or "")
    assert fill_rgb_new.endswith("CFE2FF")


def test_populate_planning_sheet_writes_rows_and_hides_unused():
    wb = Workbook()
    ws = wb.active
    _reset_planning_grid(ws)

    dfp = pd.DataFrame(
        [
            {
                "DATE": pd.Timestamp("2026-01-19"),  # lundi
                "HEURE_MIN": 600,
                "VOL_AFF": "AF 652",
                "Destination": "RUN",
                "BENEVOLE_DISP": "A MARTIN",
                "VILLE": "SAINT-DENIS",
                "IATA": "RUN",
                "ROUTING": "CDG-RUN",
                "HEURE_AFF": "10h00",
                "BE_NUM": "260001",
                "BE_COLIS": 2,
                "BE_TYPE": "MM",
                "DEPART_MAG": pd.Timestamp("2026-01-16"),
                "BE_EXP": "ASF",
                "BE_DEST": "HOP",
                "_STATUS": "normal",
            }
        ]
    )
    day_blocks = {0: (4, 8)}
    keep_rows = {4, 8}
    _populate_planning_sheet(
        ws,
        dfp=dfp,
        day_blocks=day_blocks,
        keep_rows=keep_rows,
        safe_excel=lambda value: value,
    )
    assert ws.cell(row=4, column=4).value == "A MARTIN"
    assert ws.cell(row=4, column=11).value == "260001"
    assert ws.row_dimensions[7].hidden is True
    assert ws.row_dimensions[8].hidden in (False, None)


def test_apply_planning_layout_hides_columns_and_moves_titles():
    wb = Workbook()
    ws = wb.active
    ws["A17"] = "LUNDI"
    ws["P4"] = "COLLONG"
    ws["Q4"] = "X"
    ws.row_dimensions[18].hidden = False
    ws.row_dimensions[17].hidden = False

    _apply_planning_layout(
        ws,
        hidden_columns=("B", "C"),
        auto_width_columns=("P", "Q"),
        middle_moves=((17, 4, 32),),
    )
    assert ws.column_dimensions["B"].hidden is True
    assert ws.column_dimensions["C"].hidden is True
    assert ws.column_dimensions["P"].width >= 10
    assert ws["A17"].value is None
    assert ws["A18"].value == "LUNDI"


def test_mag_lookup_keys_includes_base_and_suffixes():
    keys = _mag_lookup_keys("000123")
    assert keys[0] == "000123"
    assert "123" in keys
    assert "0123" in keys


def test_alt_key_for_sheet_uses_year_suffix():
    assert _alt_key_for_sheet("000123", "MAG CENTRAL 2025") == "250123"
    assert _alt_key_for_sheet("250123", "MAG CENTRAL 2025") is None
    assert _alt_key_for_sheet("000123", "MAG CENTRAL") is None


def test_sheet_order_for_be_prefers_be_prefix_then_preferred_year():
    names = ["MAG CENTRAL 2024", "MAG CENTRAL 2026", "MAG CENTRAL 2025"]
    ordered = _sheet_order_for_be("250001", names, preferred_year=2026)
    assert ordered[0] == "MAG CENTRAL 2025"
    assert ordered[1] == "MAG CENTRAL 2026"
    assert set(ordered) == set(names)


def test_find_mag_row_uses_alt_key_when_needed():
    sheet = "MAG CENTRAL 2025"
    target_sheet, target_row = _find_mag_row(
        be_key="000123",
        mag_sheet_names=[sheet],
        mag_indexes={sheet: {"250123": 7}},
        preferred_year=2025,
    )
    assert target_sheet == sheet
    assert target_row == 7


def test_safe_text_handles_none_and_na():
    assert _safe_text(None) == ""
    assert _safe_text(pd.NA) == ""
    assert _safe_text("  x  ") == "x"


def test_update_mag_central_dates_for_export_missing_file(tmp_path):
    missing = tmp_path / "nope.xlsx"
    used_dates, method = _update_mag_central_dates_for_export(
        df_export=pd.DataFrame(),
        week=1,
        year=2026,
        tdb_source_path=missing,
    )
    assert used_dates == {}
    assert method == "missing"


def test_update_mag_central_dates_for_export_openpyxl_fallback(tmp_path, monkeypatch):
    path = tmp_path / "tdb.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "MAG CENTRAL 2026"
    ws["A1"] = "N° BE"
    ws["A2"] = "260001"
    wb.save(path)

    monkeypatch.setattr(cp, "sync_local_file_to_onedrive", lambda *_args, **_kwargs: None)
    monkeypatch.setattr(excel_automation, "update_excel_cells", lambda *_args, **_kwargs: False)

    df_export = pd.DataFrame(
        [
            {
                "BE_KEY": "260001",
                "DATE": pd.Timestamp("2026-01-05"),
                "ID": "45.0",
                "BENEVOLE_DISP": "DUPONT",
                "VOL_AFF": "AF123",
                "HEURE_AFF": "11h30",
            }
        ]
    )
    used_dates, method = _update_mag_central_dates_for_export(
        df_export=df_export,
        week=1,
        year=2026,
        tdb_source_path=path,
    )

    assert method == "openpyxl"
    assert used_dates["260001"] == _friday_previous_week(1, 2026)

    wb2 = load_workbook(path)
    ws2 = wb2["MAG CENTRAL 2026"]

    depart_vol = ws2.cell(row=2, column=cp.MAG_CENTRAL_COL_DEPART_VOL).value
    depart_mag = ws2.cell(row=2, column=cp.MAG_CENTRAL_COL_DEPART_MAG).value
    if hasattr(depart_vol, "date"):
        depart_vol = depart_vol.date()
    if hasattr(depart_mag, "date"):
        depart_mag = depart_mag.date()

    assert depart_vol == date(2026, 1, 5)
    assert depart_mag == _friday_previous_week(1, 2026)
    assert ws2.cell(row=2, column=cp.MAG_CENTRAL_COL_ID_BENEV).value == "45"
    assert ws2.cell(row=2, column=cp.MAG_CENTRAL_COL_BENEV).value == "DUPONT"
    assert ws2.cell(row=2, column=cp.MAG_CENTRAL_COL_VOL).value == "AF123"
    assert ws2.cell(row=2, column=cp.MAG_CENTRAL_COL_HEURE).value == "11h30"

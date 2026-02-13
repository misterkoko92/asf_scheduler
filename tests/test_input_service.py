# -*- coding: utf-8 -*-
from __future__ import annotations

import pandas as pd
import pytest

import asf_app.services.input_service as input_service
import scheduler.config_paths as cp
from asf_app.services.input_service import InputLoadError, load_benev, load_tdb, load_vols


def test_load_tdb(sample_onedrive):
    data = load_tdb(cp.TABLEAU_DE_BORD)
    assert not data.df_be.empty
    assert not data.df_param_be.empty
    assert not data.df_param_dest.empty


def test_load_tdb_missing(tmp_path):
    with pytest.raises(FileNotFoundError):
        load_tdb(tmp_path / "missing.xlsx")


def test_load_benev(sample_onedrive):
    data = load_benev(cp.PLANNING_BENEVOLES)
    assert not data.df_param_benev.empty
    assert not data.df_benev.empty


def test_load_benev_missing(tmp_path):
    with pytest.raises(FileNotFoundError):
        load_benev(tmp_path / "missing.xlsx")


def test_load_vols(sample_onedrive):
    df = load_vols(cp.VOLS)
    assert not df.empty


def test_load_vols_missing(tmp_path):
    with pytest.raises(FileNotFoundError):
        load_vols(tmp_path / "missing.xlsx")


def test_load_vols_empty_file_returns_empty_df(tmp_path):
    path = tmp_path / "vols.xlsx"
    path.touch()
    df = load_vols(path)
    assert df.empty


def test_get_benev_source_message(tmp_path):
    from openpyxl import Workbook

    from asf_app.services.input_service import get_benev_source_message

    path = tmp_path / "benev.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Source"
    ws["D2"] = "27/01/2025"
    ws["E2"] = "10:30"
    wb.save(path)

    msg = get_benev_source_message(path)
    assert "27/01/25" in msg or "27/01/2025" in msg
    assert "10h30" in msg


def test_load_vols_fallbacks_to_simple_sheet_when_loader_fails(tmp_path, monkeypatch):
    path = tmp_path / "vols.xlsx"
    path.touch()

    monkeypatch.setattr(input_service, "load_vols_df", lambda **_kwargs: (_ for _ in ()).throw(ValueError("boom")))
    monkeypatch.setattr(
        input_service,
        "load_and_normalize",
        lambda *_args, **_kwargs: pd.DataFrame([{"Numero_Vol": "AF123"}]),
    )

    df = input_service.load_vols(path)
    assert not df.empty
    assert "Numero_Vol" in df.columns


def test_load_vols_raises_input_load_error_when_all_strategies_fail(tmp_path, monkeypatch):
    path = tmp_path / "vols.xlsx"
    path.touch()

    monkeypatch.setattr(input_service, "load_vols_df", lambda **_kwargs: (_ for _ in ()).throw(ValueError("boom")))
    monkeypatch.setattr(
        input_service,
        "load_and_normalize",
        lambda *_args, **_kwargs: (_ for _ in ()).throw(OSError("cannot read")),
    )

    with pytest.raises(InputLoadError):
        input_service.load_vols(path)


def test_load_vols_param_dest_fallback_to_none_when_tdb_load_fails(tmp_path, monkeypatch):
    path = tmp_path / "vols.xlsx"
    tdb_path = tmp_path / "tdb.xlsx"
    path.touch()
    tdb_path.touch()

    monkeypatch.setattr(
        input_service,
        "load_param_dest_from_path",
        lambda *_args, **_kwargs: (_ for _ in ()).throw(OSError("boom")),
    )
    captured: dict[str, object] = {}
    monkeypatch.setattr(
        input_service,
        "load_vols_df",
        lambda **kwargs: captured.update(kwargs) or pd.DataFrame([{"Numero_Vol": "AF123"}]),
    )

    df = input_service.load_vols(path, tdb_path=tdb_path)
    assert not df.empty
    assert captured["param_dest_df"] is None


def test_load_normalized_sheet_delegates_to_universal_loader(tmp_path, monkeypatch):
    path = tmp_path / "x.xlsx"
    path.touch()
    monkeypatch.setattr(
        input_service,
        "load_and_normalize",
        lambda **kwargs: pd.DataFrame([{"sheet": kwargs["sheet_name"]}]),
    )
    out = input_service.load_normalized_sheet(path, "MySheet", {"A": "B"}, header=2)
    assert out.iloc[0]["sheet"] == "MySheet"


def test_get_benev_source_message_returns_na_for_missing_inputs(tmp_path):
    from asf_app.services.input_service import get_benev_source_message

    assert get_benev_source_message(None) == "N/A"
    assert get_benev_source_message(tmp_path / "missing.xlsx") == "N/A"


def test_get_benev_source_message_returns_na_when_source_missing_or_empty(tmp_path):
    from openpyxl import Workbook

    from asf_app.services.input_service import get_benev_source_message

    no_source = tmp_path / "no_source.xlsx"
    wb1 = Workbook()
    wb1.active.title = "Sheet1"
    wb1.save(no_source)
    assert get_benev_source_message(no_source) == "N/A"

    blank_source = tmp_path / "blank_source.xlsx"
    wb2 = Workbook()
    ws = wb2.active
    ws.title = "Source"
    ws["D2"] = None
    ws["E2"] = None
    wb2.save(blank_source)
    assert get_benev_source_message(blank_source) == "N/A"


def test_get_benev_source_message_handles_loader_exception(monkeypatch, tmp_path):
    from asf_app.services.input_service import get_benev_source_message

    path = tmp_path / "benev.xlsx"
    path.touch()

    monkeypatch.setattr(
        "openpyxl.load_workbook",
        lambda *_args, **_kwargs: (_ for _ in ()).throw(OSError("boom")),
    )
    assert get_benev_source_message(path) == "N/A"

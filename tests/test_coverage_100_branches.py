# -*- coding: utf-8 -*-
from __future__ import annotations

import importlib
import sys
from pathlib import Path
from types import SimpleNamespace
import types

import pandas as pd

import asf_app.services.export_service as es
import asf_app.services.shipments_update_service as sus
import asf_app.ui.ui_communication.clean_planning_df as clean_df
import asf_app.ui.ui_communication.whatsapp_handler as wa
import asf_app.ui.ui_simulation as ui_sim
import asf_app.ui.ui_stats.stats_loader as stats_loader
import asf_app.ui.ui_stats.stats_processor as stats_processor
import asf_app.ui.ui_stats.ui_stats as ui_stats
import asf_app.ui.ui_week_data as ui_week_data
import loaders.load_params as lp
import loaders.load_shipments as load_shipments_mod
import scheduler.be_manager as be_manager
import scheduler.be_rules as be_rules
import scheduler.planning_views as planning_views
import scheduler.solver_ortools_common as solver_common
from loaders.load_vols_api import store_vols_api_sheet
from scheduler.models import Shipment
from utils.benevole_utils import count_benevoles_with_dispo
from utils.identifiers import format_be_number, format_vol_number
from utils.ui_helpers import sort_planning_df
import utils.excel_automation as excel_auto


def test_streamlit_import_fallbacks_for_ui_and_loaders(monkeypatch, tmp_path):
    import asf_app.ui.email_defaults as ui_email_defaults
    import loaders.load_shipments as ls_mod
    import loaders.load_vols as load_vols_mod

    real_streamlit = sys.modules.get("streamlit")
    monkeypatch.setitem(sys.modules, "streamlit", None)

    ui_email_defaults = importlib.reload(ui_email_defaults)
    ls_mod = importlib.reload(ls_mod)
    load_vols_mod = importlib.reload(load_vols_mod)

    assert ui_email_defaults.st is None
    assert ls_mod.st is None

    calls: dict[str, object] = {}

    monkeypatch.setattr(
        ls_mod,
        "load_shipments_df",
        lambda **kwargs: calls.setdefault("shipments", kwargs) or pd.DataFrame(),
    )
    monkeypatch.setattr(
        load_vols_mod,
        "load_vols_df",
        lambda **kwargs: calls.setdefault("vols", kwargs) or pd.DataFrame(),
    )

    _ = ls_mod.get_shipments_df_cached(
        planifiables_only=False,
        tdb_path=tmp_path / "TABLEAU_DE_BORD.xlsx",
    )
    _ = load_vols_mod.get_vols_df_cached(
        vols_path=tmp_path / "VOLS.xlsx",
        tdb_path=tmp_path / "TABLEAU_DE_BORD.xlsx",
    )
    assert "shipments" in calls
    assert "vols" in calls

    if real_streamlit is None:
        sys.modules.pop("streamlit", None)
    else:
        sys.modules["streamlit"] = real_streamlit
    importlib.reload(ui_email_defaults)
    importlib.reload(ls_mod)
    importlib.reload(load_vols_mod)


def test_load_params_internal_cached_wrappers(monkeypatch, tmp_path):
    if not hasattr(lp, "_param_be_cached"):
        return

    for fn_name in (
        "_param_be_cached",
        "_param_dest_cached",
        "_param_exp_cached",
        "_param_benev_cached",
    ):
        fn = getattr(lp, fn_name, None)
        if fn is not None and hasattr(fn, "clear"):
            fn.clear()

    called: list[tuple[str, Path]] = []
    monkeypatch.setattr(
        lp,
        "_load_param_be",
        lambda tableau_de_bord_path=None: called.append(("be", Path(tableau_de_bord_path))) or pd.DataFrame(),
    )
    monkeypatch.setattr(
        lp,
        "_load_param_dest",
        lambda tableau_de_bord_path=None: called.append(("dest", Path(tableau_de_bord_path))) or pd.DataFrame(),
    )
    monkeypatch.setattr(
        lp,
        "_load_param_exp",
        lambda tableau_de_bord_path=None: called.append(("exp", Path(tableau_de_bord_path))) or pd.DataFrame(),
    )
    monkeypatch.setattr(
        lp,
        "_load_param_benev",
        lambda planning_benevoles_path=None: called.append(("benev", Path(planning_benevoles_path))) or pd.DataFrame(),
    )

    tdb = tmp_path / "tdb.xlsx"
    benev = tmp_path / "benev.xlsx"
    lp._param_be_cached(str(tdb), 1.0)
    lp._param_dest_cached(str(tdb), 2.0)
    lp._param_exp_cached(str(tdb), 3.0)
    lp._param_benev_cached(str(benev), 4.0)

    assert called == [
        ("be", tdb),
        ("dest", tdb),
        ("exp", tdb),
        ("benev", benev),
    ]


def test_be_manager_and_rules_edge_paths(monkeypatch, tmp_path):
    normalized = be_manager.normalize_param_be(
        pd.DataFrame(
            [
                {"Type": "", "Priorite_Type": 1, "Equiv": 1},
                {"Type": "MM", "Priorite_Type": 3, "Equiv": 2},
            ]
        )
    )
    assert "MM" in normalized
    assert "AUTRE" in normalized
    assert "" not in normalized

    class _BadHeadDf(pd.DataFrame):
        @property
        def _constructor(self):
            return _BadHeadDf

        def head(self, *args, **kwargs):
            _ = args, kwargs
            raise TypeError("boom")

    monkeypatch.setattr(
        be_manager,
        "load_and_normalize",
        lambda **_kwargs: _BadHeadDf([{"Type": "MM", "Priorite_Type": 4, "Equiv": 1}]),
    )
    out = be_manager.load_param_be(use_cache=False, tdb_path=tmp_path / "TABLEAU_DE_BORD.xlsx")
    assert int(out["MM"]["Priorite_Type"]) == 4

    assert be_rules.is_empty_date(pd.NA) is True
    assert be_rules.is_planifiable_status(be_rules.STATUS_PLANIFIABLE) is True
    shipment = Shipment(
        be_numero="1",
        dest="RUN",
        nb_colis_physiques=0,
        nb_hf=0,
        priority=0,
        type_colis="MM",
    )
    assert be_rules.compute_equiv_colis(shipment, {"MM": {"Equiv": 2}, "AUTRE": {"Equiv": 1}}) == 1


def test_models_identifiers_helpers_and_benevole_utils_misc():
    ship = Shipment(
        be_numero="260001",
        dest="DLA",
        nb_colis_physiques=1,
        nb_hf=0,
        priority=1,
    )
    assert ship.be_num == "260001"
    ship.be_num = "260999"
    assert ship.be_numero == "260999"

    assert format_be_number("1234") == "001234"
    assert format_vol_number("AF0652") == "652"
    assert sort_planning_df(pd.DataFrame()).empty

    count0, start0, end0 = count_benevoles_with_dispo(pd.DataFrame())
    assert count0 == 0 and start0 is None and end0 is None

    count1, _, _ = count_benevoles_with_dispo(pd.DataFrame([{"Date": "bad"}]))
    assert count1 == 0


def test_stats_loader_and_processor_remaining_branches(monkeypatch, tmp_path):
    file_path = tmp_path / "ASFmm - PLANNING SEMAINE 2026-08.xlsx"
    file_path.write_text("dummy", encoding="utf-8")

    monkeypatch.setattr(stats_loader, "load_planning_xlsx", lambda _path: pd.DataFrame())
    assert stats_loader.load_all_plannings(tmp_path).empty

    monkeypatch.setattr(
        stats_processor,
        "coerce_datetime",
        lambda *_args, **_kwargs: (_ for _ in ()).throw(TypeError("boom")),
    )
    delay = stats_processor.compute_transfer_delay(
        pd.DataFrame([{"date": "2026-01-23", "heure": "10:00"}])
    )
    assert isinstance(delay, pd.Series)
    assert delay.empty


def test_ui_stats_load_all_plannings_skips_files_without_week(monkeypatch, tmp_path):
    file_path = tmp_path / "random.xls"
    file_path.write_text("x", encoding="utf-8")

    import scheduler.config_paths as cp

    monkeypatch.setattr(cp, "detect_onedrive_asf", lambda: tmp_path)
    monkeypatch.setattr(cp, "get_planning_dirs", lambda: [tmp_path])
    monkeypatch.setattr(ui_stats, "get_output_planning_dir", lambda: tmp_path)
    monkeypatch.setattr(ui_stats, "extract_week_version", lambda _name: (None, None))

    out = ui_stats._load_all_plannings(base_override=tmp_path)
    assert out.empty


def test_solver_common_parse_vols_without_dest_columns_executes_default_path():
    df_raw = pd.DataFrame(
        [
            {
                "Date_Vol": "16/02/26",
                "Heure_Vol": "11:00",
                "Numero_Vol": "AF822",
            }
        ]
    )
    out = solver_common.parse_vols(df_raw, dest_info={})
    assert out.empty


def test_ui_simulation_remaining_destination_and_benevole_branches():
    df_plan = pd.DataFrame(
        [
            {
                "Destination": "DLA",
                "BE_Numero": "260001",
                "BE_Nb_Colis": 1,
                "BE_Nb_Equiv": 1,
                "Date_Vol": "16/02/26",
                "Heure_Vol": "11:00",
                "Numero_Vol": "AF822",
                "_MANUEL": False,
            }
        ]
    )
    df_vols = pd.DataFrame(
        [
            {
                "Date_Vol": "16/02/26",
                "Heure_Vol": "11:00",
                "Numero_Vol": "AF822",
                "IATA": "DLA",
            }
        ]
    )
    out_dest = ui_sim._recompute_dest_stats(
        df_plan,
        df_vols_src=df_vols,
        df_paramdest=pd.DataFrame(),
        start_dt=None,
        end_dt=None,
    )
    assert int(out_dest.iloc[0]["Nb_Vols_Existant"]) == 1

    df_plan_benev = pd.DataFrame(
        [
            {"Benevole": "", "Date_Vol": "16/02/26", "Heure_Vol": "11:00", "Numero_Vol": "AF822", "BE_Numero": "1"},
            {"Benevole": "ALICE", "Date_Vol": "16/02/26", "Heure_Vol": "11:00", "Numero_Vol": "AF822", "BE_Numero": "2"},
        ]
    )
    df_dispo = pd.DataFrame(
        [
            {"Benevole": "ALICE", "Date": "16/02/26", "Heure_Arrivee": "09:00", "Heure_Depart": "12:00"},
        ]
    )
    out_benev = ui_sim._recompute_bilan_benevoles(
        df_plan_benev,
        df_dispo,
        df_parambenev=pd.DataFrame([{"Benevole": "ALICE"}]),
        start_dt=None,
        end_dt=None,
    )
    assert "ALICE" in out_benev["Benevole"].astype(str).tolist()


def test_ui_week_data_and_planning_views_remaining_branches(monkeypatch):
    state = SimpleNamespace(
        df_vols=pd.DataFrame(
            [
                {"Date_Vol": "invalid", "Heure_Vol": "11h00", "Routing": "CDG-DLA", "Numero_Vol": "AF822"},
                {"Date_Vol": "16/02/26", "Heure_Vol": "11h00", "Routing": "CDG-DLA", "Numero_Vol": "AF822", "Source": "api"},
            ]
        ),
        df_param_dest=pd.DataFrame(),
        df_be=pd.DataFrame(),
        api_start_date=None,
        api_end_date=None,
    )
    monkeypatch.setattr(ui_week_data, "load_be_moteur", lambda: (None, None))
    out = ui_week_data._prepare_flights_dataframe(state, week=None, iata_to_city={"DLA": "DOUALA"})
    assert len(out) == 1

    # Force normalize_planning_df sans colonne Routing pour exécuter le fallback.
    monkeypatch.setattr(
        planning_views,
        "normalize_planning_df",
        lambda _df: pd.DataFrame(
            [
                {
                    "Date_Vol": "16/02/26",
                    "Heure_Vol": "11:00",
                    "Numero_Vol": "822",
                    "Destination": "DLA",
                    "BE_Numero": "260001",
                    "BE_Nb_Colis": 1,
                    "BE_Nb_Equiv": 1,
                    "Benevole": "ALICE",
                    "ID": "1",
                    "BE_Type": "",
                    "BE_Expediteur": "",
                    "BE_Destinataire": "",
                    "Telephone": "",
                    "_MANUEL": False,
                    "_STATUS": "normal",
                    "UID": "u1",
                }
            ]
        ),
    )
    out_view = planning_views.build_export_view(pd.DataFrame([{"x": 1}]))
    assert "Routing" in out_view.columns


def test_ui_week_data_heure_min_coercion_branch(monkeypatch):
    original_dataframe = ui_week_data.pd.DataFrame

    def _patched_dataframe(*args, **kwargs):
        df = original_dataframe(*args, **kwargs)
        if "HEURE_MIN" not in df.columns and "Label" in df.columns:
            df["HEURE_MIN"] = "not-a-number"
        return df

    monkeypatch.setattr(ui_week_data.pd, "DataFrame", _patched_dataframe)
    monkeypatch.setattr(ui_week_data, "load_be_moteur", lambda: (None, None))
    state = SimpleNamespace(
        df_vols=pd.DataFrame(
            [
                {
                    "Date_Vol": "16/02/26",
                    "Heure_Vol": "11h00",
                    "Routing": "CDG-DLA",
                    "Numero_Vol": "AF822",
                    "Source": "api",
                }
            ]
        ),
        df_param_dest=pd.DataFrame(),
        df_be=pd.DataFrame(),
        api_start_date=None,
        api_end_date=None,
    )
    out = ui_week_data._prepare_flights_dataframe(state, week=None, iata_to_city={"DLA": "DOUALA"})
    assert "HEURE_MIN" in out.columns
    assert pd.isna(out.loc[0, "HEURE_MIN"])


def test_load_shipments_additional_missing_columns_and_auto_parambe(monkeypatch):
    monkeypatch.setattr(load_shipments_mod, "_list_mag_central_sheets", lambda *_a, **_k: [])

    def _load_and_normalize(**kwargs):
        if kwargs.get("sheet_name") == load_shipments_mod.SHEET_MAG_CENTRAL:
            return pd.DataFrame([{"BE_Numero": "260001", "BE_Nb_Colis": 1, "BE_Type": "MM"}])
        return pd.DataFrame()

    monkeypatch.setattr(load_shipments_mod, "load_and_normalize", _load_and_normalize)
    monkeypatch.setattr(
        load_shipments_mod,
        "get_param_be",
        lambda: pd.DataFrame([{"Type": "MM", "Priorite_Type": 1, "Equiv": 1}]),
    )
    monkeypatch.setattr(load_shipments_mod.be_manager, "normalize_param_be", lambda _df: {"MM": {}, "AUTRE": {}})
    monkeypatch.setattr(load_shipments_mod, "compute_be_priority", lambda *_a, **_k: 1)
    monkeypatch.setattr(load_shipments_mod, "compute_equiv_colis", lambda *_a, **_k: 1)

    out = load_shipments_mod.load_shipments_df(planifiables_only=False, tdb_path=Path("dummy.xlsx"), param_be_raw=None)
    assert not out.empty
    assert "BE_Statut" in out.columns
    assert "Destination" in out.columns


def test_whatsapp_prenom_court_fallback_branch():
    df_bene = pd.DataFrame(columns=["DATE", "BENEVOLE", "Benevole_Prenom"])
    msg = wa._build_message_for_benevole(df_bene, vols_info={}, map_iata_city={})
    assert "Bonjour" in msg


def test_shipments_update_and_load_vols_api_edge_branches(monkeypatch, tmp_path):
    class _WB:
        def __init__(self):
            self.sheetnames = [sus.cp.SHEET_MAG_CENTRAL]
            self.active = SimpleNamespace(title="Main")

    assert sus._mag_sheet_names(_WB()) == [sus.cp.SHEET_MAG_CENTRAL]

    monkeypatch.setattr(
        sus.pd,
        "isna",
        lambda _v: (_ for _ in ()).throw(TypeError("boom")),
    )
    df_export = pd.DataFrame([{"BE_Key": "260001", "_STATUS": "normal"}])
    out = sus._apply_update_to_export_df(
        df_export,
        action="Mise à jour",
        be_num="260001",
        dest_iata="DLA",
        date_new="16/02/26",
        vol_new="822",
        heure_new="11h00",
        bene_choice="ALICE",
        be_info={"BE_Nb_Colis": object()},
        plan_row_full={},
        bene_meta={"ID": "1", "Telephone": "0600000000"},
        bene_changed=True,
    )
    assert not out.empty

    import loaders.load_vols_api as lva
    from openpyxl import Workbook

    target = tmp_path / "Vols.xlsx"
    Workbook().save(target)

    wb = Workbook()
    ws = wb.active
    ws.title = "API-S08-2026"
    ws._tables = []  # force branche non-dict
    monkeypatch.setattr(wb, "save", lambda *_a, **_k: None)
    monkeypatch.setattr(ws, "add_table", lambda *_a, **_k: (_ for _ in ()).throw(ValueError("boom")))
    monkeypatch.setattr(
        sys.modules.setdefault("utils.excel_automation", importlib.import_module("utils.excel_automation")),
        "write_sheet_table",
        lambda *_a, **_k: False,
    )
    monkeypatch.setattr(lva, "load_workbook", lambda *_a, **_k: wb)
    monkeypatch.setattr(lva.cp, "sync_local_file_to_onedrive", lambda *_a, **_k: None)

    df = pd.DataFrame([{"Date_Vol": "16/02/26", "Heure_Vol": "11h00", "Numero_Vol": "AF822", "Destination": "DOUALA"}])
    sheet_name = store_vols_api_sheet(df, pd.Timestamp("2026-02-16").date(), path=target)
    assert sheet_name.startswith("API-S")


def test_export_service_additional_remaining_branches(monkeypatch, tmp_path):
    class _ScalarGetDf(pd.DataFrame):
        @property
        def _constructor(self):
            return _ScalarGetDf

        def get(self, key, default=None):
            if key == "Heure_Vol":
                return "11:00"
            return super().get(key, default)

    monkeypatch.setattr(
        es,
        "build_export_view",
        lambda *_a, **_k: _ScalarGetDf(
            [
                {
                    "Date_Vol": "2026-02-16",
                    "Numero_Vol": "AF822",
                    "Destination": "DLA",
                    "Benevole": "ALICE DOE",
                    "BE_Numero": "260001",
                    "BE_Nb_Colis": 1,
                    "BE_Type": "MM",
                    "BE_Expediteur": "ASF",
                    "BE_Destinataire": "HOP",
                    "Routing": "",
                }
            ]
        ),
    )
    out = es._prepare_export_dataframe(
        pd.DataFrame([{"x": 1}]),
        df_paramdest=None,
        df_vols=None,
        df_parambenev=None,
    )
    assert not out.empty
    assert out.loc[0, "HEURE_AFF"] == "11h00"

    out2 = es._apply_routing_fallback_from_vols(
        pd.DataFrame(
            [
                {
                    "DATE": pd.Timestamp("2026-02-16"),
                    "Numero_Vol": "AF822",
                    "Routing": "",
                }
            ]
        ),
        df_vols=pd.DataFrame(
            [{"Date_Vol": "16/02/26", "Numero_Vol": "AF999", "Routing": "CDG-RUN"}]
        ),
    )
    assert out2.loc[0, "Routing"] == ""

    assert es._mag_sheet_names(SimpleNamespace(sheetnames=["MAG CENTRAL"], active=SimpleNamespace(title="Main"))) == [
        "MAG CENTRAL"
    ]

    class _Cell:
        def __init__(self, value, row):
            self.value = value
            self.row = row

    class _Ws:
        max_row = 2

        def iter_rows(self, **_kwargs):
            return [[_Cell(None, 1)], [_Cell("260001", 2)]]

    mag_idx = es._build_mag_index(_Ws())
    assert "260001" in mag_idx
    assert es._mag_lookup_keys("") == []
    assert es._alt_key_for_sheet("123", "MAG CENTRAL 2026") is None

    planning_dir = tmp_path / "planning"
    planning_dir.mkdir(parents=True, exist_ok=True)
    current = planning_dir / "ASFmm - PLANNING SEMAINE 2026-08-01.xlsx"
    current.touch()

    class _PathErr:
        def __init__(self, path: Path):
            self._path = path
            self.name = path.name

        def stat(self):
            raise OSError("boom")

        def __fspath__(self):
            return str(self._path)

    es._archive_latest_planning_if_needed(
        increment_version=False,
        max_version=1,
        existing_files=[_PathErr(current)],
        planning_dir_final=planning_dir,
        week_final=8,
        year_final=2026,
    )


def test_export_service_update_mag_error_paths(monkeypatch, tmp_path):
    path = tmp_path / "tdb.xlsx"
    from openpyxl import Workbook

    wb = Workbook()
    wb.active.title = "MAG CENTRAL 2026"
    wb.save(path)

    monkeypatch.setattr(es, "load_workbook", lambda *_a, **_k: (_ for _ in ()).throw(OSError("boom")))
    used, method = es._update_mag_central_dates_for_export(
        df_export=pd.DataFrame([{"BE_KEY": "260001"}]),
        week=8,
        year=2026,
        tdb_source_path=path,
    )
    assert used == {}
    assert method == "read_error"

    monkeypatch.setattr(es, "load_workbook", lambda *_a, **_k: wb)
    monkeypatch.setattr(es, "_find_mag_row", lambda **_k: ("MISSING", 2))
    used2, method2 = es._update_mag_central_dates_for_export(
        df_export=pd.DataFrame([{"BE_KEY": "260001"}]),
        week=8,
        year=2026,
        tdb_source_path=path,
    )
    assert used2 == {}
    assert method2 == "no_updates"

    monkeypatch.setattr(es, "_find_mag_row", lambda **_k: ("MAG CENTRAL 2026", 1))
    monkeypatch.setitem(sys.modules, "utils.excel_automation", None)
    used3, method3 = es._update_mag_central_dates_for_export(
        df_export=pd.DataFrame([{"BE_KEY": "260001", "DATE": pd.Timestamp("2026-02-16")}]),
        week=8,
        year=2026,
        tdb_source_path=path,
    )
    assert method3 in {"openpyxl", "no_updates"}


def test_clean_planning_df_name_fallback_and_id_benevole_branch(monkeypatch):
    base_df = pd.DataFrame(
        [
            {
                "DATE": "16/02/26",
                "BENEVOLE": "ALICE DOE",
                "BENEVOLE_ID": "999",
                "ID_BENEVOLE": "999.0",
                "DESTINATION": "DOUALA",
                "NUMERO VOL": "AF 822",
                "HEURE VOL": "11h00",
                "NUMERO BE": "260001",
                "NOMBRE COLIS": 1,
                "TYPE": "MM",
                "EXPEDITEUR": "ASF",
            }
        ]
    )
    monkeypatch.setattr(clean_df, "build_comm_base", lambda _df: base_df.copy())

    out = clean_df.build_df_comm(
        pd.DataFrame([{"dummy": 1}]),
        pd.DataFrame([{"Dest_IATA": "DLA", "Dest_Ville": "DOUALA"}]),
        pd.DataFrame(
            [
                {
                    "ID": "1",
                    "Benevole": "ALICE DOE",
                    "Telephone": "0600000000",
                    "Prenom": "Alice",
                    "Prenom_Court": "A",
                    "Nom": "Doe",
                }
            ]
        ),
    )
    assert not out.empty
    assert "BENEVOLE_ID" in out.columns


def test_excel_automation_remaining_windows_and_macos_branches(monkeypatch, tmp_path):
    def _dispatch_fail(_name):
        raise RuntimeError("boom")

    win32_mod = types.ModuleType("win32com")
    client_mod = types.ModuleType("win32com.client")
    client_mod.Dispatch = _dispatch_fail  # type: ignore[attr-defined]
    win32_mod.client = client_mod  # type: ignore[attr-defined]
    monkeypatch.setitem(sys.modules, "win32com", win32_mod)
    monkeypatch.setitem(sys.modules, "win32com.client", client_mod)

    assert excel_auto._update_excel_windows(tmp_path / "a.xlsx", "S", [(1, 1, "x")]) is False
    assert excel_auto._write_table_windows(tmp_path / "a.xlsx", "S", [["x"]]) is False
    assert excel_auto._write_table_macos(tmp_path / "a.xlsx", "S", []) is False

    monkeypatch.setattr(excel_auto.sys, "platform", "win32")
    monkeypatch.setattr(excel_auto, "_write_table_windows", lambda *_a, **_k: True)
    assert excel_auto.write_sheet_table(tmp_path / "a.xlsx", "S", [["x"]]) is True

    class _Ws:
        def __call__(self, _name):
            raise RuntimeError("missing")

        def Add(self):
            return SimpleNamespace(
                Name="S",
                Cells=lambda *_a, **_k: SimpleNamespace(),
                Range=lambda *_a, **_k: SimpleNamespace(Value=None),
            )

    class _Wb:
        def __init__(self):
            self.Worksheets = _Ws()

        def Save(self):
            return None

        def Close(self, SaveChanges=True):
            _ = SaveChanges
            return None

    class _Workbooks:
        def Open(self, _path):
            return _Wb()

    class _Excel:
        def __init__(self):
            self.DisplayAlerts = True
            self.Workbooks = _Workbooks()

        def Quit(self):
            return None

    client_mod.Dispatch = lambda _name: _Excel()  # type: ignore[attr-defined]
    monkeypatch.setattr(excel_auto.sys, "platform", "win32")
    assert excel_auto.replace_sheet_table(tmp_path / "a.xlsx", "Missing", [["x"]]) is True

# -*- coding: utf-8 -*-
from __future__ import annotations

from datetime import date

import pandas as pd
from openpyxl import Workbook

import loaders.load_vols_api as lva


def test_load_be_dest_codes_returns_empty_when_shipments_are_empty(monkeypatch):
    monkeypatch.setattr(lva, "load_shipments_df", lambda *_a, **_k: pd.DataFrame())

    assert lva.load_be_dest_codes() == []


def test_store_vols_api_sheet_falls_back_after_write_sheet_table_exception(monkeypatch, tmp_path):
    path = tmp_path / "Vols.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "API-S04-2026"
    ws["C10"] = "legacy"  # force extra empty cells after reset/write
    wb.save(path)

    import utils.excel_automation as excel_auto

    monkeypatch.setattr(excel_auto, "write_sheet_table", lambda *_a, **_k: (_ for _ in ()).throw(RuntimeError("boom")))
    monkeypatch.setattr(lva.cp, "sync_local_file_to_onedrive", lambda *_a, **_k: None)

    df = pd.DataFrame([{"Date_Vol": "23/01/26"}])
    sheet_name = lva.store_vols_api_sheet(df, date(2026, 1, 23), path=path)

    assert sheet_name == "API-S04-2026"


def test_copy_api_sheet_to_tmp_swallows_copy_errors_when_source_exists(monkeypatch, tmp_path):
    src = tmp_path / "src.xlsx"
    pd.DataFrame([{"A": 1}]).to_excel(src, sheet_name="API-S04-2026", index=False)

    monkeypatch.setattr(lva, "load_workbook", lambda *_a, **_k: (_ for _ in ()).throw(OSError("boom")))

    lva.copy_api_sheet_to_tmp(
        sheet_name="API-S04-2026",
        src_path=src,
        dst_path=tmp_path / "dst.xlsx",
    )

# -*- coding: utf-8 -*-
from __future__ import annotations

from datetime import date

from openpyxl import Workbook, load_workbook

import asf_app.services.shipments_update_service as sus
import scheduler.config_paths as cp


def test_mag_lookup_keys_includes_base_and_suffixes():
    keys = sus._mag_lookup_keys("000123")
    assert keys[0] == "000123"
    assert "123" in keys
    assert "0123" in keys


def test_alt_key_for_sheet_uses_year_suffix():
    assert sus._alt_key_for_sheet("000123", "MAG CENTRAL 2025") == "250123"
    assert sus._alt_key_for_sheet("250123", "MAG CENTRAL 2025") is None
    assert sus._alt_key_for_sheet("000123", "MAG CENTRAL") is None


def test_sheet_order_prefers_year_matching_be_prefix():
    names = ["MAG CENTRAL 2024", "MAG CENTRAL 2025", "MAG CENTRAL 2026"]
    ordered = sus._sheet_order_for_be("250001", names)
    assert ordered[0] == "MAG CENTRAL 2025"
    assert set(ordered) == set(names)


def test_build_mag_index_uses_normalized_and_raw_keys():
    wb = Workbook()
    ws = wb.active
    ws.title = "MAG CENTRAL 2025"
    ws["A1"] = "N° BE"
    ws["A2"] = "000123"
    ws["A3"] = "250001"

    idx = sus._build_mag_index(ws)
    assert idx["000123"] == 2
    assert idx["123"] == 2
    assert idx["250001"] == 3


def test_find_mag_target_row_uses_alt_key_when_needed():
    sheet = "MAG CENTRAL 2025"
    target_sheet, target_row = sus._find_mag_target_row(
        be_key="000123",
        mag_sheet_names=[sheet],
        mag_indexes={sheet: {"250123": 7}},
    )
    assert target_sheet == sheet
    assert target_row == 7


def test_parse_mag_departure_date_and_previous_friday():
    dt = sus._parse_mag_departure_date("06/01/2025")
    assert dt == date(2025, 1, 6)
    assert sus._previous_iso_week_friday(dt) == date(2025, 1, 3)
    assert sus._parse_mag_departure_date("not-a-date") is None


def test_mag_sheet_names_fallback_to_default_or_active():
    wb = Workbook()
    ws = wb.active
    ws.title = "Autre"
    assert sus._mag_sheet_names(wb) == [ws.title]

    wb.create_sheet(cp.SHEET_MAG_CENTRAL)
    assert sus._mag_sheet_names(wb) == [cp.SHEET_MAG_CENTRAL]


def test_mag_lookup_keys_handles_empty_input():
    assert sus._mag_lookup_keys("") == []


def test_find_mag_target_row_skips_empty_indexes():
    target_sheet, target_row = sus._find_mag_target_row(
        be_key="250001",
        mag_sheet_names=["MAG CENTRAL 2025"],
        mag_indexes={"MAG CENTRAL 2025": {}},
    )
    assert target_sheet is None
    assert target_row is None


def test_parse_mag_departure_date_handles_empty_and_previous_friday_none():
    assert sus._parse_mag_departure_date("") is None
    assert sus._previous_iso_week_friday(None) is None

    class _BadIso:
        def isocalendar(self):
            raise ValueError("boom")

    assert sus._previous_iso_week_friday(_BadIso()) is None


def test_update_mag_central_for_be_updates_cells_with_alt_key(tmp_path, monkeypatch):
    path = tmp_path / "tdb.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "MAG CENTRAL 2025"
    ws["A1"] = "N° BE"
    ws["A2"] = "250123"
    wb.save(path)

    monkeypatch.setattr(sus.cp, "sync_local_file_to_onedrive", lambda *_args, **_kwargs: None)

    status = sus._update_mag_central_for_be(
        be_num="000123",
        action="Replanification",
        date_new="06/01/2025",
        heure_new="11:30",
        vol_new="AF123",
        bene_choice="DUPONT",
        bene_meta={"ID": "45.0"},
        tdb_source_path=path,
    )

    assert status == "updated"
    wb2 = load_workbook(path)
    ws2 = wb2["MAG CENTRAL 2025"]

    depart_vol = ws2.cell(row=2, column=cp.MAG_CENTRAL_COL_DEPART_VOL).value
    depart_mag = ws2.cell(row=2, column=cp.MAG_CENTRAL_COL_DEPART_MAG).value
    if hasattr(depart_vol, "date"):
        depart_vol = depart_vol.date()
    if hasattr(depart_mag, "date"):
        depart_mag = depart_mag.date()

    assert depart_vol == date(2025, 1, 6)
    assert depart_mag == date(2025, 1, 3)
    assert ws2.cell(row=2, column=cp.MAG_CENTRAL_COL_ID_BENEV).value == "45"
    assert ws2.cell(row=2, column=cp.MAG_CENTRAL_COL_BENEV).value == "DUPONT"
    assert ws2.cell(row=2, column=cp.MAG_CENTRAL_COL_VOL).value == "AF123"
    assert ws2.cell(row=2, column=cp.MAG_CENTRAL_COL_HEURE).value == "11h30"


def test_update_mag_central_for_be_annulation_clears_assignment(tmp_path, monkeypatch):
    path = tmp_path / "tdb.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "MAG CENTRAL 2025"
    ws["A1"] = "N° BE"
    ws["A2"] = "250123"
    ws.cell(row=2, column=cp.MAG_CENTRAL_COL_DEPART_VOL).value = date(2025, 1, 6)
    ws.cell(row=2, column=cp.MAG_CENTRAL_COL_ID_BENEV).value = "45"
    ws.cell(row=2, column=cp.MAG_CENTRAL_COL_BENEV).value = "DUPONT"
    ws.cell(row=2, column=cp.MAG_CENTRAL_COL_VOL).value = "AF123"
    ws.cell(row=2, column=cp.MAG_CENTRAL_COL_HEURE).value = "11h30"
    wb.save(path)

    monkeypatch.setattr(sus.cp, "sync_local_file_to_onedrive", lambda *_args, **_kwargs: None)

    status = sus._update_mag_central_for_be(
        be_num="250123",
        action="Annulation",
        date_new="",
        heure_new="",
        vol_new="",
        bene_choice="",
        bene_meta={},
        tdb_source_path=path,
    )

    assert status == "updated"
    wb2 = load_workbook(path)
    ws2 = wb2["MAG CENTRAL 2025"]
    assert ws2.cell(row=2, column=cp.MAG_CENTRAL_COL_DEPART_VOL).value is None
    assert ws2.cell(row=2, column=cp.MAG_CENTRAL_COL_ID_BENEV).value is None
    assert ws2.cell(row=2, column=cp.MAG_CENTRAL_COL_BENEV).value is None
    assert ws2.cell(row=2, column=cp.MAG_CENTRAL_COL_VOL).value is None
    assert ws2.cell(row=2, column=cp.MAG_CENTRAL_COL_HEURE).value is None


def test_update_mag_central_for_be_returns_not_found_and_missing_file(tmp_path):
    assert sus._update_mag_central_for_be(
        be_num="250123",
        action="Replanification",
        date_new="06/01/2025",
        heure_new="11:30",
        vol_new="AF123",
        bene_choice="DUPONT",
        bene_meta={},
        tdb_source_path=tmp_path / "missing.xlsx",
    ) == "missing"

    path = tmp_path / "tdb.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "MAG CENTRAL 2025"
    ws["A1"] = "N° BE"
    ws["A2"] = "250999"
    wb.save(path)
    assert sus._update_mag_central_for_be(
        be_num="250123",
        action="Replanification",
        date_new="06/01/2025",
        heure_new="11:30",
        vol_new="AF123",
        bene_choice="DUPONT",
        bene_meta={},
        tdb_source_path=path,
    ) == "not_found"


def test_update_mag_central_for_be_returns_read_or_write_error(tmp_path, monkeypatch):
    import openpyxl

    path = tmp_path / "tdb.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "MAG CENTRAL 2025"
    ws["A1"] = "N° BE"
    ws["A2"] = "250123"
    wb.save(path)

    monkeypatch.setattr(openpyxl, "load_workbook", lambda *_a, **_k: (_ for _ in ()).throw(ValueError("bad")))
    assert sus._update_mag_central_for_be(
        be_num="250123",
        action="Replanification",
        date_new="06/01/2025",
        heure_new="11:30",
        vol_new="AF123",
        bene_choice="DUPONT",
        bene_meta={},
        tdb_source_path=path,
    ) == "read_error"

    wb2 = Workbook()
    ws2 = wb2.active
    ws2.title = "MAG CENTRAL 2025"
    ws2["A1"] = "N° BE"
    ws2["A2"] = "250123"
    monkeypatch.setattr(openpyxl, "load_workbook", lambda *_a, **_k: wb2)
    monkeypatch.setattr(sus.cp, "sync_local_file_to_onedrive", lambda *_a, **_k: None)
    monkeypatch.setattr(wb2, "save", lambda *_a, **_k: (_ for _ in ()).throw(OSError("boom")))
    assert sus._update_mag_central_for_be(
        be_num="250123",
        action="Replanification",
        date_new="06/01/2025",
        heure_new="11:30",
        vol_new="AF123",
        bene_choice="DUPONT",
        bene_meta={},
        tdb_source_path=path,
    ) == "write_error"


def test_update_mag_central_for_be_handles_missing_openpyxl_import(tmp_path, monkeypatch):
    import builtins

    path = tmp_path / "tdb.xlsx"
    path.write_text("x", encoding="utf-8")
    original_import = builtins.__import__

    def _fake_import(name, *args, **kwargs):
        if name == "openpyxl":
            raise ImportError("missing")
        return original_import(name, *args, **kwargs)

    monkeypatch.setattr(builtins, "__import__", _fake_import)
    assert sus._update_mag_central_for_be(
        be_num="250123",
        action="Replanification",
        date_new="06/01/2025",
        heure_new="11:30",
        vol_new="AF123",
        bene_choice="DUPONT",
        bene_meta={},
        tdb_source_path=path,
    ) == "openpyxl_missing"

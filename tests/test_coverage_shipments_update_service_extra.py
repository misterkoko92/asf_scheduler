# -*- coding: utf-8 -*-
from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import Workbook

import asf_app.services.shipments_update_service as sus


class _FakeExcelFile:
    def __init__(self, sheet_names: list[str]):
        self.sheet_names = sheet_names


class _Sentinel:
    def __str__(self) -> str:  # pragma: no cover - simple helper
        return "SENTINEL"


def test_load_be_status_falls_back_when_excel_file_open_fails(monkeypatch, tmp_path):
    monkeypatch.setattr(sus.pd, "ExcelFile", lambda *_a, **_k: (_ for _ in ()).throw(ValueError("boom")))
    monkeypatch.setattr(sus, "load_and_normalize", lambda **_kwargs: pd.DataFrame())

    out = sus.load_be_status("D", tdb_path=tmp_path / "missing.xlsx")

    assert out.empty
    assert list(out.columns) == ["Week", "Year"]


def test_load_be_status_skips_empty_sheets_and_sets_default_statut(monkeypatch, tmp_path):
    monkeypatch.setattr(
        sus.pd,
        "ExcelFile",
        lambda *_a, **_k: _FakeExcelFile(["MAG CENTRAL 2025", "MAG CENTRAL 2026"]),
    )

    calls = {"count": 0}

    def _fake_loader(**_kwargs):
        calls["count"] += 1
        if calls["count"] == 1:
            return pd.DataFrame()
        return pd.DataFrame([
            {
                "BE_Numero": "250001",
                "BE_Date_Vol": "2025-01-06",
                # intentionally no BE_Statut
            }
        ])

    monkeypatch.setattr(sus, "load_and_normalize", _fake_loader)

    out = sus.load_be_status("D", tdb_path=tmp_path / "tdb.xlsx")

    assert out.empty
    assert "BE_Statut" in out.columns


def test_load_be_status_d_for_week_returns_early_when_empty(monkeypatch):
    monkeypatch.setattr(sus, "load_be_status", lambda *_a, **_k: pd.DataFrame())

    out = sus.load_be_status_d_for_week(week=4, year=2026)

    assert out.empty


def test_apply_update_to_export_df_handles_pd_isna_type_error(monkeypatch):
    base = pd.DataFrame(
        [
            {
                "BE_Numero": "250001",
                "BE_Key": "250001",
                "BE_Nb_Colis": "",
                "BE_Nb_Equiv": "",
                "BE_Type": "",
                "BE_Expediteur": "",
                "BE_Destinataire": "",
                "ID": "",
                "Telephone": "",
                "_STATUS": "normal",
            }
        ]
    )
    sentinel = _Sentinel()
    real_isna = sus.pd.isna

    def _fake_isna(value):
        if value is sentinel:
            raise TypeError("cannot test")
        return real_isna(value)

    monkeypatch.setattr(sus.pd, "isna", _fake_isna)

    out = sus._apply_update_to_export_df(
        base,
        action="Replanification",
        be_num="250001",
        dest_iata="DLA",
        date_new="06/01/2025",
        vol_new="AF123",
        heure_new="11:30",
        bene_choice="TEST",
        be_info={"BE_Nb_Colis": sentinel},
        plan_row_full={"NO_MATCH": "x"},
        bene_meta={},
        bene_changed=False,
    )

    new_row = out[out["_STATUS"] == "new"].iloc[0]
    assert str(new_row["BE_Nb_Colis"]) == "SENTINEL"


def test_alt_key_and_build_mag_index_edge_cases():
    assert sus._alt_key_for_sheet("abc", "MAG CENTRAL 2025") is None

    wb = Workbook()
    ws = wb.active
    ws.title = "MAG CENTRAL 2025"
    ws["A1"] = None
    ws["A2"] = "250001"

    idx = sus._build_mag_index(ws)
    assert idx["250001"] == 2


def test_parse_mag_departure_date_tolerates_parser_errors(monkeypatch):
    monkeypatch.setattr(sus, "parse_date_series", lambda *_a, **_k: (_ for _ in ()).throw(TypeError("bad")))

    assert sus._parse_mag_departure_date("06/01/2025") is None


class _Cell:
    def __init__(self, value=None):
        self.value = value


class _WorksheetFailDepartMag:
    def __init__(self):
        self._cells: dict[tuple[int, int], _Cell] = {}

    def cell(self, row: int, column: int):
        if column == sus.cp.MAG_CENTRAL_COL_DEPART_MAG:
            raise TypeError("bad depart mag cell")
        key = (row, column)
        if key not in self._cells:
            self._cells[key] = _Cell()
        return self._cells[key]


class _WorkbookFake:
    def __init__(self):
        self.sheetnames = ["MAG CENTRAL 2025"]
        self._ws = _WorksheetFailDepartMag()

    def __getitem__(self, name: str):
        if name == "MAG CENTRAL 2025":
            return self._ws
        raise KeyError(name)

    def save(self, _path: Path):
        return None


class _WorkbookMissingSheet:
    def __init__(self):
        self.sheetnames = ["MAG CENTRAL 2025"]

    def __getitem__(self, name: str):
        raise KeyError(name)

    def save(self, _path: Path):
        return None


def test_update_mag_central_handles_missing_sheet_entries(monkeypatch, tmp_path):
    path = tmp_path / "tdb.xlsx"
    path.write_text("x", encoding="utf-8")

    monkeypatch.setattr("openpyxl.load_workbook", lambda *_a, **_k: _WorkbookMissingSheet())
    monkeypatch.setattr(sus, "_mag_sheet_names", lambda _wb: ["MISSING_SHEET"])

    status = sus._update_mag_central_for_be(
        be_num="250001",
        action="Replanification",
        date_new="06/01/2025",
        heure_new="11:30",
        vol_new="AF123",
        bene_choice="TEST",
        bene_meta={},
        tdb_source_path=path,
    )

    assert status == "not_found"


def test_update_mag_central_tolerates_depart_mag_write_errors(monkeypatch, tmp_path):
    path = tmp_path / "tdb.xlsx"
    path.write_text("x", encoding="utf-8")

    fake_wb = _WorkbookFake()
    monkeypatch.setattr("openpyxl.load_workbook", lambda *_a, **_k: fake_wb)
    monkeypatch.setattr(sus, "_mag_sheet_names", lambda _wb: ["MAG CENTRAL 2025"])
    monkeypatch.setattr(sus, "_build_mag_index", lambda _ws: {"250001": 2})
    monkeypatch.setattr(sus.cp, "sync_local_file_to_onedrive", lambda *_a, **_k: None)

    status = sus._update_mag_central_for_be(
        be_num="250001",
        action="Replanification",
        date_new="06/01/2025",
        heure_new="11:30",
        vol_new="AF123",
        bene_choice="TEST",
        bene_meta={"ID": "45.0"},
        tdb_source_path=path,
    )

    assert status == "updated"

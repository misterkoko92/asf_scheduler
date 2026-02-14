# -*- coding: utf-8 -*-
from __future__ import annotations

from datetime import date
from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook

import asf_app.services.export_service as es
import scheduler.config_paths as cp


class _FakeMatch:
    def __init__(self, *groups: str):
        self._groups = groups

    def group(self, idx: int) -> str:
        return self._groups[idx - 1]


def _planning_rows() -> pd.DataFrame:
    return pd.DataFrame(
        [
            {
                "DATE": pd.Timestamp("2026-01-19"),
                "HEURE_MIN": 660,
                "VOL_AFF": "AF 822",
                "Destination": "DLA",
                "BENEVOLE_DISP": "A DUPONT",
                "VILLE": "DOUALA",
                "IATA": "DLA",
                "ROUTING": "CDG-DLA",
                "HEURE_AFF": "11h00",
                "BE_NUM": "260001",
                "BE_COLIS": 5,
                "BE_TYPE": "MM",
                "DEPART_MAG": pd.Timestamp("2026-01-16"),
                "BE_EXP": "ASF",
                "BE_DEST": "HOPITAL",
                "_STATUS": "normal",
            },
            {
                "DATE": pd.Timestamp("2026-01-19"),
                "HEURE_MIN": 660,
                "VOL_AFF": "AF 822",
                "Destination": "DLA",
                "BENEVOLE_DISP": "B MARTIN",
                "VILLE": "DOUALA",
                "IATA": "DLA",
                "ROUTING": "CDG-DLA",
                "HEURE_AFF": "11h00",
                "BE_NUM": "260002",
                "BE_COLIS": 3,
                "BE_TYPE": "MM",
                "DEPART_MAG": pd.Timestamp("2026-01-16"),
                "BE_EXP": "ASF",
                "BE_DEST": "HOPITAL",
                "_STATUS": "normal",
            },
            {
                "DATE": pd.Timestamp("2026-01-19"),
                "HEURE_MIN": 700,
                "VOL_AFF": "AF 948",
                "Destination": "DLA",
                "BENEVOLE_DISP": "C BERNARD",
                "VILLE": "DOUALA",
                "IATA": "DLA",
                "ROUTING": "CDG-DLA",
                "HEURE_AFF": "11h40",
                "BE_NUM": "260003",
                "BE_COLIS": 2,
                "BE_TYPE": "MM",
                "DEPART_MAG": pd.Timestamp("2026-01-16"),
                "BE_EXP": "ASF",
                "BE_DEST": "HOPITAL",
                "_STATUS": "normal",
            },
        ]
    )


def test_create_minimal_workbook_tolerates_invalid_week_year(tmp_path):
    out = tmp_path / "minimal.xlsx"

    es._create_minimal_workbook(out, week="bad", year="bad")

    wb = load_workbook(out)
    ws = wb.worksheets[0]
    assert ws["A1"].value is None


def test_extract_version_from_name_tolerates_invalid_groups(monkeypatch):
    real_search = es.re.search

    def _fake_search(pattern: str, string: str, flags: int = 0):
        if pattern.startswith(r"SEMAINE"):
            return _FakeMatch("2026", "bad", "2")
        if "v(" in pattern:
            return _FakeMatch("bad", "2026", "2")
        if pattern.startswith(r"N"):
            return _FakeMatch("bad", "2026")
        return real_search(pattern, string, flags)

    monkeypatch.setattr(es.re, "search", _fake_search)

    assert es._extract_version_from_planning_name("ASFmm - PLANNING SEMAINE 2026-05-01.xlsx", week=5, year=2026) is None


def test_add_excel_table_if_needed_returns_early_when_disabled():
    wb = Workbook()
    ws = wb.active
    ws.append(["A", "B"])
    ws.append([1, 2])

    es._add_excel_table_if_needed(ws, create_tables=False, table_name="Table_X")

    assert len(ws.tables) == 0


def test_reset_planning_grid_skips_merged_cells():
    wb = Workbook()
    ws = wb.active
    ws["D3"] = "x"
    ws.merge_cells("D4:E4")
    ws["D4"] = "merged"

    es._reset_planning_grid(ws)

    assert ws["D3"].value is None


def test_populate_planning_sheet_breaks_when_block_is_full():
    wb = Workbook()
    ws = wb.active

    es._populate_planning_sheet(
        ws,
        dfp=_planning_rows(),
        day_blocks={0: (4, 4)},
        keep_rows={4},
        safe_excel=lambda value: value,
    )

    assert ws.cell(row=4, column=11).value == "260001"


def test_move_cell_value_to_visible_middle_returns_for_edge_cases():
    wb = Workbook()
    ws = wb.active
    ws["A17"] = "LUNDI"

    ws.row_dimensions[4].hidden = True
    ws.row_dimensions[5].hidden = True
    es._move_cell_value_to_visible_middle(ws, src_row=17, start=4, end=5)
    assert ws["A17"].value == "LUNDI"

    ws["A10"] = "MARDI"
    es._move_cell_value_to_visible_middle(ws, src_row=10, start=10, end=10)
    assert ws["A10"].value == "MARDI"


def test_prepare_output_workbook_path_creates_minimal_when_template_missing(tmp_path):
    out = tmp_path / "explicit.xlsx"

    out_path, skip = es._prepare_output_workbook_path(
        week=4,
        year=2026,
        template=tmp_path / "missing-template.xlsx",
        has_template=False,
        output_path=out,
        output_dir=None,
        skip_versioning=False,
    )

    assert out_path == out
    assert out.exists()
    assert skip is True


def test_archive_latest_planning_returns_when_no_candidate(monkeypatch, tmp_path):
    planning_dir = tmp_path / "planning"
    planning_dir.mkdir(parents=True, exist_ok=True)
    existing = planning_dir / "ASFmm - PLANNING SEMAINE 2026-04-03.xlsx"
    existing.touch()

    monkeypatch.setattr(es, "_extract_version_from_planning_name", lambda *_a, **_k: None)

    es._archive_latest_planning_if_needed(
        increment_version=False,
        max_version=3,
        existing_files=[existing],
        planning_dir_final=planning_dir,
        week_final=4,
        year_final=2026,
    )

    assert existing.exists()


def test_update_mag_central_dates_returns_excel_when_automation_succeeds(monkeypatch, tmp_path):
    path = tmp_path / "tdb.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "MAG CENTRAL 2026"
    ws["A1"] = "N° BE"
    ws["A2"] = "260001"
    ws.cell(row=2, column=cp.MAG_CENTRAL_COL_DEPART_MAG).value = date(2025, 1, 1)
    wb.save(path)

    monkeypatch.setattr(cp, "sync_local_file_to_onedrive", lambda *_a, **_k: None)
    monkeypatch.setattr("utils.excel_automation.update_excel_cells", lambda *_a, **_k: True)

    df_export = pd.DataFrame(
        [
            {"BE_KEY": "", "DATE": pd.Timestamp("2026-01-19")},
            {
                "BE_KEY": "260001",
                "DATE": pd.Timestamp("2026-01-19"),
                "ID": "45.0",
                "BENEVOLE_DISP": "A DUPONT",
                "VOL_AFF": "AF822",
                "HEURE_AFF": "11h00",
            },
        ]
    )

    used_dates, method = es._update_mag_central_dates_for_export(
        df_export=df_export,
        week=4,
        year=2026,
        tdb_source_path=path,
    )

    assert method == "excel"
    used = used_dates["260001"]
    if hasattr(used, "date"):
        used = used.date()
    assert used == date(2025, 1, 1)


class _WorkbookMissingSheet:
    def __init__(self):
        self.sheetnames = ["MAG CENTRAL 2026"]

    def __getitem__(self, _name: str):
        raise KeyError("missing")


def test_update_mag_central_dates_returns_read_error_when_sheet_lookup_fails(monkeypatch, tmp_path):
    path = tmp_path / "tdb.xlsx"
    path.write_text("x", encoding="utf-8")

    monkeypatch.setattr(es, "load_workbook", lambda *_a, **_k: _WorkbookMissingSheet())

    used_dates, method = es._update_mag_central_dates_for_export(
        df_export=pd.DataFrame([{"BE_KEY": "260001", "DATE": pd.Timestamp("2026-01-19")}]),
        week=4,
        year=2026,
        tdb_source_path=path,
    )

    assert used_dates == {}
    assert method == "read_error"


def test_update_mag_central_dates_fallback_openpyxl_handles_missing_sheet_and_save_error(monkeypatch, tmp_path):
    path = tmp_path / "tdb.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "MAG CENTRAL 2026"
    ws["A1"] = "N° BE"
    ws["A2"] = "260001"
    wb.save(path)

    loaded = load_workbook(path)
    monkeypatch.setattr(es, "load_workbook", lambda *_a, **_k: loaded)
    monkeypatch.setattr(es, "_build_mag_index", lambda *_a, **_k: {"260001": 2})
    monkeypatch.setattr(es, "_build_update_items_by_sheet", lambda *_a, **_k: {"MISSING": [(2, 1, "x")]})
    monkeypatch.setattr("utils.excel_automation.update_excel_cells", lambda *_a, **_k: False)
    monkeypatch.setattr(loaded, "save", lambda *_a, **_k: (_ for _ in ()).throw(OSError("boom")))

    used_dates, method = es._update_mag_central_dates_for_export(
        df_export=pd.DataFrame([{"BE_KEY": "260001", "DATE": pd.Timestamp("2026-01-19")}]),
        week=4,
        year=2026,
        tdb_source_path=path,
    )

    assert method == "openpyxl"
    assert isinstance(used_dates, dict)


def test_write_export_workbook_data_writes_bene_sheet_when_dataframe_present(monkeypatch, tmp_path):
    wb = Workbook()
    ws_plan = wb.active
    ws_export = wb.create_sheet("Export planning")
    ws_vols = wb.create_sheet("Data Vols")
    ws_bene = wb.create_sheet("Data Benevoles")
    out_path = tmp_path / "out.xlsx"

    calls: list[tuple[str, bool]] = []

    def _spy_write(ws, df_to_write, *, create_tables, table_name=None):
        calls.append((ws.title, bool(df_to_write is not None and not getattr(df_to_write, "empty", True))))

    monkeypatch.setattr(es, "_reset_planning_grid", lambda *_a, **_k: None)
    monkeypatch.setattr(es, "_populate_planning_sheet", lambda *_a, **_k: None)
    monkeypatch.setattr(es, "_apply_planning_layout", lambda *_a, **_k: None)
    monkeypatch.setattr(es, "sanitize_dataframe_for_excel", lambda df: df)
    monkeypatch.setattr(es, "_resolve_benevoles_export_dataframe", lambda *_a, **_k: pd.DataFrame([{"ID": "1"}]))
    monkeypatch.setattr(es, "_write_dataframe_to_sheet", _spy_write)
    monkeypatch.setattr(es, "load_workbook", lambda *_a, **_k: Workbook())
    monkeypatch.setattr(es, "_week_year_from_ws_plan", lambda *_a, **_k: (4, 2026))

    week, year = es._write_export_workbook_data(
        wb=wb,
        ws_plan=ws_plan,
        ws_export=ws_export,
        ws_vols=ws_vols,
        ws_bene=ws_bene,
        out_path=out_path,
        week=4,
        year=2026,
        dfp=pd.DataFrame([{"DATE": pd.Timestamp("2026-01-19")}]),
        df_vols=pd.DataFrame(),
        df_dispos=pd.DataFrame(),
        df_parambenev=pd.DataFrame(),
        benev_path=None,
        create_tables=False,
    )

    assert (week, year) == (4, 2026)
    assert ("Data Benevoles", True) in calls

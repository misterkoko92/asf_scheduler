# utils/excel_automation.py
# -*- coding: utf-8 -*-

from __future__ import annotations

import datetime as dt
import subprocess
import sys
from pathlib import Path
from typing import Iterable

from openpyxl.utils import get_column_letter
from openpyxl.utils.datetime import to_excel

from utils.applescript_utils import applescript_escape


def _coerce_excel_value(value):
    if value is None or value == "":
        return None
    if hasattr(value, "to_pydatetime"):
        try:
            value = value.to_pydatetime()
        except Exception:
            pass
    if isinstance(value, (dt.datetime, dt.date)):
        try:
            return float(to_excel(value))
        except Exception:
            return value.strftime("%Y-%m-%d")
    return value


def _normalize_table(data: Iterable[Iterable[object]]) -> list[list[object]]:
    rows = []
    max_cols = 0
    for row in data:
        prepared = []
        for val in row:
            v = _coerce_excel_value(val)
            if v is None:
                v = ""
            prepared.append(v)
        rows.append(prepared)
        if len(prepared) > max_cols:
            max_cols = len(prepared)
    for row in rows:
        if len(row) < max_cols:
            row.extend([""] * (max_cols - len(row)))
    return rows


def _as_applescript_value(value: object) -> str:
    if value is None or value == "":
        return "\"\""
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, (int, float)):
        return str(value)
    text = applescript_escape(str(value))
    return f"\"{text}\""


def _update_excel_windows(path: Path, sheet_name: str, updates: list[tuple[int, int, object]]) -> bool:
    try:
        import win32com.client  # type: ignore

        excel = win32com.client.Dispatch("Excel.Application")
        excel.DisplayAlerts = False
        wb = excel.Workbooks.Open(str(path))
        ws = wb.Worksheets(sheet_name)
        for row, col, val in updates:
            ws.Cells(row, col).Value = val
        wb.Save()
        wb.Close(SaveChanges=True)
        excel.Quit()
        return True
    except Exception:
        return False


def _update_excel_macos(path: Path, sheet_name: str, updates: list[tuple[int, int, object]]) -> bool:
    lines = []
    for row, col, val in updates:
        addr = f"{get_column_letter(col)}{row}"
        if isinstance(val, str):
            safe = applescript_escape(val)
            lines.append(f'set value of range "{addr}" of ws to "{safe}"')
        else:
            lines.append(f'set value of range "{addr}" of ws to {val}')

    path_esc = applescript_escape(str(path))
    sheet_name_safe = applescript_escape(sheet_name)
    script = f"""
        set theFile to POSIX file "{path_esc}"
        tell application "Microsoft Excel"
            activate
            set display alerts to false
            set wb to open workbook workbook file name theFile
            set ws to worksheet "{sheet_name_safe}" of wb
            {chr(10).join(lines)}
            save wb
            close wb saving yes
            set display alerts to true
        end tell
    """
    res = subprocess.run(
        ["osascript", "-e", script],
        capture_output=True,
        text=True,
    )
    return res.returncode == 0


def update_excel_cells(path: Path, sheet_name: str, updates: Iterable[tuple[int, int, object]]) -> bool:
    """
    Update Excel cells via native Excel automation to preserve validations.
    Returns True if updates were applied, False if unsupported/failed.
    """
    prepared = []
    for row, col, value in updates:
        val = _coerce_excel_value(value)
        if val is None:
            continue
        prepared.append((int(row), int(col), val))

    if not prepared:
        return True

    if sys.platform.startswith("win"):
        return _update_excel_windows(path, sheet_name, prepared)
    if sys.platform == "darwin":
        return _update_excel_macos(path, sheet_name, prepared)
    return False


def _write_table_windows(
    path: Path,
    sheet_name: str,
    table: list[list[object]],
    *,
    clear_contents: bool = True,
) -> bool:
    try:
        import win32com.client  # type: ignore

        excel = win32com.client.Dispatch("Excel.Application")
        excel.DisplayAlerts = False
        wb = excel.Workbooks.Open(str(path))
        try:
            ws = wb.Worksheets(sheet_name)
        except Exception:
            ws = wb.Worksheets.Add()
            ws.Name = sheet_name
        if clear_contents:
            ws.UsedRange.ClearContents()
        rows = len(table)
        cols = len(table[0]) if rows else 0
        if rows and cols:
            rng = ws.Range(ws.Cells(1, 1), ws.Cells(rows, cols))
            rng.Value = table
        wb.Save()
        wb.Close(SaveChanges=True)
        excel.Quit()
        return True
    except Exception:
        return False


def _write_table_macos(
    path: Path,
    sheet_name: str,
    table: list[list[object]],
    *,
    clear_contents: bool = True,
) -> bool:
    if not table:
        return False
    rows = len(table)
    cols = len(table[0])
    last_col = get_column_letter(cols)
    lines = []
    if clear_contents:
        lines.append("clear contents of used range of ws")
    for idx, row in enumerate(table, start=1):
        row_values = ", ".join(_as_applescript_value(v) for v in row)
        addr = f"A{idx}:{last_col}{idx}"
        lines.append(f"set value of range \"{addr}\" of ws to {{{row_values}}}")

    path_esc = applescript_escape(str(path))
    sheet_name_safe = applescript_escape(sheet_name)
    script = f"""
        set theFile to POSIX file "{path_esc}"
        tell application "Microsoft Excel"
            activate
            set display alerts to false
            set wb to open workbook workbook file name theFile
            if exists worksheet "{sheet_name_safe}" of wb then
                set ws to worksheet "{sheet_name_safe}" of wb
            else
                set ws to make new worksheet at end of worksheets of wb with properties {{name:"{sheet_name_safe}"}}
            end if
            {chr(10).join(lines)}
            save wb
            close wb saving yes
            set display alerts to true
        end tell
    """
    res = subprocess.run(
        ["osascript", "-e", script],
        capture_output=True,
        text=True,
    )
    return res.returncode == 0


def write_sheet_table(path: Path, sheet_name: str, data: Iterable[Iterable[object]], *, clear_contents: bool = True) -> bool:
    """
    Write a full table to a sheet using Excel automation (creates it if missing).
    Preserves conditional formatting and data validation by clearing contents only.
    """
    table = _normalize_table(data)
    if not table:
        return False
    if sys.platform.startswith("win"):
        return _write_table_windows(path, sheet_name, table, clear_contents=clear_contents)
    if sys.platform == "darwin":
        return _write_table_macos(path, sheet_name, table, clear_contents=clear_contents)
    return False


def replace_sheet_table(path: Path, sheet_name: str, data: Iterable[Iterable[object]]) -> bool:
    """
    Replace a sheet entirely using Excel automation (delete/create then write table).
    Useful for sheets where formatting is not required but other sheets must be preserved.
    """
    table = _normalize_table(data)
    if sys.platform.startswith("win"):
        try:
            import win32com.client  # type: ignore

            excel = win32com.client.Dispatch("Excel.Application")
            excel.DisplayAlerts = False
            wb = excel.Workbooks.Open(str(path))
            try:
                wb.Worksheets(sheet_name).Delete()
            except Exception:
                pass
            ws = wb.Worksheets.Add()
            ws.Name = sheet_name
            if table:
                rows = len(table)
                cols = len(table[0])
                rng = ws.Range(ws.Cells(1, 1), ws.Cells(rows, cols))
                rng.Value = table
            wb.Save()
            wb.Close(SaveChanges=True)
            excel.Quit()
            return True
        except Exception:
            return False
    if sys.platform == "darwin":
        path_esc = applescript_escape(str(path))
        sheet_name_safe = applescript_escape(sheet_name)
        lines = []
        if table:
            cols = len(table[0])
            last_col = get_column_letter(cols)
            for idx, row in enumerate(table, start=1):
                row_values = ", ".join(_as_applescript_value(v) for v in row)
                addr = f"A{idx}:{last_col}{idx}"
                lines.append(f"set value of range \"{addr}\" of ws to {{{row_values}}}")
        script = f"""
            set theFile to POSIX file "{path_esc}"
            tell application "Microsoft Excel"
                activate
                set display alerts to false
                set wb to open workbook workbook file name theFile
                if exists worksheet "{sheet_name_safe}" of wb then
                    delete worksheet "{sheet_name_safe}" of wb
                end if
                set ws to make new worksheet at end of worksheets of wb with properties {{name:"{sheet_name_safe}"}}
                {chr(10).join(lines)}
                save wb
                close wb saving yes
                set display alerts to true
            end tell
        """
        res = subprocess.run(
            ["osascript", "-e", script],
            capture_output=True,
            text=True,
        )
        return res.returncode == 0
    return False

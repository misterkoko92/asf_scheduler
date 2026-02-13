# loaders/load_vols_api.py
# -*- coding: utf-8 -*-
"""
Loader alternatif des vols via l'API Air France (CDG -> destinations ParamDest).
N'affecte pas le loader Excel existant.
"""

from __future__ import annotations

import logging
from datetime import date
from pathlib import Path
from typing import Dict, List

import pandas as pd
from openpyxl import load_workbook

import scheduler.config_paths as cp
from asf_app.services.airfrance_api import fetch_multiple
from loaders.load_shipments import load_shipments_df
from loaders.universal_loader import load_and_normalize
from scheduler.column_map import column_map_param_dest
from scheduler.config_paths import SHEET_PARAM_DEST, TABLEAU_DE_BORD, VOLS, VOLS_SRC
from utils.datetime_utils import (
    hour_min_from_series,
    normalize_hour_str,
    parse_date_series,
    parse_time_series,
)

logger = logging.getLogger("ASF-SCHEDULER")


def _fmt_date(dt) -> str:
    try:
        return pd.to_datetime(dt, errors="coerce", dayfirst=True).strftime("%d/%m/%y")
    except (AttributeError, TypeError, ValueError):
        return ""


def load_paramdest_codes() -> pd.DataFrame:
    """Charge ParamDest depuis TABLEAU_DE_BORD."""
    return load_and_normalize(
        TABLEAU_DE_BORD,
        SHEET_PARAM_DEST,
        column_map_param_dest,
        header=0,
    )


def load_be_dest_codes() -> List[str]:
    """
    Retourne la liste des destinations IATA présentes dans les BE statut D.
    """
    try:
        df = load_shipments_df(planifiables_only=True)
        if df is None or df.empty:
            return []
        codes = pd.Series(dtype=str)
        if "Dest_IATA" in df.columns:
            codes = df["Dest_IATA"]
        elif "Destination" in df.columns:
            codes = df["Destination"]
        codes = codes.dropna().astype(str).str.upper()
        codes = codes[codes.str.len() == 3]
        return sorted(set(codes))
    except (FileNotFoundError, OSError, KeyError, RuntimeError, TypeError, ValueError):
        return []


def load_vols_api(
    start_date: date,
    end_date: date,
    *,
    time_origin_type: str | None = None,
) -> pd.DataFrame:
    """
    Retourne un DataFrame structuré comme load_vols_df() mais issu de l'API.
    time_origin_type: S/M/I/P (None -> valeur configurée dans le service API).
    """
    df_param = load_paramdest_codes()
    df_be_codes = load_be_dest_codes()
    if df_param is None or df_param.empty:
        return pd.DataFrame()

    if df_be_codes:
        codes = df_be_codes
    else:
        codes = sorted(set(df_param["Dest_IATA"].dropna().astype(str).str.upper()))

    flights = fetch_multiple(
        dest_codes=codes,
        start_date=start_date.strftime("%Y-%m-%d"),
        end_date=end_date.strftime("%Y-%m-%d"),
        time_origin_type=time_origin_type,
    )

    iata_to_city: Dict[str, str] = {
        str(r.get("Dest_IATA", "")).upper(): str(r.get("Dest_Ville", "")).upper()
        for _, r in df_param.iterrows()
    }
    iata_to_cap: Dict[str, int | None] = {
        str(r.get("Dest_IATA", "")).upper(): r.get("Max_Colis_Par_Vol")
        for _, r in df_param.iterrows()
    }

    rows: List[Dict[str, object]] = []
    for f in flights:
        route_parts = f.route.split("-") if f.route else []
        if not route_parts or len(route_parts) < 2:
            continue
        origin = route_parts[0]
        dests = route_parts[1:]
        for dest_iata in dests:
            dest_city = iata_to_city.get(dest_iata, dest_iata)
            rows.append(
                {
                    "Date_Vol": f.date_depart,
                    "Heure_Vol": f.heure_depart,
                    "Numero_Vol": f.numero_vol,
                    "Destination": dest_city,
                    "IATA": dest_iata,
                    "Routing": f.route,
                    "Max_Colis": iata_to_cap.get(dest_iata),
                    "Origine": origin,
                }
            )

    df = pd.DataFrame(rows)
    if not df.empty:
        df["Max_Colis"] = pd.to_numeric(df["Max_Colis"], errors="coerce").astype("Int64")
        df["Date_dt"] = parse_date_series(df["Date_Vol"])
        df["Date_Vol"] = df["Date_dt"].dt.strftime("%d/%m/%y")
        df["Heure_Vol_dt"] = parse_time_series(df["Heure_Vol"])
        df["Heure_Vol"] = normalize_hour_str(df["Heure_Vol"]).fillna("")
        df["HEURE_MIN"] = hour_min_from_series(df["Heure_Vol"])
        df = df.drop_duplicates(subset=["Date_Vol", "Numero_Vol", "Destination"]).reset_index(drop=True)

    # Debug logs
    try:
        logger.info("[AF API] Vols charges: %s", len(df))
        if not df.empty:
            logger.debug(
                "[AF API] Dates uniques: %s",
                sorted(df["Date_Vol"].dropna().unique().tolist()),
            )
            logger.debug(
                "[AF API] Destinations uniques: %s",
                sorted(df["IATA"].dropna().unique().tolist()),
            )
    except (AttributeError, KeyError, TypeError, ValueError):
        pass

    return df


def store_vols_api_sheet(df: pd.DataFrame, start_date: date, path: Path = VOLS_SRC) -> str:
    """
    Enregistre le résultat API dans Vols.xlsx sur une feuille dédiée.
    - Nom : API-SXX-YYYY (XX = semaine ISO de start_date, YYYY = année)
    - Si la feuille existe, elle est écrasée ; sinon elle est créée en dernière position.
    """
    if df is None:
        return ""

    week = start_date.isocalendar()[1]
    year = start_date.isocalendar()[0]
    sheet_name = f"API-S{week:02d}-{year}"

    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)

    # Nettoyage valeurs <NA> pour openpyxl
    df_to_save = df.copy()
    try:
        df_to_save = df_to_save.astype(object).where(pd.notna(df_to_save), "")
    except (TypeError, ValueError):
        pass
    table_rows = [list(df_to_save.columns)]
    table_rows.extend([list(row) for row in df_to_save.itertuples(index=False, name=None)])

    if not path.exists():
        with pd.ExcelWriter(path, engine="openpyxl") as writer:
            df_to_save.to_excel(writer, sheet_name=sheet_name, index=False)
        cp.sync_local_file_to_onedrive(path)
        return sheet_name

    try:
        from utils.excel_automation import write_sheet_table

        if write_sheet_table(path, sheet_name, table_rows):
            cp.sync_local_file_to_onedrive(path)
            return sheet_name
    except (ImportError, OSError, RuntimeError, TypeError, ValueError):
        pass

    wb = load_workbook(path)
    ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.create_sheet(sheet_name)
    max_row = max(ws.max_row, len(table_rows))
    max_col = max(ws.max_column, len(table_rows[0]) if table_rows else 0)
    if max_row and max_col:
        for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
            for cell in row:
                cell.value = None
    for r_idx, row in enumerate(table_rows, start=1):
        for c_idx, val in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=val)
    # Ajuste la largeur des colonnes (sans wrap)
    for col_cells in ws.columns:
        max_len = 0
        col_letter = col_cells[0].column_letter
        for cell in col_cells:
            if cell.value is None:
                continue
            max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_len + 2
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = cell.alignment.copy(wrap_text=False)
    # Applique un tableau pour activer les filtres
    try:
        try:
            tbls = ws._tables
            if isinstance(tbls, dict):
                tbls.clear()
            else:
                ws._tables = []
        except (AttributeError, TypeError):
            pass
        from openpyxl.worksheet.table import Table, TableStyleInfo

        last_row = ws.max_row
        last_col = ws.max_column
        end_col_letter = ws.cell(row=1, column=last_col).column_letter
        table_ref = f"A1:{end_col_letter}{last_row}"
        ws_table = Table(displayName="VolsAPI", ref=table_ref)
        style = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        ws_table.tableStyleInfo = style
        ws.add_table(ws_table)
    except (ImportError, AttributeError, OSError, TypeError, ValueError):
        pass

    wb.save(path)
    cp.sync_local_file_to_onedrive(path)
    return sheet_name


def copy_api_sheet_to_tmp(sheet_name: str, src_path: Path = VOLS_SRC, dst_path: Path = VOLS) -> None:
    """
    Copie l'onglet API-Sxx-YYYY du fichier source vers la copie TMP (VOLS).
    Écrase l'onglet s'il existe déjà dans la copie.
    """
    if not sheet_name:
        return
    try:
        src_path = Path(src_path)
        dst_path = Path(dst_path)
        if not src_path.exists():
            return
        # Charger source et destination
        wb_src = load_workbook(src_path)
        if sheet_name not in wb_src.sheetnames:
            return
        wb_dst = load_workbook(dst_path) if dst_path.exists() else None
        if wb_dst is None:
            from openpyxl import Workbook
            wb_dst = Workbook()
            # supprimer la feuille par défaut si présente
            if wb_dst.active and wb_dst.active.title == "Sheet":
                del wb_dst[wb_dst.active.title]

        ws_src = wb_src[sheet_name]
        ws_new = wb_dst[sheet_name] if sheet_name in wb_dst.sheetnames else wb_dst.create_sheet(sheet_name)

        max_row = max(ws_new.max_row, ws_src.max_row)
        max_col = max(ws_new.max_column, ws_src.max_column)
        if max_row and max_col:
            for row in ws_new.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
                for cell in row:
                    cell.value = None

        # Copier les valeurs (sans styles) pour rester simple
        for r_idx, row in enumerate(ws_src.values, start=1):
            for c_idx, val in enumerate(row, start=1):
                ws_new.cell(row=r_idx, column=c_idx, value=val)

        wb_dst.save(dst_path)
        cp.sync_local_file_to_onedrive(dst_path)
    except (FileNotFoundError, OSError, KeyError, RuntimeError, TypeError, ValueError):
        pass
